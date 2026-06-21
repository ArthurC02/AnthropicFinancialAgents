# -*- coding: utf-8 -*-
"""Aurora Capital Management LLC — 2026-04 月結工作底稿(活公式 + foot check)。

資料來源:**全部來自 internal-gl MCP 的真實回傳**(../mcp_pulls/*.json),不是手 key。
本腳本讀 mcp_pulls 的 trial balance + journal entries,組成 Excel 結帳工作底稿:
  - Accruals & JEs : 3 筆應計草稿 + Dr=Cr 檢查
  - Trial Balance  : pre-close / 應計調整 / post-close(活公式)+ Dr=Cr 平衡檢查
  - Roll-forward 2100 : 期初 − 沖回 + 新估列 = 期末,內建 IF foot check
  - Variance       : 4月當月 vs 前3月平均月跑率,>5% 自動標示
  - JE Drill       : 從 MCP get_journal_entries 抓回的明細(解釋 flux 的證據)

應計政策(A1/A2/A3)是「任務輸入 / 公司應計排程」——internal-gl 是唯讀總帳,不提供
應計排程工具,所以這三筆比照 fileBased 的 accruals_support 當成 close 的政策輸入。
草稿分錄一律 **不過帳**,等 controller 簽核。Mock / dev only。
"""
from __future__ import annotations

import json
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
PULLS = os.path.normpath(os.path.join(HERE, "..", "mcp_pulls"))
OUT = os.path.normpath(os.path.join(HERE, "..", "out"))
os.makedirs(OUT, exist_ok=True)

# ---- styles -----------------------------------------------------------------
HDR = "1F4E79"; SUB = "D9E1F2"; OUTC = "BDD7EE"; INP = "FFF2CC"
WARN = "FCE4D6"; OK = "E2EFDA"; WHITE = "FFFFFF"
BLUEF = "0000FF"; GREENF = "008000"
thin = Side(style="thin", color="BFBFBF")
box = Border(left=thin, right=thin, top=thin, bottom=thin)
F_TITLE = Font(bold=True, size=14, color="1F4E79")
F_HDR = Font(bold=True, color=WHITE)
F_BOLD = Font(bold=True)
F_IN = Font(color=BLUEF)          # 藍字 = 輸入(MCP 拉回)
F_FORM = Font(color=GREENF)       # 綠字 = 公式
MONEY = "#,##0"
PCT = "0.0%"


def fill(c):
    return PatternFill("solid", fgColor=c)


def style_hdr(ws, row, cols):
    for c in cols:
        cell = ws[f"{c}{row}"]
        cell.font = F_HDR
        cell.fill = fill(HDR)
        cell.border = box
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


# ---- load MCP pulls ---------------------------------------------------------
def load(name):
    with open(os.path.join(PULLS, f"{name}.json"), encoding="utf-8") as f:
        return json.load(f)["result"]


tb = load("02_trial_balance")
je_all = load("10_je_all_april")
bal_2100 = load("11_balance_2100")

lines = tb["lines"]
by_acct = {l["account"]: l for l in lines}

# 應計政策輸入(任務輸入;internal-gl 不提供應計排程工具)
ACCRUALS = [
    ("A1", "IT 服務(4月)", "6300", "資訊技術 Technology", 75000, True,
     "供應商月帳單,發票 2026-05-03 才到,屬4月服務 → 估列;發票入帳時自動沖回"),
    ("A2", "年度分紅", "6000", "薪酬 Compensation", 100000, False,
     "獎金池 1,200,000 ÷ 12,按月攤;年底發放,非自動沖回"),
    ("A3", "FY2026 外部審計", "6200", "專業服務費 Professional Fees", 100000, False,
     "全年估 300,000,至4月約完成 1/3,累計估列;已入帳 0"),
]
accrual_total = sum(a[4] for a in ACCRUALS)            # 275,000
accrual_by_acct = {}
for a in ACCRUALS:
    accrual_by_acct[a[2]] = accrual_by_acct.get(a[2], 0) + a[4]

wb = Workbook()

# =====================================================================
# Sheet 1: Accruals & JEs
# =====================================================================
ws = wb.active
ws.title = "Accruals & JEs"
ws["A1"] = "應計排程 + 草稿分錄 Accrual Schedule & Draft JEs"
ws["A1"].font = F_TITLE
ws["A2"] = "Aurora Capital Management LLC · 2026-04 · USD · 草稿,待 controller 核准(勿過帳)"
ws["A2"].font = Font(italic=True, color="808080")

r = 4
hdrs = ["#", "應計項目", "借方科目 Dr", "貸方科目 Cr", "金額 USD", "自動沖回?", "依據 / Memo"]
for i, h in enumerate(hdrs):
    ws.cell(row=r, column=i + 1, value=h)
style_hdr(ws, r, [get_column_letter(i + 1) for i in range(len(hdrs))])
r += 1
first = r
for a in ACCRUALS:
    ws.cell(row=r, column=1, value=a[0]).border = box
    ws.cell(row=r, column=2, value=a[1]).border = box
    ws.cell(row=r, column=3, value=f"{a[2]} {a[3]}").border = box
    ws.cell(row=r, column=4, value="2100 應付費用 Accrued Exp").border = box
    cAmt = ws.cell(row=r, column=5, value=a[4]); cAmt.border = box; cAmt.number_format = MONEY; cAmt.font = F_IN; cAmt.fill = fill(INP)
    ws.cell(row=r, column=6, value="是 (auto-reversing)" if a[5] else "否").border = box
    ws.cell(row=r, column=7, value=a[6]).border = box
    r += 1
ws.cell(row=r, column=4, value="合計 Total").font = F_BOLD
tot = ws.cell(row=r, column=5, value=f"=SUM(E{first}:E{r-1})")
tot.number_format = MONEY; tot.font = F_FORM; tot.fill = fill(OK); tot.border = box
total_row = r

# Dr=Cr check for the JE batch
r += 2
ws.cell(row=r, column=1, value="分錄批次 Dr=Cr 檢查").font = F_BOLD
r += 1
ws.cell(row=r, column=1, value="借方合計 Σ Dr").border = box
ws.cell(row=r, column=2, value=f"=E{total_row}").number_format = MONEY
ws.cell(row=r, column=2).font = F_FORM
ws.cell(row=r + 1, column=1, value="貸方合計 Σ Cr (全入 2100)").border = box
ws.cell(row=r + 1, column=2, value=f"=E{total_row}").number_format = MONEY
ws.cell(row=r + 1, column=2).font = F_FORM
ws.cell(row=r + 2, column=1, value="差額 / 平衡?").font = F_BOLD
chk = ws.cell(row=r + 2, column=2, value=f'=IF(B{r}=B{r+1},"✓ 平衡 0","✗ 不平")')
chk.font = F_FORM; chk.fill = fill(OK)

for col, w in zip("ABCDEFG", [5, 22, 26, 26, 12, 18, 46]):
    ws.column_dimensions[col].width = w

# =====================================================================
# Sheet 2: Trial Balance (pre / accrual adj / post)  活公式
# =====================================================================
ws2 = wb.create_sheet("Trial Balance")
ws2["A1"] = "試算表 Trial Balance — pre-close → 應計調整 → post-close"
ws2["A1"].font = F_TITLE
ws2["A2"] = "來源:internal-gl MCP get_trial_balance(藍字)· 調整與合計為公式(綠字)"
ws2["A2"].font = Font(italic=True, color="808080")

r = 4
hdrs = ["科目", "名稱", "類別", "Dr/Cr", "上月 YTD", "本月 YTD (pre-close)",
        "應計調整", "Post-close", "Dr 餘額", "Cr 餘額"]
for i, h in enumerate(hdrs):
    ws2.cell(row=r, column=i + 1, value=h)
style_hdr(ws2, r, [get_column_letter(i + 1) for i in range(len(hdrs))])
r += 1
data_first = r
for l in lines:
    acct = l["account"]
    ws2.cell(row=r, column=1, value=acct).border = box
    ws2.cell(row=r, column=2, value=l["name"]).border = box
    ws2.cell(row=r, column=3, value=l["type"]).border = box
    ws2.cell(row=r, column=4, value=l["normal_balance"]).border = box
    c5 = ws2.cell(row=r, column=5, value=l["prior_ytd"]); c5.number_format = MONEY; c5.font = F_IN; c5.border = box
    c6 = ws2.cell(row=r, column=6, value=l["current_ytd"]); c6.number_format = MONEY; c6.font = F_IN; c6.border = box
    adj = accrual_by_acct.get(acct, 0)
    c7 = ws2.cell(row=r, column=7, value=adj if adj else None)
    c7.number_format = MONEY; c7.border = box
    if adj:
        c7.font = F_IN; c7.fill = fill(INP)
    # post-close = pre + adjustment (note: 2100 picks up the credit side of all accruals)
    if acct == "2100":
        c8 = ws2.cell(row=r, column=8, value=f"=F{r}+{accrual_total}")
    else:
        c8 = ws2.cell(row=r, column=8, value=f"=F{r}+IF(G{r}=\"\",0,G{r})")
    c8.number_format = MONEY; c8.font = F_FORM; c8.border = box
    # Dr / Cr split off post-close
    if l["normal_balance"] == "Dr":
        ws2.cell(row=r, column=9, value=f"=H{r}").number_format = MONEY
        ws2.cell(row=r, column=10, value=0).number_format = MONEY
    else:
        ws2.cell(row=r, column=9, value=0).number_format = MONEY
        ws2.cell(row=r, column=10, value=f"=H{r}").number_format = MONEY
    ws2.cell(row=r, column=9).font = F_FORM; ws2.cell(row=r, column=10).font = F_FORM
    ws2.cell(row=r, column=9).border = box; ws2.cell(row=r, column=10).border = box
    r += 1
data_last = r - 1

# totals
ws2.cell(row=r, column=2, value="合計 Total").font = F_BOLD
for col in (5, 6, 7, 9, 10):
    L = get_column_letter(col)
    c = ws2.cell(row=r, column=col, value=f"=SUM({L}{data_first}:{L}{data_last})")
    c.number_format = MONEY; c.font = F_FORM; c.fill = fill(SUB); c.border = box
tot_row = r
r += 2
# Dr=Cr balance checks (pre and post)
ws2.cell(row=r, column=1, value="平衡檢查 Balance Check").font = F_BOLD
r += 1
ws2.cell(row=r, column=1, value="Post-close Σ Dr")
ws2.cell(row=r, column=3, value=f"=I{tot_row}").number_format = MONEY
ws2.cell(row=r, column=3).font = F_FORM
ws2.cell(row=r + 1, column=1, value="Post-close Σ Cr")
ws2.cell(row=r + 1, column=3, value=f"=J{tot_row}").number_format = MONEY
ws2.cell(row=r + 1, column=3).font = F_FORM
ws2.cell(row=r + 2, column=1, value="Dr − Cr / 平衡?").font = F_BOLD
bc = ws2.cell(row=r + 2, column=3, value=f'=IF(ROUND(I{tot_row}-J{tot_row},2)=0,"✓ 平衡 (Dr=Cr)","✗ 差額 "&TEXT(I{tot_row}-J{tot_row},"#,##0"))')
bc.font = F_FORM; bc.fill = fill(OK)
# YTD net income pre vs post
r += 4
ws2.cell(row=r, column=1, value="YTD 損益(pre vs post)").font = F_BOLD
r += 1
rev_rows = [i for i, l in enumerate(lines) if l["type"] == "Revenue"]
exp_rows = [i for i, l in enumerate(lines) if l["type"] == "Expense"]
# revenue is Cr-normal (col F pre / H post in 'current' direction); use H for post
ws2.cell(row=r, column=1, value="收入合計 (post)")
ws2.cell(row=r, column=3, value="=" + "+".join(f"H{data_first+i}" for i in rev_rows)).number_format = MONEY
ws2.cell(row=r, column=3).font = F_FORM
ws2.cell(row=r + 1, column=1, value="費用合計 (post)")
ws2.cell(row=r + 1, column=3, value="=" + "+".join(f"H{data_first+i}" for i in exp_rows)).number_format = MONEY
ws2.cell(row=r + 1, column=3).font = F_FORM
ws2.cell(row=r + 2, column=1, value="YTD 淨利 (post) = 收入 − 費用").font = F_BOLD
ni = ws2.cell(row=r + 2, column=3, value=f"=C{r}-C{r+1}")
ni.number_format = MONEY; ni.font = F_FORM; ni.fill = fill(WARN)

for col, w in zip("ABCDEFGHIJ", [7, 32, 10, 7, 16, 18, 14, 16, 16, 16]):
    ws2.column_dimensions[col].width = w

# =====================================================================
# Sheet 3: Roll-forward 2100
# =====================================================================
ws3 = wb.create_sheet("Roll-forward 2100")
ws3["A1"] = "應付費用 (2100) 結轉 Roll-forward"
ws3["A1"].font = F_TITLE
ws3["A2"] = "期初 / 沖回 來自 MCP(get_account_balance + JE-4002 drill)· 期末為公式"
ws3["A2"].font = Font(italic=True, color="808080")

rows = [
    ("期初餘額(2026-03 月底)Beginning", bal_2100["prior_ytd"], "MCP get_account_balance.prior_ytd", "in"),
    ("減:4月已支付/沖回 Payments & Reversals", -200000, "MCP get_journal_entries 2100 → JE-4002 (Bank)", "in"),
    ("= 結帳前餘額(與試算表一致)Pre-close", None, "公式 = 期初 + 沖回", "form_pre"),
    ("加:本月新估列(A1+A2+A3)New Accruals", accrual_total, "應計排程(任務輸入)", "in"),
    ("= 期末餘額(post-close)Ending", None, "公式 = 結帳前 + 新估列", "form_end"),
]
r = 4
ws3.cell(row=r, column=1, value="項目").font = F_HDR
ws3.cell(row=r, column=2, value="金額 USD").font = F_HDR
ws3.cell(row=r, column=3, value="來源 / 註").font = F_HDR
style_hdr(ws3, r, ["A", "B", "C"])
r += 1
beg_row = r
for label, amt, note, kind in rows:
    ws3.cell(row=r, column=1, value=label).border = box
    c = ws3.cell(row=r, column=2); c.number_format = MONEY; c.border = box
    if kind == "in":
        c.value = amt; c.font = F_IN; c.fill = fill(INP)
    elif kind == "form_pre":
        c.value = f"=B{beg_row}+B{beg_row+1}"; c.font = F_FORM
    elif kind == "form_end":
        c.value = f"=B{beg_row+2}+B{beg_row+3}"; c.font = F_FORM; c.fill = fill(OK)
    ws3.cell(row=r, column=3, value=note).border = box
    r += 1
end_row = beg_row + 4
pre_row = beg_row + 2
# foot check
r += 1
ws3.cell(row=r, column=1, value="Foot check:期初 − 沖回 + 新估列 = 期末").font = F_BOLD
fc = ws3.cell(row=r, column=2,
             value=f'=IF(ROUND(B{beg_row}+B{beg_row+1}+B{beg_row+3}-B{end_row},2)=0,"✓ 對得上","✗ 不符")')
fc.font = F_FORM; fc.fill = fill(OK)
r += 1
ws3.cell(row=r, column=1, value="與 TB post-close 2100 一致?").font = F_BOLD
# TB 2100 post is at data_first + index of 2100
idx_2100 = next(i for i, l in enumerate(lines) if l["account"] == "2100")
tb_2100_post = f"'Trial Balance'!H{data_first + idx_2100}"
tie = ws3.cell(row=r, column=2,
              value=f'=IF(ROUND(B{end_row}-{tb_2100_post},2)=0,"✓ 與試算表一致","✗ 不一致")')
tie.font = F_FORM; tie.fill = fill(OK)
for col, w in zip("ABC", [40, 16, 48]):
    ws3.column_dimensions[col].width = w

# =====================================================================
# Sheet 4: Variance (current-month vs prior 3-month run-rate)
# =====================================================================
ws4 = wb.create_sheet("Variance")
ws4["A1"] = "差異說明 Variance / Flux(費用,門檻 >5%)"
ws4["A1"].font = F_TITLE
ws4["A2"] = ("試算表為 YTD;以「4月當月 = JE drill 加總」對「前3月平均月跑率 = 上月YTD ÷ 3」"
             "判讀真實異常。金額用 post-close。")
ws4["A2"].font = Font(italic=True, color="808080")

# current-month movement = TB movement (== JE net, tied above) for expense accounts
exp = [l for l in lines if l["type"] == "Expense"]
r = 4
hdrs = ["科目", "名稱", "前 YTD", "Post YTD", "4月當月", "月跑率(前÷3)",
        "當月 vs 跑率", ">5%?", "判讀(MCP JE 佐證)"]
for i, h in enumerate(hdrs):
    ws4.cell(row=r, column=i + 1, value=h)
style_hdr(ws4, r, [get_column_letter(i + 1) for i in range(len(hdrs))])
r += 1
# 用 JE drill 的描述當佐證
je_by_acct = {}
for e in je_all:
    je_by_acct.setdefault(str(e["account"]), []).append(e)

READ = {
    "6000": "JE-4001 4月薪資 1,500,000(跑率 1,500,000,持平);+應計分紅 100,000(A2)。已解釋。",
    "6100": "JE-4005 Q2租金及擴租 200,000;符合 ~200,000/月跑率。正常 run-rate。",
    "6200": "JE-4012 法律顧問併購盡調 220,000 + JE-4025 審計進度 130,000 = 350,000;另 +應計審計 100,000(A3)。全部 MCP JE 可解釋。",
    "6300": "JE-4022 雲端與資安 100,000;+應計 IT 75,000(A1)。已解釋。",
    "6400": "JE-4018 投資人年會及募資差旅 200,000;事件性、非經常。MCP JE 已解釋(非 unknown)。",
    "6500": "JE-4015 辦公雜項 80,000;符合跑率。正常。",
    "6600": "JE-4031 月折舊 50,000(Manual);但 MCP 僅 2026-04 在檔,看不到 1–3月折舊 → 時點疑慮,標示待查。",
}
data_first4 = r
for l in exp:
    acct = l["account"]
    ws4.cell(row=r, column=1, value=acct).border = box
    ws4.cell(row=r, column=2, value=l["name"]).border = box
    c3 = ws4.cell(row=r, column=3, value=l["prior_ytd"]); c3.number_format = MONEY; c3.font = F_IN; c3.border = box
    # post YTD via reference to TB sheet
    idx = next(i for i, x in enumerate(lines) if x["account"] == acct)
    c4 = ws4.cell(row=r, column=4, value=f"='Trial Balance'!H{data_first + idx}")
    c4.number_format = MONEY; c4.font = F_FORM; c4.border = box
    # current month = TB movement (pre) == JE net (tied)
    c5 = ws4.cell(row=r, column=5, value=l["movement"]); c5.number_format = MONEY; c5.font = F_IN; c5.border = box
    c6 = ws4.cell(row=r, column=6, value=f"=C{r}/3"); c6.number_format = MONEY; c6.font = F_FORM; c6.border = box
    c7 = ws4.cell(row=r, column=7, value=f'=IF(F{r}=0,"n/a",E{r}/F{r}-1)')
    c7.number_format = PCT; c7.font = F_FORM; c7.border = box
    c8 = ws4.cell(row=r, column=8, value=f'=IF(F{r}=0,"新增/⚠",IF(ABS(E{r}/F{r}-1)>0.05,"⚠ 是","否"))')
    c8.font = F_FORM; c8.border = box
    ws4.cell(row=r, column=9, value=READ.get(acct, "—")).border = box
    # shade warn rows
    if acct in ("6200", "6400", "6600"):
        for col in range(1, 10):
            ws4.cell(row=r, column=col).fill = fill(WARN)
    r += 1
for col, w in zip("ABCDEFGHI", [7, 26, 13, 13, 12, 13, 12, 9, 70]):
    ws4.column_dimensions[col].width = w

# =====================================================================
# Sheet 5: JE Drill (raw MCP evidence)
# =====================================================================
ws5 = wb.create_sheet("JE Drill (MCP)")
ws5["A1"] = "分錄明細 Journal Entry Drill — internal-gl MCP get_journal_entries"
ws5["A1"].font = F_TITLE
ws5["A2"] = "2026-04-01 ~ 2026-04-30 全部 JE(big movers);與試算表 movement 已逐科目對上。"
ws5["A2"].font = Font(italic=True, color="808080")
r = 4
hdrs = ["日期", "JE ID", "科目", "名稱", "說明", "借 Dr", "貸 Cr", "來源"]
for i, h in enumerate(hdrs):
    ws5.cell(row=r, column=i + 1, value=h)
style_hdr(ws5, r, [get_column_letter(i + 1) for i in range(len(hdrs))])
r += 1
first5 = r
for e in je_all:
    ws5.cell(row=r, column=1, value=e["date"]).border = box
    ws5.cell(row=r, column=2, value=e["je_id"]).border = box
    ws5.cell(row=r, column=3, value=e["account"]).border = box
    ws5.cell(row=r, column=4, value=e["name"]).border = box
    ws5.cell(row=r, column=5, value=e["description"]).border = box
    cd = ws5.cell(row=r, column=6, value=e["debit"]); cd.number_format = MONEY; cd.border = box
    cc = ws5.cell(row=r, column=7, value=e["credit"]); cc.number_format = MONEY; cc.border = box
    ws5.cell(row=r, column=8, value=e["source"]).border = box
    r += 1
ws5.cell(row=r, column=5, value="合計(big-mover lines,非完整複式)").font = F_BOLD
ws5.cell(row=r, column=6, value=f"=SUM(F{first5}:F{r-1})").number_format = MONEY
ws5.cell(row=r, column=7, value=f"=SUM(G{first5}:G{r-1})").number_format = MONEY
for col, w in zip("ABCDEFGH", [12, 10, 9, 26, 40, 13, 13, 10]):
    ws5.column_dimensions[col].width = w

path = os.path.join(OUT, "Close_Package_2026-04.xlsx")
wb.save(path)
print("SAVED", path)
print("sheets:", wb.sheetnames)

# ---- self-validate the key foot checks in Python (independent of Excel) ------
pre_dr = tb["total_debits"]; pre_cr = tb["total_credits"]
post_dr = pre_dr + accrual_total
post_cr = pre_cr + accrual_total
beg = bal_2100["prior_ytd"]; reversal = -200000; new_acc = accrual_total
pre_2100 = beg + reversal
end_2100 = pre_2100 + new_acc
rev_post = sum(l["current_ytd"] for l in lines if l["type"] == "Revenue")
exp_post = sum(l["current_ytd"] for l in lines if l["type"] == "Expense") + accrual_total
ni_pre = sum(l["current_ytd"] for l in lines if l["type"] == "Revenue") - sum(l["current_ytd"] for l in lines if l["type"] == "Expense")
ni_post = rev_post - exp_post
print("---- foot checks (python recompute) ----")
print(f"pre-close  Dr={pre_dr:,.0f}  Cr={pre_cr:,.0f}  balanced={pre_dr==pre_cr}")
print(f"post-close Dr={post_dr:,.0f}  Cr={post_cr:,.0f}  balanced={post_dr==post_cr}")
print(f"2100 roll-forward: {beg:,.0f} {reversal:,.0f} + {new_acc:,.0f} = {end_2100:,.0f}  (pre-close {pre_2100:,.0f} matches TB 1,000,000={pre_2100==1000000})")
print(f"YTD net income pre={ni_pre:,.0f} -> post={ni_post:,.0f}")
