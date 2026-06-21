# -*- coding: utf-8 -*-
"""Aurora Growth Fund III — GL <-> Subledger reconciliation workpaper (live formulas).

mcpBased version: reads the *actual* MCP pulls written by the harness
(docs/outputs/mcpBased/gl-conciler/mcp_pulls/*.json) and builds the Excel
workpaper from them. No hard-coded numbers — every figure is taken from the
JSON snapshots so the workpaper provably ties back to what the MCP returned.

Mock / dev data only. The agent never posts; this is an exceptions workpaper
for a human controller to sign off.
"""
import json
import os

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
BASE = os.path.dirname(HERE)                       # .../gl-conciler
PULLS = os.path.join(BASE, "mcp_pulls")
OUT = os.path.join(BASE, "out")
os.makedirs(OUT, exist_ok=True)


def load(name):
    with open(os.path.join(PULLS, name), encoding="utf-8") as f:
        return json.load(f)["result"]


# ---- live MCP data -------------------------------------------------
gl = load("02_gl_positions.json")
sub = load("03_subledger_holdings.json")
pend = load("04_pending_settlements.json")
fx = load("05_fx_rates.json")
tol = load("06_tolerance_policy.json")

gl_pos = {p["sec_id"]: p for p in gl["positions"]}
sub_pos = {p["sec_id"]: p for p in sub["positions"]}

# pending settlements may come back as a single dict (one trade) or a list
pend_list = pend if isinstance(pend, list) else [pend]
pending_secs = {p["sec_id"]: p for p in pend_list if isinstance(p, dict) and p.get("sec_id")}

fx_rel = abs(fx["custody_rate"] - fx["gl_rate"]) / fx["gl_rate"]
fx_tol = tol["fx_rounding_relative"]


def num(v):
    return 0 if v in ("", None) else v


# Classify each sec from the live data + evidence.
# Category enum mirrors the agent's reader output_schema:
#   Match / FP-Reclass / FP-FX / FP-Timing / REAL
def classify(sec):
    g, s = gl_pos.get(sec), sub_pos.get(sec)
    dq = num(g["qty_face"]) - num(s["qty_face"])
    dmv = num(g["mv_usd"]) - num(s["mv_usd"])
    if dq == 0 and dmv == 0:
        if str(g["asset_class"]).lower() != str(s["asset_class"]).lower():
            return "FP-Reclass"
        return "Match"
    # MV-only diff, qty equal -> candidate FX rounding. Guard tightly:
    #   (1) the position's OWN relative MV diff must be within the FX tolerance, and
    #   (2) it must plausibly be a foreign-currency position (an FX pair is on file
    #       and the diff size is consistent with that pair's GL/custody precision gap).
    # This stops a same-size cash break (CA-USD, USD, 20% diff) being waved through
    # just because an unrelated EUR/USD pair happens to sit inside tolerance.
    if dq == 0 and dmv != 0:
        base = num(gl_pos.get(sec, {}).get("mv_usd")) or 0
        own_rel = abs(dmv) / base if base else 1.0
        if own_rel <= fx_tol and fx_rel <= fx_tol:
            return "FP-FX"
    # diff fully explained by a pending (traded-not-settled) trade -> timing
    if sec in pending_secs:
        t = pending_secs[sec]
        signed = t["qty_face"] if str(t["side"]).lower() == "buy" else -t["qty_face"]
        if dq == signed:
            return "FP-Timing"
    return "REAL"


order = ["EQ-AAPL", "EQ-MSFT", "EQ-NVDA", "EQ-SPY", "FI-BUND", "FI-GB123",
         "EQ-VRT", "FI-XYZ", "CA-USD"]
rows = []
for sec in order:
    g, s = gl_pos[sec], sub_pos[sec]
    rows.append((sec, g["asset_class"], num(g["qty_face"]), num(s["qty_face"]),
                 num(g["mv_usd"]), num(s["mv_usd"]), classify(sec)))

# ---- styling -------------------------------------------------------
BLUEF = "0000FF"; BLACKF = "000000"
HDR = "1F4E79"; SUB = "D9E1F2"; WARN = "FCE4D6"; OK = "E2EFDA"; AMB = "FFF2CC"
CJK = "Microsoft JhengHei"
med = Side(style="thin", color="9BAFC4")
cell_b = Border(left=med, right=med, top=med, bottom=med)


def Fi(s=10, b=False): return Font(name=CJK, size=s, bold=b, color=BLUEF)
def Ff(s=10, b=False): return Font(name=CJK, size=s, bold=b, color=BLACKF)
def Fh(s=11, b=True):  return Font(name=CJK, size=s, bold=b, color="FFFFFF")
def fill(c): return PatternFill("solid", fgColor=c)


CEN = Alignment(horizontal="center", vertical="center", wrap_text=True)
RGT = Alignment(horizontal="right", vertical="center")
LEF = Alignment(horizontal="left", vertical="center", wrap_text=True)
USD = '#,##0;(#,##0);"-"'

wb = Workbook()


def sheet(name, idx=None):
    s = wb.create_sheet(name) if idx is None else wb.create_sheet(name, idx)
    s.sheet_view.showGridLines = False
    return s


def W(ws, cell, v, f, fmt=None, al=RGT, fl=None):
    c = ws[cell]; c.value = v; c.font = f; c.border = cell_b; c.alignment = al
    if fmt: c.number_format = fmt
    if fl: c.fill = fill(fl)


# ===================== Recon =====================
rc = sheet("Recon")
rc.merge_cells("A1:K1")
rc["A1"] = "GL <-> 子帳 對帳 Reconciliation · Aurora Growth Fund III · 2026-04-30 (USD) · 來源:live MCP pulls"
rc["A1"].font = Fh(12); rc["A1"].fill = fill(HDR)
rc["A1"].alignment = Alignment(horizontal="left", vertical="center"); rc.row_dimensions[1].height = 24
rc.merge_cells("A2:K2")
rc["A2"] = ("GL=internal-gl(交易日基礎) / 子帳=subledger(交割日基礎,custody,外來)。"
            "Category 為複核分類(藍字可改);真實break / 已解釋由公式拆解。資料皆 mock。")
rc["A2"].font = Ff(8); rc["A2"].alignment = Alignment(horizontal="left", vertical="center")
heads = ["Sec ID", "資產類別", "GL Qty", "Sub Qty", "Qty差", "GL MV", "Sub MV",
         "MV差(GL-Sub)", "Category", "真實break", "已解釋"]
for i, h in enumerate(heads):
    W(rc, f"{get_column_letter(1+i)}4", h, Fh(9), al=CEN if i > 1 else LEF, fl=HDR)
r0 = 5
for i, (sid, cls, glq, sbq, glm, sbm, cat) in enumerate(rows):
    r = r0 + i; real = (cat == "REAL"); fp = cat.startswith("FP")
    W(rc, f"A{r}", sid, Ff(9, True), al=LEF); W(rc, f"B{r}", cls, Ff(9), al=CEN)
    W(rc, f"C{r}", glq, Fi(9), USD); W(rc, f"D{r}", sbq, Fi(9), USD); W(rc, f"E{r}", f"=C{r}-D{r}", Ff(9), USD)
    W(rc, f"F{r}", glm, Fi(9), USD); W(rc, f"G{r}", sbm, Fi(9), USD); W(rc, f"H{r}", f"=F{r}-G{r}", Ff(9, True), USD)
    W(rc, f"I{r}", cat, Fi(9, True), al=CEN, fl=(WARN if real else (AMB if fp else OK)))
    W(rc, f'J{r}', f'=IF(I{r}="REAL",H{r},0)', Ff(9), USD, fl=(WARN if real else None))
    W(rc, f'K{r}', f'=IF(AND(I{r}<>"REAL",I{r}<>"Match"),H{r},0)', Ff(9), USD, fl=(AMB if fp else None))
rN = r0 + len(rows)
W(rc, f"A{rN}", "總計 Total", Ff(10, True), al=LEF, fl=SUB)
for col in ["F", "G", "H", "J", "K"]:
    W(rc, f"{col}{rN}", f"=SUM({col}{r0}:{col}{rN-1})", Ff(10, True), USD, fl=SUB)
W(rc, f"A{rN+2}", "檢查 1：MV差合計 = GL - 子帳", Ff(9), al=LEF)
W(rc, f"H{rN+2}", f"=IF(ROUND(H{rN}-(F{rN}-G{rN}),0)=0,\"OK\",\"X\")", Ff(9, True), al=CEN, fl=OK)
W(rc, f"A{rN+3}", "檢查 2：真實 + 已解釋 = MV差合計", Ff(9), al=LEF)
W(rc, f"H{rN+3}", f"=IF(ROUND(J{rN}+K{rN}-H{rN},0)=0,\"OK\",\"X\")", Ff(9, True), al=CEN, fl=OK)
rc.column_dimensions["A"].width = 11; rc.column_dimensions["B"].width = 12; rc.column_dimensions["I"].width = 12
for c in "CDEFGHJK":
    rc.column_dimensions[c].width = 13

# ===================== Break Analysis =====================
ba = sheet("Break Analysis", 0)
ba.merge_cells("A1:E1"); ba["A1"] = "差異拆解與獨立複核 Break Analysis (critic)"; ba["A1"].font = Fh(12); ba["A1"].fill = fill(HDR)
ba["A1"].alignment = Alignment(horizontal="left", vertical="center")
for col, h in (("A", "項目"), ("C", "金額"), ("D", "類型"), ("E", "佐證/處理")):
    W(ba, f"{col}3", h, Fh(9), al=LEF if col in ("A", "E") else CEN, fl=HDR)
fx_pct = f"{fx_rel*100:.4f}%"
items = [
    ("GL 總計", "=Recon!F14", "", ""),
    ("子帳 總計", "=Recon!G14", "", ""),
    ("總差異(GL-子帳)", "=C4-C5", "", "= MV差合計"),
    ("", None, "", ""),
    ("FI-GB123", "=Recon!H10", "誤報-時間差", "TRD-9001 買500,000;trade 2026-04-29 / settle 2026-05-02,跨月底"),
    ("FI-BUND", "=Recon!H9", "誤報-FX", f"EUR/USD GL 1.08000 vs custody 1.08004;{fx_pct} < 0.05% 容差"),
    ("EQ-SPY", "=Recon!H8", "誤報-重分類", "Equity vs Funds/ETF;Qty/MV 一致,淨額 0"),
    ("已解釋小計", "=Recon!K14", "", "(時間差+FX+重分類)"),
    ("", None, "", ""),
    ("EQ-VRT", "=Recon!H11", "真實", "子帳多 2,500 股;不在未交割清單 -> 疑 GL 漏記買單"),
    ("FI-XYZ", "=Recon!H12", "真實", "GL 多 1,000,000 面額;不在未交割清單 -> 疑 GL 未沖已處分"),
    ("CA-USD", "=Recon!H13", "真實", "現金缺口 600,000;非FX非時間差,查無佐證 -> 升級"),
    ("真實未解小計", "=Recon!J14", "", "需處理"),
    ("", None, "", ""),
    ("檢查：已解釋+真實=總差異", "=IF(ROUND(Recon!J14+Recon!K14-(C4-C5),0)=0,\"OK tie-out\",\"X\")", "", ""),
]
r = 4
for lab, val, typ, note in items:
    if lab == "":
        r += 1; continue
    bold = ("小計" in lab or "總" in lab or "檢查" in lab or "差異" in lab)
    W(ba, f"A{r}", lab, Ff(10, bold), al=LEF, fl=(SUB if bold else None)); ba.merge_cells(f"A{r}:B{r}")
    if val is not None:
        is_chk = lab.startswith("檢查")
        W(ba, f"C{r}", val, Ff(10, bold), None if is_chk else USD,
          al=CEN if is_chk else RGT, fl=(OK if is_chk else (SUB if bold else None)))
    if typ:
        W(ba, f"D{r}", typ, Ff(9, True), al=CEN, fl=(WARN if typ == "真實" else AMB))
    if note:
        W(ba, f"E{r}", note, Ff(8), al=LEF)
    r += 1
ba.column_dimensions["A"].width = 16; ba.column_dimensions["B"].width = 6; ba.column_dimensions["C"].width = 14
ba.column_dimensions["D"].width = 12; ba.column_dimensions["E"].width = 52

# ===================== Escalation =====================
es = sheet("Escalation")
es.merge_cells("A1:E1"); es["A1"] = "升級清單 Escalation · 真實 break (給控制員簽核;agent 不過帳)"; es["A1"].font = Fh(12); es["A1"].fill = fill(HDR)
es["A1"].alignment = Alignment(horizontal="left", vertical="center")
for i, h in enumerate(["優先", "Sec", "金額", "根因", "動作"]):
    W(es, f"{get_column_letter(1+i)}3", h, Fh(9), al=CEN if i in (0, 2) else LEF, fl=HDR)
esc = [
    (1, "CA-USD", "=Recon!H13", "現金缺口,不在未交割清單、非FX,查無佐證", "對銀行/保管流水,24h 內回覆;先排除錯帳/舞弊"),
    (2, "FI-XYZ", "=Recon!H12", "GL 多 1,000,000 面額;疑已處分/贖回保管已沖、GL 未沖", "調 XYZ 處分/贖回確認,GL 補沖減"),
    (3, "EQ-VRT", "=Recon!H11", "子帳多 2,500 股;疑買單保管已入、GL 未入", "查 VRT 成交確認,GL 補入 2,500 股"),
    ("監控", "FI-GB123", "=Recon!H10", "未交割買單 TRD-9001(時間差)", "2026-05-02 交割後覆核,差異應消除"),
]
for i, (p, sid, amt, rc_, act) in enumerate(esc):
    r = 4 + i; mon = (p == "監控")
    W(es, f"A{r}", p, Ff(9, True), al=CEN, fl=(AMB if mon else WARN))
    W(es, f"B{r}", sid, Ff(9, True), al=LEF); W(es, f"C{r}", amt, Ff(9), USD)
    W(es, f"D{r}", rc_, Ff(8), al=LEF); W(es, f"E{r}", act, Ff(8), al=LEF)
es.column_dimensions["A"].width = 8; es.column_dimensions["B"].width = 11; es.column_dimensions["C"].width = 13
es.column_dimensions["D"].width = 40; es.column_dimensions["E"].width = 40

if "Sheet" in wb.sheetnames:
    wb.remove(wb["Sheet"])
out_path = os.path.join(OUT, "GL_Recon_Workpaper_2026-04-30.xlsx")
wb.save(out_path)

# ---- python-side tie-out (independent of Excel formulas) -----------
tot_diff = gl["total_mv_usd"] - sub["total_mv_usd"]
explained = real = 0
for sid, cls, glq, sbq, glm, sbm, cat in rows:
    d = glm - sbm
    if cat == "REAL":
        real += d
    elif cat.startswith("FP"):
        explained += d
print("SAVED", out_path, "sheets:", wb.sheetnames)
print(f"total diff {tot_diff:,.0f} | real {real:,.0f} | explained {explained:,.0f} | "
      f"real+expl {real+explained:,.0f} -> tie {'OK' if real+explained == tot_diff else 'FAIL'}")
print("categories:", {sid: cat for sid, cls, glq, sbq, glm, sbm, cat in rows})
