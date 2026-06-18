# -*- coding: utf-8 -*-
"""Aurora Growth Fund III — GL↔Subledger reconciliation workpaper (live formulas)."""
import os
os.makedirs("./out", exist_ok=True)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BLUEF="0000FF"; BLACKF="000000"; GREENF="008000"
HDR="1F4E79"; SUB="D9E1F2"; OUT="BDD7EE"; INP="F2F2F2"; WHITE="FFFFFF"; WARN="FCE4D6"; OK="E2EFDA"; AMB="FFF2CC"
CJK="Microsoft JhengHei"
med=Side(style="thin",color="9BAFC4"); cell_b=Border(left=med,right=med,top=med,bottom=med)
def Fi(s=10,b=False): return Font(name=CJK,size=s,bold=b,color=BLUEF)
def Ff(s=10,b=False): return Font(name=CJK,size=s,bold=b,color=BLACKF)
def Fh(s=11,b=True):  return Font(name=CJK,size=s,bold=b,color=WHITE)
def fill(c): return PatternFill("solid",fgColor=c)
CEN=Alignment(horizontal="center",vertical="center",wrap_text=True)
RGT=Alignment(horizontal="right",vertical="center"); LEF=Alignment(horizontal="left",vertical="center",wrap_text=True)
USD='#,##0;(#,##0);"-"'

wb=Workbook()
def sheet(name,idx=None):
    s=wb.create_sheet(name) if idx is None else wb.create_sheet(name,idx); s.sheet_view.showGridLines=False; return s
def W(ws,cell,v,f,fmt=None,al=RGT,fl=None):
    c=ws[cell]; c.value=v; c.font=f; c.border=cell_b; c.alignment=al
    if fmt:c.number_format=fmt
    if fl:c.fill=fill(fl)
def hdr(ws,rng,t):
    a=rng.split(":")[0]; ws[a]=t; ws[a].font=Fh(11); ws.merge_cells(rng)
    for row in ws[rng]:
        for c in row: c.fill=fill(HDR); c.border=cell_b
    ws[a].alignment=Alignment(horizontal="left",vertical="center")

# ===================== Recon =====================
rc=sheet("Recon")
rc.merge_cells("A1:K1"); rc["A1"]="GL ↔ 子帳 對帳 Reconciliation · Aurora Growth Fund III · 2026-04-30 (USD)"
rc["A1"].font=Fh(13); rc["A1"].fill=fill(HDR); rc["A1"].alignment=Alignment(horizontal="left",vertical="center"); rc.row_dimensions[1].height=24
rc.merge_cells("A2:K2"); rc["A2"]="GL=交易日基礎 / 子帳=交割日基礎(custody,外來)。Category 為複核分類(藍字輸入);Real/Explained 由公式拆解。"
rc["A2"].font=Ff(8); rc["A2"].alignment=Alignment(horizontal="left",vertical="center")
heads=["Sec ID","資產類別","GL Qty","Sub Qty","Qty差","GL MV","Sub MV","MV差(GL−Sub)","Category","真實break","已解釋"]
for i,h in enumerate(heads): W(rc,f"{get_column_letter(1+i)}4",h,Fh(9),al=CEN if i>1 else LEF,fl=HDR)
rows=[
 ("EQ-AAPL","Equity",50000,50000,10000000,10000000,"Match"),
 ("EQ-MSFT","Equity",20000,20000,8000000,8000000,"Match"),
 ("EQ-NVDA","Equity",10000,10000,12000000,12000000,"Match"),
 ("EQ-SPY","Equity",5000,5000,2000000,2000000,"FP-Reclass"),
 ("FI-BUND","Fixed Income",8333333,8333333,9000000,9000300,"FP-FX"),
 ("FI-GB123","Fixed Income",15500000,15000000,15500000,15000000,"FP-Timing"),
 ("EQ-VRT","Equity",50000,52500,5000000,5250000,"REAL"),
 ("FI-XYZ","Fixed Income",20000000,19000000,20000000,19000000,"REAL"),
 ("CA-USD","Cash",0,0,3000000,2400000,"REAL"),
]
r0=5
for i,(sid,cls,glq,sbq,glm,sbm,cat) in enumerate(rows):
    r=r0+i; real=(cat=="REAL"); fp=cat.startswith("FP")
    W(rc,f"A{r}",sid,Ff(9,True),al=LEF); W(rc,f"B{r}",cls,Ff(9),al=CEN)
    W(rc,f"C{r}",glq,Fi(9),USD); W(rc,f"D{r}",sbq,Fi(9),USD); W(rc,f"E{r}",f"=C{r}-D{r}",Ff(9),USD)
    W(rc,f"F{r}",glm,Fi(9),USD); W(rc,f"G{r}",sbm,Fi(9),USD); W(rc,f"H{r}",f"=F{r}-G{r}",Ff(9,True),USD)
    W(rc,f"I{r}",cat,Fi(9,True),al=CEN,fl=(WARN if real else (AMB if fp else OK)))
    W(rc,f'J{r}',f'=IF(I{r}="REAL",H{r},0)',Ff(9),USD,fl=(WARN if real else None))
    W(rc,f'K{r}',f'=IF(AND(I{r}<>"REAL",I{r}<>"Match"),H{r},0)',Ff(9),USD,fl=(AMB if fp else None))
rN=r0+len(rows)
W(rc,f"A{rN}","總計 Total",Ff(10,True),al=LEF,fl=SUB)
for col in ["F","G","H","J","K"]:
    W(rc,f"{col}{rN}",f"=SUM({col}{r0}:{col}{rN-1})",Ff(10,True),USD,fl=SUB)
# tie-out checks
W(rc,f"A{rN+2}","檢查 1：MV差合計 = GL − 子帳",Ff(9),al=LEF)
W(rc,f"H{rN+2}",f"=IF(ROUND(H{rN}-(F{rN}-G{rN}),0)=0,\"✓\",\"✗\")",Ff(9,True),al=CEN,fl=OK)
W(rc,f"A{rN+3}","檢查 2：真實 + 已解釋 = MV差合計",Ff(9),al=LEF)
W(rc,f"H{rN+3}",f"=IF(ROUND(J{rN}+K{rN}-H{rN},0)=0,\"✓\",\"✗\")",Ff(9,True),al=CEN,fl=OK)
rc.column_dimensions["A"].width=11; rc.column_dimensions["B"].width=12; rc.column_dimensions["I"].width=12
for c in "CDEFGHJK": rc.column_dimensions[c].width=13

# ===================== Break Analysis =====================
ba=sheet("Break Analysis",0)
ba.merge_cells("A1:E1"); ba["A1"]="差異拆解與獨立複核 Break Analysis"; ba["A1"].font=Fh(12); ba["A1"].fill=fill(HDR)
ba["A1"].alignment=Alignment(horizontal="left",vertical="center")
W(ba,"A3","項目",Fh(9),al=LEF,fl=HDR); W(ba,"C3","金額",Fh(9),al=CEN,fl=HDR); W(ba,"D3","類型",Fh(9),al=CEN,fl=HDR); W(ba,"E3","佐證/處理",Fh(9),al=LEF,fl=HDR)
items=[("GL 總計","=Recon!F14","",""),("子帳 總計","=Recon!G14","",""),
 ("總差異(GL−子帳)","=C4-C5","","= MV差合計"),
 ("",None,"",""),
 ("FI-GB123","=Recon!H10","誤報-時間差","TRD-9001 未交割買單,5/2 交割後消除"),
 ("FI-BUND","=Recon!H9","誤報-FX","0.003% < 0.05% 容差"),
 ("EQ-SPY","=Recon!H8","誤報-重分類","僅分類欄不同,Qty/MV 一致"),
 ("已解釋小計","=Recon!K14","","(時間差+FX+重分類)"),
 ("",None,"",""),
 ("EQ-VRT","=Recon!H11","真實","子帳多2,500股,疑GL漏記買單"),
 ("FI-XYZ","=Recon!H12","真實","GL多1,000,000面額,疑GL未沖處分"),
 ("CA-USD","=Recon!H13","真實","現金缺口600,000,查無佐證→升級"),
 ("真實未解小計","=Recon!J14","","需處理"),
 ("",None,"",""),
 ("檢查：已解釋+真實=總差異","=IF(ROUND(Recon!J14+Recon!K14-(C4-C5),0)=0,\"✓ tie-out\",\"✗\")","",""),
]
r=4
for lab,val,typ,note in items:
    if lab=="":
        r+=1; continue
    bold=("小計" in lab or "總" in lab or "檢查" in lab or "差異" in lab)
    W(ba,f"A{r}",lab,Ff(10,bold),al=LEF,fl=(SUB if bold else None)); ba.merge_cells(f"A{r}:B{r}")
    if val is not None:
        is_chk=lab.startswith("檢查")
        W(ba,f"C{r}",val,Ff(10,bold),None if is_chk else USD,al=CEN if is_chk else RGT,fl=(OK if is_chk else (SUB if bold else None)))
    if typ: W(ba,f"D{r}",typ,Ff(9,True),al=CEN,fl=(WARN if typ=="真實" else AMB))
    if note: W(ba,f"E{r}",note,Ff(8),al=LEF)
    r+=1
ba.column_dimensions["A"].width=16; ba.column_dimensions["B"].width=6; ba.column_dimensions["C"].width=14
ba.column_dimensions["D"].width=12; ba.column_dimensions["E"].width=40

# ===================== Escalation =====================
es=sheet("Escalation")
es.merge_cells("A1:E1"); es["A1"]="升級清單 Escalation · 真實 break"; es["A1"].font=Fh(12); es["A1"].fill=fill(HDR)
es["A1"].alignment=Alignment(horizontal="left",vertical="center")
for i,h in enumerate(["優先","Sec","金額","根因","動作"]):
    W(es,f"{get_column_letter(1+i)}3",h,Fh(9),al=CEN if i in(0,2) else LEF,fl=HDR)
esc=[(1,"CA-USD","=Recon!H13","現金缺口,不在未交割清單、非FX,查無佐證","對銀行/保管流水,24h 內回覆;先排除錯帳/舞弊"),
 (2,"FI-XYZ","=Recon!H12","GL 多 1,000,000 面額;疑已處分/贖回保管已沖、GL 未沖","調 XYZ 處分/贖回確認,GL 補沖減"),
 (3,"EQ-VRT","=Recon!H11","子帳多 2,500 股;疑買單保管已入、GL 未入","查 VRT 成交確認,GL 補入 2,500 股"),
 ("監控","FI-GB123","=Recon!H10","未交割買單(時間差)","2026-05-02 交割後覆核,差異應消除")]
for i,(p,sid,amt,rc_,act) in enumerate(esc):
    r=4+i; mon=(p=="監控")
    W(es,f"A{r}",p,Ff(9,True),al=CEN,fl=(AMB if mon else WARN))
    W(es,f"B{r}",sid,Ff(9,True),al=LEF); W(es,f"C{r}",amt,Ff(9),USD)
    W(es,f"D{r}",rc_,Ff(8),al=LEF); W(es,f"E{r}",act,Ff(8),al=LEF)
es.column_dimensions["A"].width=8; es.column_dimensions["B"].width=11; es.column_dimensions["C"].width=13
es.column_dimensions["D"].width=34; es.column_dimensions["E"].width=36

if "Sheet" in wb.sheetnames: wb.remove(wb["Sheet"])
wb.save("./out/GL_Recon_Workpaper_2026-04-30.xlsx")
print("SAVED ./out/GL_Recon_Workpaper_2026-04-30.xlsx  sheets:", wb.sheetnames)
