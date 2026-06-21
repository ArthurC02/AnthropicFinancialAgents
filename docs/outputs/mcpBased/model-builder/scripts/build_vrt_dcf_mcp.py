# -*- coding: utf-8 -*-
"""Vertiv (VRT) DCF — MCP-sourced build per model-builder dcf-model skill.

資料來源（全部來自實際 MCP pull，存在 ../mcp_pulls/*.json）:
  - Daloopa get_line_items("VRT")  → 歷史財務（FY2025 營收 8.0、營業利益 1.5、backlog 15.0、
                                     分部 Americas/APAC/EMEA、產品線 Power/Thermal/Rack/Services）
  - CapIQ  get_quote("VRT")        → 現價 320、市值 122bn（→ 隱含股數 0.38125bn）
  - CapIQ  get_estimates("VRT")    → FY26 營收 10.9 / EBITDA margin 22% / EPS 3.60；FY27 營收 13.5 / margin 23%
  - CapIQ  get_comps(...)          → VRT EV/EBITDA 51x、fwd P/E 46x；同業 NVT 32x、SU.PA 21x（football field）

色碼慣例（dcf-model skill）：藍字=輸入(input)、黑字=公式(formula)、綠字=跨頁連結(link)。
計算格一律公式、不寫死。Daloopa 未提供之項目（淨負債、D&A、capex、NWC、beta…）標 [ASSUMPTION]。
兩個分頁：DCF（含 5x5 WACC×g 敏感度表）+ WACC（CAPM 推導）。
Mock / dev only — 數字皆為假資料。
"""
import os, json
HERE = os.path.dirname(os.path.abspath(__file__))
OUT  = os.path.join(HERE, "..", "out")
PULLS = os.path.join(HERE, "..", "mcp_pulls")
os.makedirs(OUT, exist_ok=True)

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter

# ----------------------------------------------------------------- read MCP pulls
def load(name):
    with open(os.path.join(PULLS, name), encoding="utf-8") as f:
        return json.load(f)["result"]

dal = load("02_daloopa_vrt_all_line_items.json")
quote = load("09_capiq_quote_vrt.json")
est = load("10_capiq_estimates_vrt.json")

def dal_get(stmt, item):
    for r in dal:
        if r["statement"] == stmt and r["line_item"] == item:
            return r["value"]
    return None

def est_get(period, metric):
    for r in est:
        if r["period"] == period and r["metric"] == metric:
            return r["value"]
    return None

REV_FY25 = dal_get("income_statement", "Revenue")          # 8.0
OPINC_FY25 = dal_get("income_statement", "Operating Profit")# 1.5
BACKLOG = dal_get("balance_sheet", "Backlog")              # 15.0
PRICE = quote["price"]                                      # 320.0
MKTCAP = quote["mktcap_usd_bn"]                             # 122
SHARES = round(MKTCAP / PRICE, 5)                           # 0.38125  (derived link)
REV_FY26E = est_get("FY2026", "revenue")                   # 10.9
REV_FY27E = est_get("FY2027", "revenue")                   # 13.5
EBITDA_M_FY26 = est_get("FY2026", "ebitda_margin")         # 0.22
OPMARGIN_FY25 = round(OPINC_FY25 / REV_FY25, 4)            # 0.1875
G_FY26 = round(REV_FY26E / REV_FY25 - 1, 4)               # +0.3625 (Daloopa 8.0 -> CapIQ 10.9)
G_FY27 = round(REV_FY27E / REV_FY26E - 1, 4)              # +0.2385

print(f"MCP inputs: REV25={REV_FY25} OP25={OPINC_FY25} (m={OPMARGIN_FY25:.2%}) backlog={BACKLOG} "
      f"price={PRICE} mcap={MKTCAP} -> shares={SHARES} | FY26E={REV_FY26E} (g={G_FY26:.2%}) "
      f"FY27E={REV_FY27E} (g={G_FY27:.2%})")

# ----------------------------------------------------------------- palette
BLUEF="0000FF"; BLACKF="000000"; GREENF="008000"
HDR="1F4E79"; SUB="D9E1F2"; OUTC="BDD7EE"; INP="FFF8E1"; WHITE="FFFFFF"; CTR="BDD7EE"; SRC="E2EFDA"
CJK="Microsoft JhengHei"
med=Side(style="thin",color="9BAFC4")
cell_b=Border(left=med,right=med,top=med,bottom=med)
def Fi(sz=10,b=False): return Font(name=CJK,size=sz,bold=b,color=BLUEF)   # input 藍
def Ff(sz=10,b=False): return Font(name=CJK,size=sz,bold=b,color=BLACKF)  # formula 黑
def Fl(sz=10,b=False): return Font(name=CJK,size=sz,bold=b,color=GREENF)  # link 綠
def Fh(sz=11,b=True):  return Font(name=CJK,size=sz,bold=b,color=WHITE)   # header
def fill(c): return PatternFill("solid",fgColor=c)
CEN=Alignment(horizontal="center",vertical="center",wrap_text=True)
RGT=Alignment(horizontal="right",vertical="center")
LEF=Alignment(horizontal="left",vertical="center",wrap_text=True)
PCT="0.0%"; PCTS="0.0%;(0.0%)"; BN='$#,##0.00;($#,##0.00);"-"'; SH2="$#,##0.00"; MULT='0.0"x"'; NUM="0.000"

wb=Workbook()

# ============================================================ WACC sheet
wc=wb.create_sheet("WACC"); wc.sheet_view.showGridLines=False
def wput(cell,val,font,fmt=None,al=RGT,bd=True,fl=None,cmt=None):
    c=wc[cell]; c.value=val; c.font=font
    if fmt:c.number_format=fmt
    c.alignment=al
    if bd:c.border=cell_b
    if fl:c.fill=fill(fl)
    if cmt:c.comment=Comment(cmt,"DCF skill")
    return c
def whdr(rng,text):
    a=rng.split(":")[0]; wc[a]=text; wc[a].font=Fh(11); wc.merge_cells(rng)
    for row in wc[rng]:
        for c in row: c.fill=fill(HDR); c.border=cell_b
    wc[a].alignment=Alignment(horizontal="left",vertical="center")

wc.merge_cells("B1:D1"); wc["B1"]="Vertiv (VRT) — WACC（資金成本,CAPM）"; wc["B1"].font=Fh(13); wc["B1"].fill=fill(HDR)
wc["B1"].alignment=Alignment(horizontal="left",vertical="center"); wc.row_dimensions[1].height=24
wc.merge_cells("B2:D2"); wc["B2"]="[ASSUMPTION] 區塊：Daloopa/CapIQ 未提供 beta、負債、現金、稅率,以下藍字為分析師輸入假設"
wc["B2"].font=Font(name=CJK,size=8,color=BLACKF); wc["B2"].alignment=Alignment(horizontal="left",vertical="center")

whdr("B4:D4","權益成本 COST OF EQUITY")
rows=[("無風險利率 (10Y UST) [ASSUMPTION]",0.043,PCT,Fi(),"[ASSUMPTION] MCP 未提供;US 10Y Treasury 2026-06 約 4.3%"),
      ("Beta (5年月beta) [ASSUMPTION]",1.45,NUM,Fi(),"[ASSUMPTION] MCP 未提供;高 beta AI 基建股,5 年月 beta 約 1.4-1.5。註:CapIQ comps 顯示 VRT 1 年報酬 +173%,屬高波動"),
      ("股權風險溢酬 ERP [ASSUMPTION]",0.05,PCT,Fi(),"[ASSUMPTION] MCP 未提供;市場慣用 5.0%")]
r=5
for lab,v,fmt,fn,cm in rows:
    wput(f"B{r}",lab,Ff(),al=LEF,fl=INP); wput(f"C{r}",v,fn,fmt,fl=INP,cmt=cm); r+=1
wput("B8","權益成本 Cost of Equity",Ff(10,True),al=LEF,fl=SUB); wput("C8","=C5+C6*C7",Ff(10,True),PCT,fl=SUB)

whdr("B10:D10","負債成本 COST OF DEBT")
wput("B11","稅前負債成本 [ASSUMPTION]",Ff(),al=LEF,fl=INP); wput("C11",0.055,Fi(),PCT,fl=INP,cmt="[ASSUMPTION] MCP 未提供;BB+/BBB- 信評隱含約 5.5%")
wput("B12","稅率 (連結 DCF)",Ff(),al=LEF); wput("C12","=DCF!$C$33",Fl(),PCT)
wput("B13","稅後負債成本",Ff(10,True),al=LEF,fl=SUB); wput("C13","=C11*(1-C12)",Ff(10,True),PCT,fl=SUB)

whdr("B15:D15","資本結構 CAPITAL STRUCTURE ($bn)")
wput("B16","股價 Stock Price (連結 DCF)",Ff(),al=LEF); wput("C16","=DCF!$C$9",Fl(),SH2)
wput("B17","股數 bn (連結 DCF)",Ff(),al=LEF); wput("C17","=DCF!$C$10",Fl(),NUM)
wput("B18","市值 Market Cap",Ff(),al=LEF); wput("C18","=C16*C17",Ff(),BN)
wput("B19","總負債 Total Debt [ASSUMPTION]",Ff(),al=LEF,fl=INP); wput("C19",2.90,Fi(),BN,fl=INP,cmt="[ASSUMPTION] Daloopa 未提供;VRT 資產負債表總負債約 $2.9bn")
wput("B20","現金 Cash & Equiv [ASSUMPTION]",Ff(),al=LEF,fl=INP); wput("C20",1.00,Fi(),BN,fl=INP,cmt="[ASSUMPTION] Daloopa 未提供;現金約 $1.0bn")
wput("B21","淨負債 Net Debt",Ff(10,True),al=LEF,fl=SUB); wput("C21","=C19-C20",Ff(10,True),BN,fl=SUB)
wput("B22","企業價值 EV (市值法)",Ff(),al=LEF); wput("C22","=C18+C21",Ff(),BN)

wput("B24","WACC 計算",Fh(10),al=LEF,fl=HDR)
wput("C24","權重",Fh(10),al=CEN,fl=HDR); wput("D24","成本",Fh(10),al=CEN,fl=HDR); wput("E24","貢獻",Fh(10),al=CEN,fl=HDR)
wput("B25","權益 Equity",Ff(),al=LEF); wput("C25","=C18/C22",Ff(),PCT); wput("D25","=C8",Ff(),PCT); wput("E25","=C25*D25",Ff(),PCT)
wput("B26","負債 Debt",Ff(),al=LEF); wput("C26","=C21/C22",Ff(),PCT); wput("D26","=C13",Ff(),PCT); wput("E26","=C26*D26",Ff(),PCT)
wput("B28","WACC",Ff(11,True),al=LEF,fl=OUTC); wc.merge_cells("C28:D28")
wput("C28","=E25+E26",Ff(11,True),PCT,al=CEN,fl=OUTC)
wc.column_dimensions["A"].width=2; wc.column_dimensions["B"].width=34
for c in "CDE": wc.column_dimensions[c].width=12

# ============================================================ DCF sheet
ws=wb.create_sheet("DCF",0); ws.sheet_view.showGridLines=False
YRS=["FY25A","FY26E","FY27E","FY28E","FY29E","FY30E"]  # cols C..H
def put(cell,val,font,fmt=None,al=RGT,bd=True,fl=None,cmt=None):
    c=ws[cell]; c.value=val; c.font=font
    if fmt:c.number_format=fmt
    c.alignment=al
    if bd:c.border=cell_b
    if fl:c.fill=fill(fl)
    if cmt:c.comment=Comment(cmt,"DCF skill")
    return c
def hdr(rng,text):
    a=rng.split(":")[0]; ws[a]=text; ws[a].font=Fh(11); ws.merge_cells(rng)
    for row in ws[rng]:
        for c in row: c.fill=fill(HDR); c.border=cell_b
    ws[a].alignment=Alignment(horizontal="left",vertical="center")

# -- header --
ws.merge_cells("B1:H1"); ws["B1"]="Vertiv Holdings (VRT) — DCF 估值模型（MCP 來源）"; ws["B1"].font=Fh(14); ws["B1"].fill=fill(HDR)
ws["B1"].alignment=Alignment(horizontal="left",vertical="center"); ws.row_dimensions[1].height=26
ws.merge_cells("B2:H2"); ws["B2"]=("Ticker: VRT | 日期: 2026-06-18 | 會計年底: 12月 | 單位: US$ bn(每股除外) | 期中折現慣例 | "
                                   "歷史=Daloopa,假設對齊 CapIQ estimates | mock/dev,數字皆為假資料")
ws["B2"].font=Font(name=CJK,size=8,color=BLACKF); ws["B2"].alignment=Alignment(horizontal="left",vertical="center")

# -- case selector --
put("B4","情境選擇 (1熊 2基 3牛)",Ff(10,True),al=LEF,fl=SUB)
put("C4",2,Fi(10,True),"0",al=CEN,fl=INP,cmt="1=Bear, 2=Base, 3=Bull;改此格切換情境")
put("B5","目前情境",Ff(),al=LEF); put("C5",'=IF($C$4=1,"熊 Bear",IF($C$4=2,"基 Base","牛 Bull"))',Ff(10,True),al=CEN)

# -- source data block (MCP provenance) --
hdr("B7:H7","MCP 來源資料 SOURCE DATA（綠=Daloopa/CapIQ 拉回 | 藍=假設）")
put("B8","FY25 營收 Revenue (Daloopa)",Ff(),al=LEF); put("C8",REV_FY25,Fl(),BN,fl=SRC,cmt="Daloopa get_line_items VRT income_statement Revenue = 8.0 USD bn")
put("B9","現價 Stock Price (CapIQ quote)",Ff(),al=LEF); put("C9",PRICE,Fl(),SH2,fl=SRC,cmt="CapIQ get_quote VRT price = 320.00")
put("B10","隱含股數 (CapIQ mcap/price)",Ff(),al=LEF); put("C10",f"={MKTCAP}/C9",Ff(),NUM,cmt=f"CapIQ get_quote: mktcap {MKTCAP}bn / price -> 隱含稀釋股數;Daloopa/CapIQ 未直接提供股數")
put("B11","FY26E 營收 (CapIQ est)",Ff(),al=LEF); put("C11",REV_FY26E,Fl(),BN,fl=SRC,cmt="CapIQ get_estimates VRT FY2026 revenue = 10.9 USD bn")
put("B12","FY27E 營收 (CapIQ est)",Ff(),al=LEF); put("C12",REV_FY27E,Fl(),BN,fl=SRC,cmt="CapIQ get_estimates VRT FY2027 revenue = 13.5 USD bn")
put("B13","FY25 營業利益率 (Daloopa)",Ff(),al=LEF); put("C13",f"={OPINC_FY25}/C8",Ff(),PCT,cmt="Daloopa Operating Profit 1.5 / Revenue 8.0 = 18.75%")

# -- scenario assumption blocks (years in D..H) --
hdr("B15:H15","情境假設 SCENARIO ASSUMPTIONS（營收成長 / EBIT 利潤率）[藍=ASSUMPTION]")
def yearhdr(rrow):
    put(f"B{rrow}","假設項目",Ff(9,True),al=LEF,fl=SUB)
    for i,y in enumerate(YRS[1:]):
        put(f"{get_column_letter(4+i)}{rrow}",y,Ff(9,True),al=CEN,fl=SUB)
# FY26/FY27 成長率錨定 CapIQ estimates;FY28-30 為遞減假設
base_g=[G_FY26,G_FY27,.15,.11,.08]; base_m=[.195,.205,.215,.22,.225]
bear_g=[G_FY26-.06,G_FY27-.06,.10,.07,.05]; bear_m=[.185,.19,.195,.20,.20]
bull_g=[G_FY26+.05,G_FY27+.05,.20,.15,.11]; bull_m=[.20,.215,.23,.24,.245]
def arow(rrow,label,vals,cmt=None):
    put(f"B{rrow}",label,Ff(),al=LEF)
    for i,v in enumerate(vals):
        c=put(f"{get_column_letter(4+i)}{rrow}",round(v,4),Fi(),PCT,fl=INP)
        if cmt and i==0: c.comment=Comment(cmt,"DCF skill")
put("B16","熊 BEAR CASE",Ff(10,True),al=LEF,fl=SUB); ws.merge_cells("B16:H16")
yearhdr(17)
arow(18,"  營收成長 %",bear_g,"[ASSUMPTION] Bear: 雲端資本支出消化、競爭加劇;FY26/27 在 CapIQ 隱含成長 -6pt")
arow(19,"  EBIT 利潤率 %",bear_m,"[ASSUMPTION] Bear margin")
put("B20","基 BASE CASE",Ff(10,True),al=LEF,fl=SUB); ws.merge_cells("B20:H20")
yearhdr(21)
arow(22,"  營收成長 %",base_g,f"FY26 +{G_FY26:.1%}、FY27 +{G_FY27:.1%} = CapIQ get_estimates 隱含成長(8.0->10.9->13.5);FY28-30 為 [ASSUMPTION] 遞減")
arow(23,"  EBIT 利潤率 %",base_m,"[ASSUMPTION] EBIT 由 ~18.75%(Daloopa FY25 實績)擴張至 ~22.5%(營運槓桿+液冷組合);參考 CapIQ FY26 EBITDA margin 22%")
put("B24","牛 BULL CASE",Ff(10,True),al=LEF,fl=SUB); ws.merge_cells("B24:H24")
yearhdr(25)
arow(26,"  營收成長 %",bull_g,"[ASSUMPTION] Bull: 液冷滲透加速、份額擴張;FY26/27 在 CapIQ 隱含成長 +5pt")
arow(27,"  EBIT 利潤率 %",bull_m,"[ASSUMPTION] Bull margin")
# Selected (consolidation) via CHOOSE
put("B28","選定情境 SELECTED（=CHOOSE 由選擇器）",Ff(10,True),al=LEF,fl=OUTC); ws.merge_cells("B28:H28")
for i in range(5):
    col=get_column_letter(4+i)
    put(f"{col}29",f"=CHOOSE($C$4,{col}18,{col}22,{col}26)",Ff(),PCT,fl=OUTC)
    put(f"{col}30",f"=CHOOSE($C$4,{col}19,{col}23,{col}27)",Ff(),PCT,fl=OUTC)
put("B29","  選定營收成長 %",Ff(),al=LEF,fl=OUTC)
put("B30","  選定 EBIT 利潤率 %",Ff(),al=LEF,fl=OUTC)

# -- global assumptions --
hdr("B32:H32","全域假設 GLOBAL ASSUMPTIONS（與情境無關）[藍=ASSUMPTION]")
put("B33","稅率 Tax Rate",Ff(),al=LEF); put("C33",0.21,Fi(),PCT,fl=INP,cmt="[ASSUMPTION] MCP 未提供;US 法定+州約 21%")
put("B34","稅率 (供 WACC 連結)",Ff(),al=LEF); put("C34","=C33",Ff(),PCT)
put("B35","D&A % 營收",Ff(),al=LEF); put("C35",0.026,Fi(),PCT,fl=INP,cmt="[ASSUMPTION] Daloopa 未提供;VRT 歷史 D&A 約營收 2.6%")
put("B36","Capex % 營收",Ff(),al=LEF); put("C36",0.023,Fi(),PCT,fl=INP,cmt="[ASSUMPTION] Daloopa 未提供;資本輕,capex 約營收 2.3%")
put("B37","ΔNWC % 增額營收",Ff(),al=LEF); put("C37",0.08,Fi(),PCT,fl=INP,cmt="[ASSUMPTION] Daloopa 未提供;營運資金約增額營收 8%(backlog $15bn 履約需備料)")
put("B38","終值成長 g",Ff(),al=LEF); put("C38",0.03,Fi(),PCT,fl=INP,cmt="[ASSUMPTION] 約長期 GDP;限制 g<WACC")
put("B39","WACC (連結 WACC 分頁)",Ff(10,True),al=LEF,fl=SUB); put("C39","=WACC!$C$28",Fl(10,True),PCT,fl=SUB)

# -- projections C..H (C=FY25A) --
hdr("B41:H41","財務預測 PROJECTIONS (US$ bn)")
put("B42","項目 \\ 年度",Ff(9,True),al=LEF,fl=SUB)
for i,y in enumerate(YRS):
    put(f"{get_column_letter(3+i)}42",y,Ff(9,True),al=CEN,fl=SUB)
# Revenue (C linked to source block C8 = Daloopa)
put("B43","營收 Revenue",Ff(10,True),al=LEF)
put("C43","=C8",Fl(10,True),BN,cmt="連結 MCP 來源 C8 = Daloopa FY25 營收 8.0")
for i in range(5):
    col=get_column_letter(4+i); prev=get_column_letter(3+i)
    put(f"{col}43",f"={prev}43*(1+{col}29)",Ff(10,True),BN)
put("B44","  營收成長 %",Ff(),al=LEF)
for i in range(5):
    col=get_column_letter(4+i); prev=get_column_letter(3+i)
    put(f"{col}44",f"={col}43/{prev}43-1",Ff(),PCT)
put("B45","  cross-check FY26/27 vs CapIQ",Ff(8),al=LEF)
put("D45","=D43-C11",Ff(8),BN,cmt="模型 FY26 營收 − CapIQ FY26E 10.9,基準應≈0")
put("E45","=E43-C12",Ff(8),BN,cmt="模型 FY27 營收 − CapIQ FY27E 13.5,基準應≈0")
put("B46","  EBIT 利潤率 %",Ff(),al=LEF)
put("C46","=C13",Fl(),PCT,cmt="連結 C13 = Daloopa FY25 營業利益率 18.75%")
for i in range(5):
    col=get_column_letter(4+i); put(f"{col}46",f"={col}30",Ff(),PCT)
put("B47","EBIT",Ff(10,True),al=LEF)
for i in range(6):
    col=get_column_letter(3+i); put(f"{col}47",f"={col}43*{col}46",Ff(10,True),BN)
put("B48","  減：稅 Taxes",Ff(),al=LEF)
for i in range(6):
    col=get_column_letter(3+i); put(f"{col}48",f"={col}47*$C$33",Ff(),BN)
put("B49","NOPAT",Ff(10,True),al=LEF)
for i in range(6):
    col=get_column_letter(3+i); put(f"{col}49",f"={col}47-{col}48",Ff(10,True),BN)
put("B50","  加：D&A",Ff(),al=LEF)
for i in range(6):
    col=get_column_letter(3+i); put(f"{col}50",f"={col}43*$C$35",Ff(),BN)
put("B51","  減：Capex",Ff(),al=LEF)
for i in range(6):
    col=get_column_letter(3+i); put(f"{col}51",f"={col}43*$C$36",Ff(),BN)
put("B52","  減：ΔNWC",Ff(),al=LEF)
put("C52",0,Ff(),BN)
for i in range(5):
    col=get_column_letter(4+i); prev=get_column_letter(3+i)
    put(f"{col}52",f"=({col}43-{prev}43)*$C$37",Ff(),BN)
put("B53","無槓桿自由現金流 FCFF",Ff(10,True),al=LEF,fl=SUB)
for i in range(6):
    col=get_column_letter(3+i); put(f"{col}53",f"={col}49+{col}50-{col}51-{col}52",Ff(10,True),BN,fl=SUB)

# -- discounting (forecast D..H) --
hdr("B55:H55","折現與終值 DISCOUNTING & TERMINAL VALUE")
put("B56","期數 (期中) Period",Ff(),al=LEF)
periods=[0.5,1.5,2.5,3.5,4.5]
for i,p in enumerate(periods):
    put(f"{get_column_letter(4+i)}56",p,Fi(),"0.0",fl=INP,cmt="期中折現慣例 mid-year convention" if i==0 else None)
put("B57","折現因子 Discount Factor",Ff(),al=LEF)
for i in range(5):
    col=get_column_letter(4+i); put(f"{col}57",f"=1/(1+$C$39)^{col}56",Ff(),"0.000")
put("B58","FCFF 現值 PV",Ff(10,True),al=LEF)
for i in range(5):
    col=get_column_letter(4+i); put(f"{col}58",f"={col}53*{col}57",Ff(10,True),BN)

# -- valuation summary (col C) --
hdr("B60:H60","估值總結 VALUATION SUMMARY (US$ bn)")
put("B61","終年 FCFF × (1+g) = 終值 FCFF",Ff(),al=LEF); put("C61","=H53*(1+$C$38)",Ff(),BN)
put("B62","終值 TV (永續成長法)",Ff(),al=LEF); put("C62","=C61/($C$39-$C$38)",Ff(),BN)
put("B63","終值現值 PV of TV",Ff(),al=LEF); put("C63","=C62/(1+$C$39)^H56",Ff(),BN)
put("B64","顯性期 PV 合計 ΣPV FCFF",Ff(),al=LEF); put("C64","=SUM(D58:H58)",Ff(),BN)
put("B65","企業價值 Enterprise Value",Ff(10,True),al=LEF,fl=SUB); put("C65","=C64+C63",Ff(10,True),BN,fl=SUB)
put("B66","  減：淨負債 Net Debt (連結 WACC)",Ff(),al=LEF); put("C66","=WACC!$C$21",Fl(),BN)
put("B67","股權價值 Equity Value",Ff(10,True),al=LEF,fl=SUB); put("C67","=C65-C66",Ff(10,True),BN,fl=SUB)
put("B68","  稀釋股數 (bn)",Ff(),al=LEF); put("C68","=C10",Ff(),NUM)
put("B69","隱含每股價值 Implied Price",Ff(11,True),al=LEF,fl=OUTC); put("C69","=C67/C68",Ff(11,True),SH2,fl=OUTC)
put("B70","現價 Current Price (CapIQ)",Ff(),al=LEF); put("C70","=C9",Fl(),SH2)
put("B71","隱含上漲/(下跌)",Ff(11,True),al=LEF,fl=OUTC); put("C71","=C69/C70-1",Ff(11,True),PCTS,fl=OUTC)
# checks
put("B73","檢查：終值占 EV %",Ff(),al=LEF); put("C73","=C63/C65",Ff(),PCT,cmt="健康區間 50-70%;偏高代表高度仰賴終值")
put("B74","檢查：隱含 EV/EBITDA (FY26E)",Ff(),al=LEF); put("C74","=C65/(D47+D50)",Ff(),MULT,cmt="DCF 隱含 EV / FY26E EBITDA(EBIT+D&A)")
put("B75","對照：CapIQ comps VRT EV/EBITDA",Ff(),al=LEF); put("C75",51,Fi(),MULT,fl=SRC,cmt="CapIQ get_comps VRT ev_ebitda = 51x(市場現價隱含)")

# -- football field: comps-implied EV vs DCF (cross-check) --
hdr("B77:H77","足球場 FOOTBALL FIELD：DCF vs CapIQ comps 倍數隱含估值")
put("B78","方法",Ff(9,True),al=LEF,fl=SUB); put("C78","倍數",Ff(9,True),al=CEN,fl=SUB)
put("D78","基準",Ff(9,True),al=CEN,fl=SUB); put("E78","隱含 EV",Ff(9,True),al=CEN,fl=SUB)
put("F78","隱含股權",Ff(9,True),al=CEN,fl=SUB); put("G78","隱含股價",Ff(9,True),al=CEN,fl=SUB); put("H78","vs 現價",Ff(9,True),al=CEN,fl=SUB)
# FY26E EBITDA base = D47+D50
def ff_row(rr,label,mult,mult_src_cmt):
    put(f"B{rr}",label,Ff(),al=LEF)
    put(f"C{rr}",mult,Fi(),MULT,fl=SRC,cmt=mult_src_cmt)
    put(f"D{rr}","=D47+D50",Ff(),BN,cmt="FY26E EBITDA = EBIT + D&A")
    put(f"E{rr}",f"=C{rr}*D{rr}",Ff(),BN)
    put(f"F{rr}",f"=E{rr}-WACC!$C$21",Ff(),BN)
    put(f"G{rr}",f"=F{rr}/$C$10",Ff(),SH2)
    put(f"H{rr}",f"=G{rr}/$C$9-1",Ff(),PCTS)
ff_row(79,"VRT 自身 EV/EBITDA",51,"CapIQ get_comps VRT ev_ebitda = 51x")
ff_row(80,"nVent (NVT) EV/EBITDA",32,"CapIQ get_comps NVT ev_ebitda = 32x")
ff_row(81,"Schneider (SU.PA) EV/EBITDA",21,"CapIQ get_comps SU.PA ev_ebitda = 21x")
put("B82","DCF 隱含(基準情境)",Ff(10,True),al=LEF,fl=OUTC)
put("C82","=C74",Ff(10,True),MULT,fl=OUTC,cmt="本 DCF 隱含的 EV/EBITDA")
put("E82","=C65",Ff(10,True),BN,fl=OUTC); put("F82","=C67",Ff(10,True),BN,fl=OUTC)
put("G82","=C69",Ff(10,True),SH2,fl=OUTC); put("H82","=C71",Ff(10,True),PCTS,fl=OUTC)
put("B83","註：comps 倍數=現價市場隱含;DCF 倍數遠低 → 市場用遠低於本模型 WACC 的折現率、或更長成長跑道定價。",Ff(8),al=LEF); ws.merge_cells("B83:H83")

# -- sensitivity: WACC x terminal growth (5x5, base-centered) --
hdr("B85:H85","敏感度分析 SENSITIVITY：隱含每股價值 — WACC × 終值成長 g")
put("B86","WACC↓ / g→",Ff(9,True),al=CEN,fl=SUB)
g_off=[-0.01,-0.005,0,0.005,0.01]
for i,off in enumerate(g_off):
    col=get_column_letter(4+i)
    put(f"{col}86",f"=$C$38+({off})",Fi(9,True),PCT,al=CEN,fl=SUB)
w_off=[-0.01,-0.005,0,0.005,0.01]
for r,off in enumerate(w_off):
    rr=87+r
    put(f"C{rr}",f"=$C$39+({off})",Fi(9,True),PCT,al=CEN,fl=SUB)
    for cidx in range(5):
        col=get_column_letter(4+cidx)
        f=(f"=(SUMPRODUCT($D$53:$H$53,1/(1+$C{rr})^$D$56:$H$56)"
           f"+($H$53*(1+{col}$86)/($C{rr}-{col}$86))/(1+$C{rr})^$H$56"
           f"-WACC!$C$21)/$C$10")
        center = (r==2 and cidx==2)
        put(f"{col}{rr}",f,Ff(10,center),SH2,al=CEN,fl=(CTR if center else (INP if (r+cidx)%2 else WHITE)))
put("B92",("註:① 中心格(粗體高亮)=基準情境,應等於上方隱含每股價值 C69。② 表格採目前情境(C4)之 FCFF 流;切換情境後同步更新。"
           "③ g<WACC 為必要條件。④ 歷史財務全部來自 Daloopa MCP;FY26/27 成長率對齊 CapIQ estimates;其餘藍字為 [ASSUMPTION]。"),
    Ff(8,False),al=LEF); ws.merge_cells("B92:H94")

# widths
ws.column_dimensions["A"].width=2; ws.column_dimensions["B"].width=32
for c in "CDEFGH": ws.column_dimensions[c].width=12

if "Sheet" in wb.sheetnames: wb.remove(wb["Sheet"])
outpath=os.path.join(OUT,"VRT_DCF_Model.xlsx")
wb.save(outpath)
print("SAVED", outpath, " sheets:", wb.sheetnames)
