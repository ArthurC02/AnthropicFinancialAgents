# -*- coding: utf-8 -*-
"""AI 資料中心散熱 — comps 工作簿（從 MCP 拉到的 JSON 建表，統計基準 + 離群值標記）。

資料來源:capiq MCP get_comps(sector="AI Data Center Cooling") 的快照
  docs/outputs/mcpBased/market-researcher/mcp_pulls/01_capiq_get_comps.json
n/a 一律保留為文字 "n/a"(不臆造),且不納入統計。Mock / dev only。
"""
import json
import os
import statistics as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
PULLS = os.path.join(ROOT, "mcp_pulls")
OUT = os.path.join(ROOT, "out")
os.makedirs(OUT, exist_ok=True)

NAVY = "0B1F3A"; BLUE = "1F4E79"; ACCENT = "2E9BD6"; LGREY = "ECEFF3"
TEAL = "14A38C"; AMBER = "E89A1C"; RED = "C03A2B"; WHITE = "FFFFFF"; MIDG = "5A5A5A"
CJK = "Microsoft JhengHei"

thin = Side(style="thin", color="D0D5DD")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
def F(sz=10, b=False, color="222A35"): return Font(name=CJK, size=sz, bold=b, color=color)
def fill(c): return PatternFill("solid", fgColor=c)
CEN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEF = Alignment(horizontal="left", vertical="center", wrap_text=True)


def num_or_none(v):
    """MCP 把缺漏值回成字串 'n/a' / ''。轉成 None 代表缺漏,其餘保留數值。"""
    if isinstance(v, (int, float)):
        return v
    if v is None:
        return None
    s = str(v).strip().lower()
    if s in ("", "n/a", "na", "nan"):
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


# ---- 讀 MCP 快照 ----
with open(os.path.join(PULLS, "01_capiq_get_comps.json"), encoding="utf-8") as f:
    pull = json.load(f)
res = pull["result"]
comps = res["comps"]
as_of = res.get("as_of", "")
caveat = res.get("caveat", "")
sector = res.get("sector", "")
COUNT = res.get("count", len(comps))
assert COUNT == 14, f"comps count 應為 14,實得 {COUNT}"
assert len(comps) == 14, f"comps rows 應為 14,實得 {len(comps)}"

wb = Workbook()

# ============================================================= Comps sheet
ws = wb.active; ws.title = "Comps"
ws.sheet_view.showGridLines = False

ws.merge_cells("A1:M1")
ws["A1"] = "AI 資料中心散熱 — 同業比較 (Comps)　【MCP 即時拉取 · mock/dev】"
ws["A1"].font = F(16, True, WHITE); ws["A1"].fill = fill(NAVY)
ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[1].height = 30
ws.merge_cells("A2:M2")
ws["A2"] = (f"來源:capiq MCP get_comps(sector='{sector}') · as_of {as_of} · "
            f"{COUNT} 家 · 數字為即時變動之估計值,僅供方向參考 · 代碼校正:Auras=3324、AVC=3017")
ws["A2"].font = F(9, False, "FFFFFF"); ws["A2"].fill = fill(BLUE)
ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[2].height = 18

headers = ["公司", "代碼", "上市地", "市值\n(USD bn)", "營收\nYoY", "營益率", "Fwd\nP/E",
           "EV/\nEBITDA", "EV/\nRevenue", "1年\n報酬", "散熱\n純度", "純標的?", "備註"]
hrow = 4
for ci, h in enumerate(headers, start=1):
    c = ws.cell(hrow, ci, h); c.font = F(10, True, WHITE); c.fill = fill(NAVY)
    c.alignment = CEN; c.border = border

# 從 MCP 列轉成表格列;數值欄用 None 代表 n/a(原樣保留,不臆造)
FIELDS = ["company", "ticker", "listing", "mktcap_usd_bn", "rev_yoy", "op_margin",
          "fwd_pe", "ev_ebitda", "ev_rev", "ret_1y", "thermal_purity", "pure_play", "note"]
NUMERIC = {4, 5, 6, 7, 8, 9, 10}  # 1-based 欄位:需轉數值
pct_cols = {5, 6, 10}

data = []
na_count = 0
for row in comps:
    out = []
    for ci, fld in enumerate(FIELDS, start=1):
        v = row.get(fld)
        if ci in NUMERIC:
            n = num_or_none(v)
            if n is None:
                na_count += 1
            out.append(n)
        else:
            out.append(v)
    data.append(out)

r = hrow + 1
for row in data:
    for ci, val in enumerate(row, start=1):
        show = val if val is not None else "n/a"
        c = ws.cell(r, ci, show)
        c.border = border
        c.fill = fill(LGREY if (r - hrow) % 2 == 0 else WHITE)
        if ci == 1:
            c.font = F(10, True); c.alignment = LEF
        elif ci == 13:
            c.font = F(9, False, MIDG); c.alignment = LEF
        else:
            c.alignment = CEN
            if val is None:
                c.font = F(9, False, MIDG)  # n/a 灰字
            elif ci == 10 and isinstance(val, (int, float)):
                c.font = F(10, True, RED if val < 0 else TEAL)
            else:
                c.font = F(10)
            if isinstance(val, (int, float)):
                if ci in pct_cols:
                    c.number_format = "+0.0%;-0.0%"
                elif ci == 4:
                    c.number_format = "$#,##0.0"
                elif ci in (7, 8, 9):
                    c.number_format = '0.0"x"'
    r += 1

# ---- 統計基準(只納入有數值者) ----
stat_start = r + 1
ws.cell(stat_start, 1, "統計基準 (Statistical Benchmarking)　— 僅納入有數值之欄位,n/a 不計").font = F(12, True, NAVY)
labels = ["平均 Mean", "中位 Median", "上四分位 Q3", "下四分位 Q1", "有效樣本 n"]
def colvals(ci):
    return [row[ci-1] for row in data if isinstance(row[ci-1], (int, float))]
def q(vals, p):
    vals = sorted(vals)
    if not vals: return None
    k = (len(vals)-1)*p; fl = int(k); cdiff = k-fl
    if fl+1 < len(vals): return vals[fl] + (vals[fl+1]-vals[fl])*cdiff
    return vals[fl]
stat_cols = [4, 5, 6, 7, 8, 9, 10]
hr = stat_start + 1
ws.cell(hr, 3, "指標").font = F(10, True, WHITE); ws.cell(hr, 3).fill = fill(BLUE)
ws.cell(hr, 3).alignment = CEN; ws.cell(hr, 3).border = border
for ci in stat_cols:
    c = ws.cell(hr, ci, headers[ci-1].replace("\n", " ")); c.font = F(9, True, WHITE)
    c.fill = fill(BLUE); c.alignment = CEN; c.border = border
for li, lab in enumerate(labels):
    rr = hr + 1 + li
    lc = ws.cell(rr, 3, lab); lc.font = F(10, True); lc.alignment = LEF
    lc.border = border; lc.fill = fill(LGREY if li % 2 == 0 else WHITE)
    for ci in stat_cols:
        vals = colvals(ci)
        if lab.startswith("平均"): v = st.mean(vals) if vals else None
        elif lab.startswith("中位"): v = st.median(vals) if vals else None
        elif lab.startswith("上四"): v = q(vals, 0.75)
        elif lab.startswith("下四"): v = q(vals, 0.25)
        else: v = len(vals)
        c = ws.cell(rr, ci, v if v is not None else "n/a")
        c.alignment = CEN; c.font = F(10); c.border = border
        c.fill = fill(LGREY if li % 2 == 0 else WHITE)
        if lab.startswith("有效"):
            c.number_format = "0"
        elif isinstance(v, (int, float)):
            if ci in pct_cols: c.number_format = "+0.0%;-0.0%"
            elif ci == 4: c.number_format = "$#,##0.0"
            else: c.number_format = '0.0"x"'

# ---- n/a 缺漏盤點 ----
nas = hr + len(labels) + 3
ws.cell(nas, 1, "缺漏盤點 (Missing-data Audit) — 共 %d 個 n/a 數值格" % na_count).font = F(12, True, NAVY)
miss_by_co = []
for row in comps:
    miss = [headers[ci-1].replace("\n", "") for ci, fld in enumerate(FIELDS, start=1)
            if ci in NUMERIC and num_or_none(row.get(fld)) is None]
    if miss:
        miss_by_co.append((f"{row['company']} ({row['ticker']})", "、".join(miss)))
nrr = nas + 1
for (co, miss) in miss_by_co:
    ws.cell(nrr, 1, "▍").font = F(11, True, AMBER)
    cc = ws.cell(nrr, 2, co); cc.font = F(10, True, "222A35")
    ws.merge_cells(start_row=nrr, start_column=4, end_row=nrr, end_column=13)
    dc = ws.cell(nrr, 4, "缺:" + miss); dc.font = F(9.5, False, MIDG); dc.alignment = LEF
    nrr += 1
ws.cell(nrr, 2, "處理原則").font = F(10, True, NAVY)
ws.merge_cells(start_row=nrr, start_column=4, end_row=nrr, end_column=13)
ws.cell(nrr, 4, "免費資料源未提供之倍數一律留 n/a,不臆造、不內插;統計僅納入有數值者。").font = F(9.5, False, MIDG)
ws.cell(nrr, 4).alignment = LEF

# ---- 離群值標記(數字來自 MCP) ----
ofs = nrr + 2
ws.cell(ofs, 1, "離群值標記 (Outliers) — 數字皆來自 MCP 拉取").font = F(12, True, NAVY)
# 動態抓出可報告的極值
def get(t, fld):
    for row in comps:
        if row["ticker"] == t:
            return num_or_none(row.get(fld))
    return None
outliers = [
    (f"Envicool — Fwd P/E ~{get('002837.SZ','fwd_pe'):.0f}x", "全表最高,AI 敘事驅動 + Q1 淨利 −82%,估值風險最高", RED),
    (f"Delta 台達電 — 1年報酬 +{get('2308.TW','ret_1y')*100:.0f}%", "再評價最劇烈,注意追高風險", AMBER),
    (f"Supermicro — 1年報酬 {get('SMCI','ret_1y')*100:.0f}%", "全表唯一負報酬;治理history + 低毛利 OEM 折價", RED),
    (f"Vertiv — EV/EBITDA ~{get('VRT','ev_ebitda'):.0f}x", "散熱純度溢價最明顯(最乾淨曝險)", TEAL),
    (f"Lotes — 營益率 {get('3533.TW','op_margin')*100:.0f}%", "高毛利元件 vs. 低毛利 ODM/OEM,EV/Rev 不可直接互比", BLUE),
    ("Parker / Asetek", "散熱曝險過小(PH ~1%)或休眠(ASTK),倍數非散熱溢價 — 純散熱統計時宜剔除", MIDG),
]
orr = ofs + 1
for (t, d, col) in outliers:
    ws.cell(orr, 1, "▍").font = F(11, True, col)
    cc = ws.cell(orr, 2, t); cc.font = F(10, True, col)
    ws.merge_cells(start_row=orr, start_column=4, end_row=orr, end_column=13)
    dc = ws.cell(orr, 4, d); dc.font = F(9.5, False, MIDG); dc.alignment = LEF
    orr += 1

widths = [16, 11, 7, 10, 8, 8, 8, 9, 9, 9, 9, 8, 40]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A5"

# ============================================================= Notes sheet
ns = wb.create_sheet("Notes & Sources")
ns.sheet_view.showGridLines = False
ns.column_dimensions["A"].width = 4
ns.column_dimensions["B"].width = 120
ns.merge_cells("A1:B1")
ns["A1"] = "方法論、注意事項與來源（MCP 版）"; ns["A1"].font = F(14, True, WHITE); ns["A1"].fill = fill(NAVY)
ns.row_dimensions[1].height = 26
notes = [
    ("H", "資料來源（本工作簿全部數字）"),
    ("",  f"capiq MCP(mock) get_comps(sector='{sector}') 的回傳快照,共 {COUNT} 家,as_of {as_of}。"),
    ("",  "原始 JSON:mcp_pulls/01_capiq_get_comps.json(harness 真實呼叫 stdio MCP 後留存)。"),
    ("",  f"MCP caveat 原文:{caveat}"),
    ("H", "缺漏(n/a)處理"),
    ("",  f"本次 14 家 × 7 個數值欄位中,共 {na_count} 格為 n/a(MCP 回字串 'n/a',代表免費源未提供)。"),
    ("",  "處理原則:一律保留 n/a,不臆造、不內插、不跨公司借值;統計(平均/中位/四分位)僅納入有數值者。"),
    ("",  "EV/EBITDA 缺最多(Delta、Auras、AVC、Envicool 4 家);台股/陸股中小型市值倍數本就難從免費源取得。"),
    ("H", "方法論注意"),
    ("",  "EV/EBITDA 與 EV/Revenue 不可跨『元件 vs. 系統/ODM』直接比較:ODM/OEM(Wiwynn、SMCI)為低毛利轉手收入,EV/Rev 天生偏低。"),
    ("",  "純散熱基準建議剔除 Parker(散熱僅約 1%)、Asetek(業務休眠)再計。"),
    ("H", "關鍵警示(來自 MCP note 欄)"),
    ("",  "Parker (PH):資料中心散熱僅約 1% 營收 — 倍數反映航太/工業品質,非散熱溢價。"),
    ("",  "Asetek (ASTK):DC 液冷業務休眠 — 不應視為 AI 散熱標的(op_margin/fwd_pe/ev_ebitda 皆 n/a)。"),
    ("",  "Supermicro (SMCI):伺服器 OEM,散熱為綁定功能;治理爭議已解。"),
    ("",  "Envicool (002837):最接近純散熱,但 Q1 淨利年減約 82%,估值(~77x fwd P/E)高度由 AI 敘事驅動。"),
    ("",  "代碼校正:Auras 奇鋐 = 3324(TPEx)、AVC = 3017(TWSE)— 中文名易混淆,下單前務必核對。"),
    ("H", "免責聲明"),
    ("",  "本工作簿之數字全部來自 mock(假)MCP server,僅供內部研究/示範,不構成投資建議,不對外發布。"),
]
rr = 3
for (tag, text) in notes:
    if tag == "H":
        c = ns.cell(rr, 2, text); c.font = F(12, True, BLUE)
        rr += 1
    else:
        c = ns.cell(rr, 2, "•  " + text); c.font = F(10, False, "222A35"); c.alignment = LEF
        ns.row_dimensions[rr].height = 16
        rr += 1

dest = os.path.join(OUT, "AI_DataCenter_Cooling_Comps.xlsx")
wb.save(dest)
print(f"SAVED {dest}")
print(f"comps rows={len(comps)}  count(MCP)={COUNT}  n/a numeric cells={na_count}")
