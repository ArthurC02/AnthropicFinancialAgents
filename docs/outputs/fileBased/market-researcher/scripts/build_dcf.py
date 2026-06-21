# -*- coding: utf-8 -*-
"""Vertiv (VRT) DCF — live-formula workbook (headless)."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

NAVY="0B1F3A"; BLUE="1F4E79"; ACCENT="2E9BD6"; LGREY="ECEFF3"; INPUTF="FFF6E6"
TEAL="14A38C"; AMBER="E89A1C"; RED="C03A2B"; WHITE="FFFFFF"; MIDG="5A5A5A"; DARK="222A35"
CJK="Microsoft JhengHei"
thin=Side(style="thin",color="D0D5DD"); border=Border(left=thin,right=thin,top=thin,bottom=thin)
def F(sz=10,b=False,color=DARK): return Font(name=CJK,size=sz,bold=b,color=color)
def fill(c): return PatternFill("solid",fgColor=c)
CEN=Alignment(horizontal="center",vertical="center",wrap_text=True)
RGT=Alignment(horizontal="right",vertical="center")
LEF=Alignment(horizontal="left",vertical="center",wrap_text=True)
PCT="0.0%"; PCTS="+0.0%;-0.0%"; USD='$#,##0.00'; BN='$#,##0.0'; MULT='0.0"x"'

wb=Workbook(); ws=wb.active; ws.title="DCF"; ws.sheet_view.showGridLines=False
YRS=["FY25A","FY26E","FY27E","FY28E","FY29E","FY30E"]  # D..I

def put(cell,val,font=None,fmt=None,al=None,bd=False,fl=None):
    c=ws[cell]; c.value=val
    c.font=font or F()
    if fmt: c.number_format=fmt
    if al: c.alignment=al
    if bd: c.border=border
    if fl: c.fill=fill(fl)
    return c

# Title
ws.merge_cells("B1:I1"); put("B1","Vertiv Holdings (VRT) — DCF 估值模型",F(15,True,WHITE),al=Alignment(horizontal="left",vertical="center"),fl=NAVY)
ws.row_dimensions[1].height=26
ws.merge_cells("B2:I2"); put("B2","單位：US$ bn（除每股/倍數）· 活公式，開啟即重算 · 資料時點 2026/6 · 僅供研究參考",F(9,False,WHITE),al=Alignment(horizontal="left",vertical="center"),fl=BLUE)

# ---- Inputs panel (K:L) ----
put("K1","輸入 INPUTS",F(11,True,WHITE),fl=NAVY); ws.merge_cells("K1:L1")
inputs=[("情境 (1熊 2基 3牛)",2,"0"),("WACC",0.10,PCT),("終值成長 g",0.035,PCT),
        ("終值 EV/EBITDA",22.0,MULT),("稅率",0.21,PCT),("D&A %營收",0.026,PCT),
        ("Capex %營收",0.023,PCT),("ΔNWC %增額營收",0.08,PCT),("淨負債 $bn",2.0,BN),
        ("股數 (bn)",0.382,"0.000"),("現價 $",320.0,USD)]
r=2
for (lab,val,fmt) in inputs:
    put(f"K{r}",lab,F(9.5,False,DARK),al=LEF,bd=True,fl=LGREY)
    put(f"L{r}",val,F(10,True,BLUE),fmt=fmt,al=CEN,bd=True,fl=INPUTF)
    r+=1
# named cells map
SCN="$L$2"; WACC="$L$3"; G="$L$4"; EXM="$L$5"; TAX="$L$6"; DAr="$L$7"; CXr="$L$8"; NWCr="$L$9"; ND="$L$10"; SH="$L$11"; PX="$L$12"

# ---- Projection header ----
hr=7
put(f"B{hr}","項目 \\ 年度",F(10,True,WHITE),al=LEF,bd=True,fl=NAVY)
for i,y in enumerate(YRS):
    col=get_column_letter(4+i)
    put(f"{col}{hr}",y,F(10,True,WHITE),al=CEN,bd=True,fl=NAVY)

def prow(rr,label,fmt,bold=False,fillc=None,topband=False):
    put(f"B{rr}",label,F(10,bold,DARK),al=LEF,bd=True,fl=fillc or (LGREY if rr%2 else WHITE))
    for i in range(6):
        col=get_column_letter(4+i)
        ws[f"{col}{rr}"].number_format=fmt; ws[f"{col}{rr}"].border=border
        ws[f"{col}{rr}"].alignment=RGT; ws[f"{col}{rr}"].font=F(10,bold)
        ws[f"{col}{rr}"].fill=fill(fillc or (LGREY if rr%2 else WHITE))

# rows
R_REV,R_G,R_M,R_EBIT,R_TAX,R_NOP,R_DA,R_CX,R_NWC,R_FCFF,R_T,R_DF,R_PV,R_EBITDA = 8,9,10,11,12,13,14,15,16,17,18,19,20,22
prow(R_REV,"營收 Revenue",BN,True)
prow(R_G,"  營收成長 %",PCTS)
prow(R_M,"  EBIT 利潤率 %",PCT)
prow(R_EBIT,"EBIT (營業利益)",BN,True)
prow(R_TAX,"  減：稅 (NOPAT 稅)",BN)
prow(R_NOP,"NOPAT",BN,True)
prow(R_DA,"  加：折舊攤銷 D&A",BN)
prow(R_CX,"  減：資本支出 Capex",BN)
prow(R_NWC,"  減：營運資金增加 ΔNWC",BN)
prow(R_FCFF,"無槓桿自由現金流 FCFF",BN,True,fillc="E3F1F8")
prow(R_T,"  期數 t",'0')
prow(R_DF,"  折現因子",'0.000')
prow(R_PV,"FCFF 現值 PV",BN,True)
prow(R_EBITDA,"備忘：EBITDA",BN)

# FY25A actuals (col D)
put("D8",10.84,F(10,True),fmt=BN,al=RGT,bd=True); ws["D8"].fill=fill(WHITE)
put("D9",0.28,F(10),fmt=PCTS,al=RGT,bd=True)
put("D10",0.190,F(10),fmt=PCT,al=RGT,bd=True); ws["D10"].fill=fill(WHITE)
ws["D11"]=f"=D8*D10"; ws["D14"]=f"=D8*{DAr}"; ws["D22"]=f"=D11+D14"
for cc in ["D9","D11","D14","D22"]:
    ws[cc].fill=fill(LGREY if int(cc[1:])%2 else WHITE)

# scenario tables (rows 27-34)
put("B26","情境假設 SCENARIO INPUTS（可編輯）",F(11,True,NAVY))
put("B27","營收成長 (FY26→FY30)",F(9.5,True,DARK),al=LEF,bd=True,fl=BLUE); ws["B27"].font=F(9.5,True,WHITE)
for i,y in enumerate(YRS[1:]):
    put(f"{get_column_letter(5+i)}27",y,F(9.5,True,WHITE),al=CEN,bd=True,fl=BLUE)
growth={"熊 Bear":[.22,.15,.10,.07,.05],"基 Base":[.27,.22,.16,.12,.09],"牛 Bull":[.32,.28,.22,.17,.12]}
margin={"熊 Bear":[.195,.20,.205,.21,.21],"基 Base":[.20,.21,.22,.225,.23],"牛 Bull":[.205,.22,.235,.245,.25]}
rr=28
for nm,vals in growth.items():
    put(f"B{rr}",nm,F(9.5,True),al=LEF,bd=True,fl=INPUTF)
    for i,v in enumerate(vals): put(f"{get_column_letter(5+i)}{rr}",v,F(9.5,False,BLUE),fmt=PCTS,al=CEN,bd=True,fl=INPUTF)
    rr+=1
put("B31","EBIT 利潤率 (FY26→FY30)",F(9.5,True,WHITE),al=LEF,bd=True,fl=BLUE)
for i,y in enumerate(YRS[1:]):
    put(f"{get_column_letter(5+i)}31",y,F(9.5,True,WHITE),al=CEN,bd=True,fl=BLUE)
rr=32
for nm,vals in margin.items():
    put(f"B{rr}",nm,F(9.5,True),al=LEF,bd=True,fl=INPUTF)
    for i,v in enumerate(vals): put(f"{get_column_letter(5+i)}{rr}",v,F(9.5,False,BLUE),fmt=PCT,al=CEN,bd=True,fl=INPUTF)
    rr+=1

# forecast formulas E..I
for i in range(5):
    col=get_column_letter(5+i); prev=get_column_letter(4+i)
    gcol=get_column_letter(5+i)
    ws[f"{col}{R_G}"]=f"=CHOOSE({SCN},{gcol}28,{gcol}29,{gcol}30)"
    ws[f"{col}{R_M}"]=f"=CHOOSE({SCN},{gcol}32,{gcol}33,{gcol}34)"
    ws[f"{col}{R_REV}"]=f"={prev}{R_REV}*(1+{col}{R_G})"
    ws[f"{col}{R_EBIT}"]=f"={col}{R_REV}*{col}{R_M}"
    ws[f"{col}{R_TAX}"]=f"={col}{R_EBIT}*{TAX}"
    ws[f"{col}{R_NOP}"]=f"={col}{R_EBIT}-{col}{R_TAX}"
    ws[f"{col}{R_DA}"]=f"={col}{R_REV}*{DAr}"
    ws[f"{col}{R_CX}"]=f"={col}{R_REV}*{CXr}"
    ws[f"{col}{R_NWC}"]=f"=({col}{R_REV}-{prev}{R_REV})*{NWCr}"
    ws[f"{col}{R_FCFF}"]=f"={col}{R_NOP}+{col}{R_DA}-{col}{R_CX}-{col}{R_NWC}"
    ws[f"{col}{R_T}"]=i+1
    ws[f"{col}{R_DF}"]=f"=1/(1+{WACC})^{col}{R_T}"
    ws[f"{col}{R_PV}"]=f"={col}{R_FCFF}*{col}{R_DF}"
    ws[f"{col}{R_EBITDA}"]=f"={col}{R_EBIT}+{col}{R_DA}"

# ---- Valuation bridge (B/D rows 36-58) ----
put("B36","估值橋接 VALUATION BRIDGE",F(12,True,NAVY))
def vrow(rr,lab,formula,fmt,bold=False,hl=None,note=""):
    put(f"B{rr}",lab,F(10,bold,DARK),al=LEF,bd=True,fl=hl or (LGREY if rr%2 else WHITE))
    c=put(f"D{rr}",formula,F(11 if bold else 10,bold,BLUE if bold else DARK),fmt=fmt,al=RGT,bd=True,fl=hl or (LGREY if rr%2 else WHITE))
    if note: put(f"E{rr}",note,F(8.5,False,MIDG),al=LEF); ws.merge_cells(f"E{rr}:I{rr}")
    return c
vrow(37,"顯性期 PV 合計 (FCFF)","=SUM(E20:I20)",BN,note="FY26–FY30 折現後現金流加總")
vrow(38,"終年 EBITDA (FY30E)","=I22",BN)
vrow(39,"終值 TV (Exit Multiple 法)","=I22*"+EXM,BN,note="= 終年 EBITDA × 終值 EV/EBITDA 倍數")
vrow(40,"終值現值 PV of TV","=D39*I19",BN)
vrow(41,"企業價值 EV","=D37+D40",BN,bold=True,hl="E3F1F8")
vrow(42,"  減：淨負債 Net Debt","="+ND,BN)
vrow(43,"股權價值 Equity Value","=D41-D42",BN,bold=True,hl="E3F1F8")
vrow(44,"  股數 (bn)","="+SH,"0.000")
vrow(45,"每股價值 (Exit Mult.)","=D43/D44",USD,bold=True,hl="D7F0E8")
vrow(46,"現價 Current Price","="+PX,USD)
c=vrow(47,"上漲/下跌空間","=D45/D46-1",PCTS,bold=True,hl="D7F0E8")
put("B49","交叉驗證 CROSS-CHECK",F(12,True,NAVY))
vrow(50,"終值 (永續成長法 Gordon)","=I17*(1+"+G+")/("+WACC+"-"+G+")",BN,note="= FCFF₃₀×(1+g)/(WACC−g)")
vrow(51,"  其現值","=D50*I19",BN)
vrow(52,"企業價值 EV (永續法)","=D37+D51",BN,bold=True)
vrow(53,"每股價值 (永續法)","=(D52-"+ND+")/"+SH,USD,bold=True,hl="D7F0E8")
vrow(54,"DCF 隱含 EV/EBITDA (FY26E)","=D41/E22",MULT,note="與同業中位/Vertiv 現倍數比較")
vrow(55,"終值占 EV 比重","=D40/D41",PCT,note="健康區間約 50–70%")
vrow(56,"DCF 隱含 P/E (vs FY26E NOPAT)","=D43/(E13)",MULT,note="粗略；NOPAT 近似淨利")

# widths
ws.column_dimensions["A"].width=2
ws.column_dimensions["B"].width=30
ws.column_dimensions["C"].width=2
for i in range(6): ws.column_dimensions[get_column_letter(4+i)].width=11
ws.column_dimensions["J"].width=2
ws.column_dimensions["K"].width=18; ws.column_dimensions["L"].width=11

# ================= Sensitivity sheet =================
se=wb.create_sheet("Sensitivity"); se.sheet_view.showGridLines=False
def sput(cell,val,font=None,fmt=None,al=None,bd=False,fl=None):
    c=se[cell]; c.value=val; c.font=font or F()
    if fmt:c.number_format=fmt
    if al:c.alignment=al
    if bd:c.border=border
    if fl:c.fill=fill(fl)
    return c
se.merge_cells("B1:H1"); sput("B1","敏感度分析 SENSITIVITY（基於目前情境之 FCFF 流）",F(14,True,WHITE),al=Alignment(horizontal="left",vertical="center"),fl=NAVY)
se.row_dimensions[1].height=24
FCFF="DCF!$E$17:$I$17"; PER="DCF!$E$18:$I$18"; T5="DCF!$I$18"; TEBIT="DCF!$I$22"; F5="DCF!$I$17"
NDr="DCF!$L$10"; SHr="DCF!$L$11"; PXr="DCF!$L$12"

# Table 1: WACC (rows) vs Exit Multiple (cols)
sput("B3","每股價值：WACC × 終值 EV/EBITDA",F(11,True,NAVY))
waccs=[0.085,0.09,0.095,0.10,0.105,0.11,0.115]
exits=[16,18,20,22,24,26]
sput("B5","WACC ↓ / 倍數 →",F(9.5,True,WHITE),al=CEN,bd=True,fl=BLUE)
for j,m in enumerate(exits):
    sput(f"{get_column_letter(3+j)}5",m,F(10,True,WHITE),fmt=MULT,al=CEN,bd=True,fl=BLUE)
for i,w in enumerate(waccs):
    rr=6+i
    sput(f"B{rr}",w,F(10,True,WHITE),fmt=PCT,al=CEN,bd=True,fl=BLUE)
    for j,m in enumerate(exits):
        col=get_column_letter(3+j)
        f=f"=(SUMPRODUCT({FCFF},1/(1+{w})^{PER})+({TEBIT}*{m})/(1+{w})^{T5}-{NDr})/{SHr}"
        c=sput(f"{col}{rr}",f,F(10),fmt=USD,al=CEN,bd=True,fl=(LGREY if i%2 else WHITE))

# Table 2: WACC vs terminal growth (perpetuity)
sput("B15","每股價值：WACC × 終值成長 g（永續法）",F(11,True,NAVY))
gs=[0.025,0.03,0.035,0.04,0.045]
sput("B17","WACC ↓ / g →",F(9.5,True,WHITE),al=CEN,bd=True,fl=BLUE)
for j,g in enumerate(gs):
    sput(f"{get_column_letter(3+j)}17",g,F(10,True,WHITE),fmt=PCT,al=CEN,bd=True,fl=BLUE)
for i,w in enumerate(waccs):
    rr=18+i
    sput(f"B{rr}",w,F(10,True,WHITE),fmt=PCT,al=CEN,bd=True,fl=BLUE)
    for j,g in enumerate(gs):
        col=get_column_letter(3+j)
        f=f"=(SUMPRODUCT({FCFF},1/(1+{w})^{PER})+(({F5}*(1+{g}))/({w}-{g}))/(1+{w})^{T5}-{NDr})/{SHr}"
        sput(f"{col}{rr}",f,F(10),fmt=USD,al=CEN,bd=True,fl=(LGREY if i%2 else WHITE))
sput("B26","註：敏感度表使用 DCF 分頁『目前情境』之 FCFF 與終年 EBITDA。切換情境(DCF!L2)後此表同步更新。",F(9,False,MIDG),al=LEF)
se.column_dimensions["B"].width=16
for i in range(6): se.column_dimensions[get_column_letter(3+i)].width=11

# ================= Summary sheet =================
sm=wb.create_sheet("Summary"); sm.sheet_view.showGridLines=False
def mput(cell,val,font=None,fmt=None,al=None,bd=False,fl=None):
    c=sm[cell]; c.value=val; c.font=font or F()
    if fmt:c.number_format=fmt
    if al:c.alignment=al
    if bd:c.border=border
    if fl:c.fill=fill(fl)
    return c
sm.merge_cells("B1:F1"); mput("B1","Vertiv (VRT) — DCF 估值總結",F(15,True,WHITE),al=Alignment(horizontal="left",vertical="center"),fl=NAVY)
sm.row_dimensions[1].height=28
sm.merge_cells("B2:F2"); mput("B2","目前情境見 DCF!L2（1熊/2基/3牛，預設 2 基準）· 切換後本頁同步更新",F(9,False,WHITE),al=Alignment(horizontal="left",vertical="center"),fl=BLUE)
rows=[("每股價值（終值倍數法）","=DCF!D45",USD,"D7F0E8"),
      ("每股價值（永續成長法）","=DCF!D53",USD,"E3F1F8"),
      ("現價","=DCF!L12",USD,WHITE),
      ("上漲/下跌空間","=DCF!D45/DCF!L12-1",PCTS,"D7F0E8"),
      ("企業價值 EV ($bn)","=DCF!D41",BN,WHITE),
      ("股權價值 ($bn)","=DCF!D43",BN,WHITE),
      ("DCF 隱含 EV/EBITDA (FY26E)","=DCF!D54",MULT,LGREY),
      ("終值占 EV 比重","=DCF!D55",PCT,LGREY),
      ("WACC","=DCF!L3",PCT,WHITE),
      ("終值成長 g","=DCF!L4",PCT,WHITE),
      ("終值 EV/EBITDA","=DCF!L5",MULT,WHITE)]
rr=4
for (lab,f,fmt,fl) in rows:
    mput(f"B{rr}",lab,F(10.5,True,DARK),al=LEF,bd=True,fl=fl)
    mput(f"D{rr}",f,F(11,True,BLUE),fmt=fmt,al=RGT,bd=True,fl=fl)
    rr+=1
mput("B16","Comps 交叉驗證（同業，資料約 2026/6）",F(11,True,NAVY))
comp=[("同業 EV/EBITDA 中位（散熱相關）","~23x"),("Vertiv 現行 EV/EBITDA","~51x"),
      ("同業 Fwd P/E 中位","~28x"),("解讀","DCF 隱含倍數應落在『同業中位 ~ Vertiv 溢價』之間；高於 51x 須檢視成長假設是否過樂觀")]
rr=17
for (a,b) in comp:
    mput(f"B{rr}",a,F(10,False,DARK),al=LEF,bd=True,fl=LGREY if rr%2 else WHITE)
    mput(f"D{rr}",b,F(10,True,DARK),al=LEF,bd=True,fl=LGREY if rr%2 else WHITE); sm.merge_cells(f"D{rr}:H{rr}")
    rr+=1
mput("B23","關鍵假設與風險",F(11,True,NAVY))
notes=["營收：FY26 +27%（指引中點 ~$13.75B）→ FY30 +9%；液冷segment 約 40% CAGR 為主要上行來源。",
       "利潤率：EBIT 由 ~19% 擴張至 ~23%（營運槓桿 + 高階液冷組合）。",
       "終值倍數 22x 為『正常化』後估計（Vertiv 現 ~51x 不可持續）；保守採同業偏高水準。",
       "風險：① 雲端資本支出 2027–28 消化 ② 整併/價格競爭 ③ 電網瓶頸拖累建置 ④ 高估值對成長落空敏感。",
       "代碼校正提醒：本檔為 VRT；台股同業 Auras=3324、AVC=3017。"]
rr=24
for n in notes:
    mput(f"B{rr}","•  "+n,F(9.8,False,DARK),al=LEF); sm.merge_cells(f"B{rr}:H{rr}"); sm.row_dimensions[rr].height=16; rr+=1
sm.column_dimensions["B"].width=30
for c in "CDEFGH": sm.column_dimensions[c].width=12

wb.save("./out/Vertiv_VRT_DCF.xlsx")
print("SAVED ./out/Vertiv_VRT_DCF.xlsx  sheets:", wb.sheetnames)
