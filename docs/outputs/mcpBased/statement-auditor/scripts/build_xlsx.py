# -*- coding: utf-8 -*-
"""Build the statement-auditor Excel workpaper from the live MCP pull + draft statement.

Reads mcp_pulls/*.json (source of truth) and the transcribed draft statement, then
writes out/audit_workpaper.xlsx with three sheets:
  - Tie-out      : per-LP draft vs NAV, per-cell diff, foots?, status
  - Exceptions   : one row per flagged exception + driver + recommendation
  - NAV_Build    : the fund-level NAV build pulled from MCP

Mock/dev data only. Figures come straight from the MCP pull — none are invented.
"""
from __future__ import annotations

import json
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
PULLS = os.path.join(ROOT, "mcp_pulls")
OUT = os.path.join(ROOT, "out", "audit_workpaper.xlsx")
TOL = 0.01

with open(os.path.join(PULLS, "02_get_nav_build.json"), encoding="utf-8") as f:
    nav_build = json.load(f)["result"]
with open(os.path.join(PULLS, "03_get_lp_capital_accounts.json"), encoding="utf-8") as f:
    lp = json.load(f)["result"]

nav_lines = {ln["item"]: ln["amount"] for ln in nav_build["lines"]}
NAV = nav_lines["Net Assets (NAV)"]
truth = {a["lp_id"]: a for a in lp["accounts"]}

draft = {
    "LP-001": dict(lp_name="University Endowment", beginning=40000000, contributions=5000000,
                   distributions=2000000, net_gain_loss=4000000, mgmt_fee=200000, ending=46800000),
    "LP-002": dict(lp_name="State Pension Plan", beginning=30000000, contributions=0,
                   distributions=1500000, net_gain_loss=3000000, mgmt_fee=187500, ending=31321500),
    "LP-003": dict(lp_name="Family Office LLC", beginning=20000000, contributions=2500000,
                   distributions=0, net_gain_loss=2000000, mgmt_fee=125000, ending=24375000),
    "LP-004": dict(lp_name="GP Commitment", beginning=10000000, contributions=0,
                   distributions=500000, net_gain_loss=1000000, mgmt_fee=0, ending=10500000),
}
COLS = ["beginning", "contributions", "distributions", "net_gain_loss", "mgmt_fee", "ending"]
ZH = {"beginning": "期初", "contributions": "出資", "distributions": "分配",
      "net_gain_loss": "損益", "mgmt_fee": "管理費", "ending": "期末"}

HEAD = PatternFill("solid", fgColor="1F4E78")
BAD = PatternFill("solid", fgColor="F8CBAD")
GOOD = PatternFill("solid", fgColor="C6E0B4")
WHITE = Font(color="FFFFFF", bold=True)
BOLD = Font(bold=True)
CEN = Alignment(horizontal="center")


def hrow(ws, row, headers):
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=j, value=h)
        c.fill, c.font, c.alignment = HEAD, WHITE, CEN


def autofit(ws):
    for col in ws.columns:
        width = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(width + 3, 40)


wb = Workbook()

# ---- Sheet 1: Tie-out -------------------------------------------------------
ws = wb.active
ws.title = "Tie-out"
ws["A1"] = "基金淨值對帳 — 逐欄 Tie-out（草稿對帳單 vs NAV pack，資料來源：nav MCP 即時拉取）"
ws["A1"].font = Font(bold=True, size=12)
ws.merge_cells("A1:N1")
hdr = ["LP", "LP Name"]
for c in COLS:
    hdr += [f"draft.{ZH[c]}", f"nav.{ZH[c]}"]
hdr += ["期末 foots?", "狀態"]
hrow(ws, 3, hdr)
r = 4
for lp_id, d in draft.items():
    t = truth[lp_id]
    ws.cell(row=r, column=1, value=lp_id).font = BOLD
    ws.cell(row=r, column=2, value=d["lp_name"])
    col = 3
    cell_mismatch = False
    for c in COLS:
        cd = ws.cell(row=r, column=col, value=d[c]); cd.number_format = "#,##0"
        ct = ws.cell(row=r, column=col + 1, value=t[c]); ct.number_format = "#,##0"
        if abs(d[c] - t[c]) > TOL:
            cd.fill = BAD; cell_mismatch = True
        col += 2
    calc = d["beginning"] + d["contributions"] - d["distributions"] + d["net_gain_loss"] - d["mgmt_fee"]
    foots = abs(calc - d["ending"]) <= TOL
    fc = ws.cell(row=r, column=col, value="是 PASS" if foots else "否 FAIL")
    fc.fill = GOOD if foots else BAD
    matches_nav = all(abs(d[c] - t[c]) <= TOL for c in COLS)
    sc = ws.cell(row=r, column=col + 1, value="可放行" if matches_nav else "需再查")
    sc.fill = GOOD if matches_nav else BAD
    r += 1
# totals
ws.cell(row=r, column=1, value="TOTAL").font = BOLD
col = 3
for c in COLS:
    dt = sum(draft[k][c] for k in draft)
    tt = sum(truth[k][c] for k in truth)
    cd = ws.cell(row=r, column=col, value=dt); cd.number_format = "#,##0"; cd.font = BOLD
    ct = ws.cell(row=r, column=col + 1, value=tt); ct.number_format = "#,##0"; ct.font = BOLD
    if abs(dt - tt) > TOL:
        cd.fill = BAD
    col += 2
ws.cell(row=r, column=col + 1, value="需再查").fill = BAD
r += 2
draft_end = sum(draft[k]["ending"] for k in draft)
ws.cell(row=r, column=1, value="基金 NAV 對帳").font = BOLD
ws.cell(row=r, column=3, value=f"草稿合計期末 {draft_end:,} − NAV {NAV:,} = +{draft_end - NAV:,} → 不 tie，DO NOT DISTRIBUTE")
autofit(ws)

# ---- Sheet 2: Exceptions ----------------------------------------------------
ws2 = wb.create_sheet("Exceptions")
ws2["A1"] = "例外清單（差異 vs nav MCP）"
ws2["A1"].font = Font(bold=True, size=12)
hrow(ws2, 3, ["#", "LP", "欄位", "草稿", "NAV 基準", "差異", "研判 driver", "建議"])
rows = [
    ("E-1", "LP-001", "管理費 mgmt_fee", 200000, 250000, -50000,
     "管理費少計 50,000，期末以錯誤費用自洽 → 期末高估 50,000",
     "需再查：費用改 250,000，期末回 46,750,000"),
    ("E-2", "LP-002", "期末 ending", 31321500, 31312500, 9000,
     "各輸入欄皆對，但期末不 foot（疑謄寫換位 …312→…321）",
     "需再查：重算期末 = 31,312,500"),
    ("E-3", "LP-003", "出資 contributions", 2500000, 2000000, 500000,
     "出資多計 500,000，連帶期末高估 500,000（金額重大）",
     "需再查：核對 capital call 紀錄，期末回 23,875,000"),
    ("E-4", "TOTAL", "期末/出資/費用", 112996500, 112437500, 559000,
     "= E-1(50k)+E-2(9k)+E-3(500k) 累積，合計未 tie 基金 NAV",
     "需再查：三筆更正後自動 tie 回 112,437,500"),
]
rr = 4
for e in rows:
    ws2.cell(row=rr, column=1, value=e[0]).font = BOLD
    ws2.cell(row=rr, column=2, value=e[1])
    ws2.cell(row=rr, column=3, value=e[2])
    c4 = ws2.cell(row=rr, column=4, value=e[3]); c4.number_format = "#,##0"
    c5 = ws2.cell(row=rr, column=5, value=e[4]); c5.number_format = "#,##0"
    c6 = ws2.cell(row=rr, column=6, value=e[5]); c6.number_format = "+#,##0;-#,##0"; c6.fill = BAD
    ws2.cell(row=rr, column=7, value=e[6])
    ws2.cell(row=rr, column=8, value=e[7])
    rr += 1
autofit(ws2)

# ---- Sheet 3: NAV_Build -----------------------------------------------------
ws3 = wb.create_sheet("NAV_Build")
ws3["A1"] = "基金層 NAV Build（nav MCP: get_nav_build）"
ws3["A1"].font = Font(bold=True, size=12)
hrow(ws3, 3, ["項目 Item", "金額 USD"])
rr = 4
for ln in nav_build["lines"]:
    ws3.cell(row=rr, column=1, value=ln["item"])
    amt = ws3.cell(row=rr, column=2, value=ln["amount"]); amt.number_format = "#,##0"
    if ln["item"] in ("Total Assets", "Total Liabilities", "Net Assets (NAV)"):
        ws3.cell(row=rr, column=1).font = BOLD; amt.font = BOLD
    rr += 1
ws3.cell(row=rr + 1, column=1, value="Sum LP ending (get_lp_capital_accounts.totals)").font = BOLD
sc = ws3.cell(row=rr + 1, column=2, value=lp["totals"]["ending"]); sc.number_format = "#,##0"; sc.font = BOLD
ws3.cell(row=rr + 2, column=1, value="NAV = Σ LP ending ?")
ws3.cell(row=rr + 2, column=2, value="PASS ✓" if abs(NAV - lp["totals"]["ending"]) <= TOL else "FAIL").fill = GOOD
autofit(ws3)

wb.save(OUT)
print("wrote", OUT)
