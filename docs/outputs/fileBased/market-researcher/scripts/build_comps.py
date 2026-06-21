# -*- coding: utf-8 -*-
"""AI data center cooling — comps workbook with statistical benchmarking."""
import statistics as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter

NAVY   = "0B1F3A"; BLUE = "1F4E79"; ACCENT = "2E9BD6"; LGREY = "ECEFF3"
TEAL   = "14A38C"; AMBER = "E89A1C"; RED = "C03A2B"; WHITE = "FFFFFF"; MIDG = "5A5A5A"
CJK = "Microsoft JhengHei"

wb = Workbook()

thin = Side(style="thin", color="D0D5DD")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
def F(sz=10, b=False, color="222A35"): return Font(name=CJK, size=sz, bold=b, color=color)
def fill(c): return PatternFill("solid", fgColor=c)
CEN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEF = Alignment(horizontal="left", vertical="center", wrap_text=True)

# ============================================================= Comps sheet
ws = wb.active; ws.title = "Comps"
ws.sheet_view.showGridLines = False

# Title band
ws.merge_cells("A1:N1")
ws["A1"] = "AI 資料中心散熱 — 同業比較 (Comps)"
ws["A1"].font = F(16, True, WHITE); ws["A1"].fill = fill(NAVY); ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[1].height = 30
ws.merge_cells("A2:N2")
ws["A2"] = "資料約 2026/6/17–18 · 數字為即時變動之估計值，僅供方向參考 · 代碼校正：Auras=3324、AVC=3017"
ws["A2"].font = F(9, False, "FFFFFF"); ws["A2"].fill = fill(BLUE); ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[2].height = 18

headers = ["公司", "代碼", "上市地", "市值\n(USD bn)", "營收\nYoY", "營益率", "Fwd\nP/E",
           "EV/\nEBITDA", "EV/\nRevenue", "1年\n報酬", "散熱\n純度", "純標的?", "備註"]
hrow = 4
for ci, h in enumerate(headers, start=1):
    c = ws.cell(hrow, ci, h); c.font = F(10, True, WHITE); c.fill = fill(NAVY)
    c.alignment = CEN; c.border = border

# data: name, ticker, listing, mktcap, revYoY, opmargin, fwdpe, ev_ebitda, ev_rev, ret1y, purity, pureplay, note
# numeric cols stored as numbers where possible (None = n/a)
data = [
    ["Vertiv", "VRT", "US", 122, 0.28, 0.188, 46, 51, 11.3, 1.73, "純(高)", "是", "散熱領頭羊；backlog >$15B；NVIDIA 參考設計"],
    ["nVent", "NVT", "US", 28, 0.11, 0.162, 35, 32, 6.7, 1.43, "多元", "否", "液冷為成長引擎；backlog 約 3×"],
    ["Schneider", "SU.PA", "EU", 160, 0.05, 0.174, 28, 21, 4.3, 0.27, "多元", "否", "整併中老二(Motivair)；grid-to-chip"],
    ["Supermicro", "SMCI", "US", 18, 1.23, 0.045, 9, 16, 0.8, -0.36, "OEM", "否", "DLC 領先但毛利薄；治理爭議已解"],
    ["Parker Hannifin", "PH", "US", 119, 0.06, 0.217, 28, 23, 6.1, 0.43, "~1%散熱", "否", "⚠ 散熱僅約 1% 營收；快接頭供應"],
    ["Flex", "FLEX", "US", 52, 0.08, 0.054, 31, 27, 2.0, 2.26, "EMS", "否", "將分拆 CPI(含 JetCool)為新純標的"],
    ["Delta 台達電", "2308.TW", "TW", 187, 0.33, 0.178, 46, None, None, 4.54, "多元", "否", "電源+散熱一體；液冷約 9% 營收"],
    ["Auras 奇鋐", "3324.TWO", "TW", 3.2, 0.48, 0.141, 21, None, None, 1.32, "純散熱", "是", "純散熱高 beta；2026 指引 +60–70%"],
    ["AVC", "3017.TW", "TW", 31, 1.07, None, 23, None, None, 2.17, "DC散熱", "近似", "台灣熱模組龍頭；能見度到 2029"],
    ["Wiwynn 緯穎", "6669.TW", "TW", 30, 1.29, 0.066, 13.5, 13.6, 0.9, 1.03, "ODM", "否", "伺服器 ODM；散熱內嵌於機櫃"],
    ["Sunon 建準", "2421.TW", "TW", 1.3, 0.24, 0.158, 13.4, 9.9, 1.8, 0.41, "多元", "否", "風扇廠；冷板/CDU 下半年放量"],
    ["Lotes 嘉澤", "3533.TW", "TW", 8.2, 0.12, 0.30, 22.8, 18.5, 7.0, 0.67, "連接器", "否", "連接器核心；切入 UQD 快接"],
    ["Envicool 英維克", "002837.SZ", "CN", 15, 0.32, 0.086, 77, None, None, 2.37, "純(最近)", "近似", "⚠ Q1 淨利 −82%；估值高"],
    ["Asetek", "ASTK.OL", "NO", 0.064, -0.21, None, None, None, 1.4, 1.05, "~0%散熱", "否", "⚠ DC 業務休眠；非 AI 散熱標的"],
]
# numeric column indices (1-based): mktcap4, revYoY5, opm6, fwdpe7, evebitda8, evrev9, ret10
pct_cols = {5, 6, 10}
r = hrow + 1
for row in data:
    for ci, val in enumerate(row, start=1):
        c = ws.cell(r, ci, val if val is not None else "n/a")
        c.border = border
        c.fill = fill(LGREY if (r - hrow) % 2 == 0 else WHITE)
        if ci == 1:
            c.font = F(10, True); c.alignment = LEF
        elif ci == 13:
            c.font = F(9, False, MIDG); c.alignment = LEF
        else:
            c.alignment = CEN
            if ci == 10 and isinstance(val, (int, float)):
                c.font = F(10, True, RED if val < 0 else TEAL)
            else:
                c.font = F(10)
            if val is not None:
                if ci in pct_cols and isinstance(val, (int, float)):
                    c.number_format = "+0.0%;-0.0%"
                elif ci == 4 and isinstance(val, (int, float)):
                    c.number_format = "$#,##0.0"
                elif ci in (7, 8, 9) and isinstance(val, (int, float)):
                    c.number_format = '0.0"x"'
    r += 1

# ---- Statistical benchmarking block ----
stat_start = r + 1
ws.cell(stat_start, 1, "統計基準 (Statistical Benchmarking)").font = F(12, True, NAVY)
labels = ["平均 Mean", "中位 Median", "上四分位 Q3", "下四分位 Q1"]
# collect numeric per column
def colvals(ci):
    out = []
    for row in data:
        v = row[ci-1]
        if isinstance(v, (int, float)):
            out.append(v)
    return out
def q(vals, p):
    vals = sorted(vals)
    if not vals: return None
    k = (len(vals)-1)*p; f = int(k); cdiff = k-f
    if f+1 < len(vals): return vals[f] + (vals[f+1]-vals[f])*cdiff
    return vals[f]
stat_cols = [4,5,6,7,8,9,10]
# header row for stats
hr = stat_start + 1
ws.cell(hr, 3, "指標").font = F(10, True, WHITE); ws.cell(hr,3).fill = fill(BLUE); ws.cell(hr,3).alignment=CEN; ws.cell(hr,3).border=border
for idx, ci in enumerate(stat_cols, start=4):
    c = ws.cell(hr, ci, headers[ci-1].replace("\n"," ")); c.font=F(9,True,WHITE); c.fill=fill(BLUE); c.alignment=CEN; c.border=border
for li, lab in enumerate(labels):
    rr = hr + 1 + li
    lc = ws.cell(rr, 3, lab); lc.font=F(10,True); lc.alignment=LEF; lc.border=border; lc.fill=fill(LGREY if li%2==0 else WHITE)
    for ci in stat_cols:
        vals = colvals(ci)
        if lab.startswith("平均"): v = st.mean(vals)
        elif lab.startswith("中位"): v = st.median(vals)
        elif lab.startswith("上四"): v = q(vals, 0.75)
        else: v = q(vals, 0.25)
        c = ws.cell(rr, ci, v); c.alignment=CEN; c.font=F(10); c.border=border; c.fill=fill(LGREY if li%2==0 else WHITE)
        if ci in pct_cols: c.number_format = "+0.0%;-0.0%"
        elif ci == 4: c.number_format = "$#,##0.0"
        else: c.number_format = '0.0"x"'

# ---- Outlier flags ----
ofs = hr + len(labels) + 3
ws.cell(ofs, 1, "離群值標記 (Outliers)").font = F(12, True, NAVY)
outliers = [
    ("Envicool — Fwd P/E ~77x", "遠高於同業中位，AI 敘事驅動 + Q1 淨利 −82%，估值風險最高", RED),
    ("Delta 台達電 — 1年報酬 +454%", "AI 基礎設施再評價最劇烈，注意追高風險", AMBER),
    ("Supermicro — 1年報酬 −36%", "唯一負報酬；治理history + 低毛利的 OEM 折價", RED),
    ("Vertiv — EV/EBITDA ~51x", "散熱純度溢價最明顯（最乾淨曝險）", TEAL),
    ("Lotes — 營益率 30%", "高毛利元件 vs. 低毛利 ODM/OEM，EV/Rev 不可直接互比", BLUE),
    ("Parker / Asetek", "散熱曝險過小或休眠，倍數非散熱溢價 — 統計時宜剔除", MIDG),
]
orr = ofs + 1
for (t, d, col) in outliers:
    ws.cell(orr, 1, "▍").font = F(11, True, col)
    cc = ws.cell(orr, 2, t); cc.font = F(10, True, col)
    ws.merge_cells(start_row=orr, start_column=4, end_row=orr, end_column=13)
    dc = ws.cell(orr, 4, d); dc.font = F(9.5, False, MIDG); dc.alignment = LEF
    orr += 1

# column widths
widths = [16, 10, 7, 10, 8, 8, 8, 9, 9, 9, 9, 8, 40]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A5"

# ============================================================= Notes sheet
ns = wb.create_sheet("Notes & Sources")
ns.sheet_view.showGridLines = False
ns.column_dimensions["A"].width = 4
ns.column_dimensions["B"].width = 120
ns.merge_cells("A1:B1")
ns["A1"] = "方法論、注意事項與來源"; ns["A1"].font = F(14, True, WHITE); ns["A1"].fill = fill(NAVY)
ns.row_dimensions[1].height = 26
notes = [
    ("H", "方法論注意"),
    ("",  "市場規模/倍數來自第三方研究與公開資訊；數字為即時變動之估計值，僅供方向參考。"),
    ("",  "EV/EBITDA 與 EV/Revenue 不可跨『元件 vs. 系統/ODM』直接比較：ODM/OEM 為低毛利轉手收入，EV/Rev 天生偏低。"),
    ("",  "台股/中國中小型（Auras、AVC、Delta、Envicool）部分 EV 倍數與『散熱營收占比』為估計值，免費資料源缺漏處標 n/a。"),
    ("",  "統計基準（平均/中位/四分位）僅納入有數值之欄位；建議剔除 Parker、Asetek 等散熱曝險過小/休眠者再做純散熱基準。"),
    ("H", "關鍵警示"),
    ("",  "Parker (PH)：資料中心散熱僅約 1% 營收 — 倍數反映航太/工業品質，非散熱溢價。"),
    ("",  "Asetek (ASTK)：DC 液冷業務休眠（最後 HPC 出貨約 2022 Q1）— 不應視為 AI 散熱標的。"),
    ("",  "Supermicro (SMCI)：伺服器 OEM，散熱為綁定功能；2024–25 會計/治理事件已於 2026/1 恢復 Nasdaq 合規。"),
    ("",  "Envicool (002837)：最接近純散熱，但 2026 Q1 淨利年減約 82%，估值（~77x fwd P/E）高度由 AI 敘事驅動。"),
    ("",  "代碼校正：Auras 奇鋐 = 3324（TPEx）、AVC = 3017（TWSE）— 中文名易混淆，下單前務必核對。"),
    ("H", "主要來源"),
    ("",  "Dell'Oro Group（DCPI / 液冷市場 / GPU TDP）· Omdia（市場規模 CAGR）· IDTechEx（技術分段份額）"),
    ("",  "Uptime Institute 2025（DLC 採用率、PUE）· JLL / McKinsey / SemiAnalysis（需求/機櫃密度）"),
    ("",  "NVIDIA GTC/CES（路線圖）· 公司 IR / SEC 8-K / Digitimes / TrendForce / stockanalysis.com（公司財務）"),
    ("H", "免責聲明"),
    ("",  "本工作簿彙整自第三方研究與公開資訊，僅供內部研究參考，不構成投資建議。"),
]
rr = 3
for (tag, text) in notes:
    if tag == "H":
        c = ns.cell(rr, 2, text); c.font = F(12, True, BLUE)
        rr += 1
    else:
        c = ns.cell(rr, 2, "•  " + text); c.font = F(10, False, "222A35"); c.alignment = LEF
        ns.row_dimensions[rr].height = 16
        rr += 1

wb.save("./out/AI資料中心散熱_Comps.xlsx")
print("SAVED ./out/AI資料中心散熱_Comps.xlsx")
