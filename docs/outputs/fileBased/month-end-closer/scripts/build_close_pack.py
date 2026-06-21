# -*- coding: utf-8 -*-
"""Aurora Capital Management LLC — 2026-04 month-end close workbook (live formulas)."""
import os
os.makedirs("./out", exist_ok=True)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BLUEF="0000FF"; BLACKF="000000"; GREENF="008000"
HDR="1F4E79"; SUB="D9E1F2"; OUT="BDD7EE"; INP="F2F2F2"; WHITE="FFFFFF"; WARN="FCE4D6"; OK="E2EFDA"
CJK="Microsoft JhengHei"
med=Side(style="thin",color="9BAFC4"); cell_b=Border(left=med,right=med,top=med,bottom=med)
def Fi(s=10,b=False): return Font(name=CJK,size=s,bold=b,color=BLUEF)
def Ff(s=10,b=False): return Font(name=CJK,size=s,bold=b,color=BLACKF)
def Fl(s=10,b=False): return Font(name=CJK,size=s,bold=b,color=GREENF)
def Fh(s=11,b=True):  return Font(name=CJK,size=s,bold=b,color=WHITE)
def fill(c): return PatternFill("solid",fgColor=c)
CEN=Alignment(horizontal="center",vertical="center",wrap_text=True)
RGT=Alignment(horizontal="right",vertical="center"); LEF=Alignment(horizontal="left",vertical="center",wrap_text=True)
USD='#,##0;(#,##0);"-"'; PCT="+0.0%;(0.0%)"

wb=Workbook()
def sheet(name,idx=None):
    s=wb.create_sheet(name) if idx is None else wb.create_sheet(name,idx); s.sheet_view.showGridLines=False; return s
def W(ws,cell,v,f,fmt=None,al=RGT,fl=None):
    c=ws[cell]; c.value=v; c.font=f; c.border=cell_b; c.alignment=al
    if fmt:c.number_format=fmt
    if fl:c.fill=fill(fl)
def hdr(ws,rng,t):
    a=rng.split(":")[0]; ws[a]=t; ws[a].font=Fh(11); ws.merge_cells(rng)
    for row in ws[rng]:
        for c in row: c.fill=fill(HDR); c.border=cell_b
    ws[a].alignment=Alignment(horizontal="left",vertical="center")

# ===================== Trial Balance =====================
tb=sheet("Trial Balance")
tb.merge_cells("A1:H1"); tb["A1"]="試算表 Trial Balance — Aurora Capital Management LLC · 2026-04 (USD)"
tb["A1"].font=Fh(13); tb["A1"].fill=fill(HDR); tb["A1"].alignment=Alignment(horizontal="left",vertical="center"); tb.row_dimensions[1].height=24
heads=["科目","名稱","類別","Dr/Cr","上月 03","本月 04 pre","應計調整","結帳後 post"]
for i,h in enumerate(heads): W(tb,f"{get_column_letter(1+i)}3",h,Fh(9),al=CEN if i>1 else LEF,fl=HDR)
rows=[
 ("1000","現金 Cash","資產","Dr",8000000,7550000,0),
 ("1100","應收管理費","資產","Dr",3000000,3200000,0),
 ("1200","預付費用","資產","Dr",600000,600000,0),
 ("1500","固定資產淨額","資產","Dr",1400000,1350000,0),
 ("2000","應付帳款","負債","Cr",900000,1100000,0),
 ("2100","應付費用","負債","Cr",1200000,1000000,275000),
 ("2200","遞延收入","負債","Cr",500000,400000,0),
 ("3000","成員資本","權益","Cr",8000000,8000000,0),
 ("3100","期初保留盈餘","權益","Cr",2000000,2000000,0),
 ("4000","管理費收入","收入","Cr",6300000,8500000,0),
 ("4100","其他收入","收入","Cr",300000,380000,0),
 ("6000","薪酬","費用","Dr",4500000,6000000,100000),
 ("6100","租金","費用","Dr",600000,800000,0),
 ("6200","專業服務費","費用","Dr",400000,750000,100000),
 ("6300","資訊技術","費用","Dr",300000,400000,75000),
 ("6400","差旅","費用","Dr",150000,350000,0),
 ("6500","其他營運費用","費用","Dr",250000,330000,0),
 ("6600","折舊","費用","Dr",0,50000,0),
]
r0=4
for i,(ac,nm,ty,dc,pr,pre,adj) in enumerate(rows):
    r=r0+i
    W(tb,f"A{r}",ac,Ff(9),al=CEN); W(tb,f"B{r}",nm,Ff(9),al=LEF); W(tb,f"C{r}",ty,Ff(9),al=CEN); W(tb,f"D{r}",dc,Ff(9),al=CEN)
    W(tb,f"E{r}",pr,Fi(9),USD); W(tb,f"F{r}",pre,Fi(9),USD)
    W(tb,f"G{r}",adj if adj else None,Fi(9,True) if adj else Ff(9),USD,fl=(SUB if adj else None))
    W(tb,f"H{r}",f"=F{r}+N(G{r})",Ff(9,True),USD)   # post-close = preclose + adj
rN=r0+len(rows)
# Dr/Cr split for balance check (helper cols J,K) using post-close
W(tb,"I3","",Fh(9),fl=HDR);
W(tb,"J3","Dr(post)",Fh(9),al=CEN,fl=HDR); W(tb,"K3","Cr(post)",Fh(9),al=CEN,fl=HDR)
for i in range(len(rows)):
    r=r0+i
    W(tb,f"J{r}",f'=IF(D{r}="Dr",H{r},0)',Ff(9),USD); W(tb,f"K{r}",f'=IF(D{r}="Cr",H{r},0)',Ff(9),USD)
# totals
W(tb,f"B{rN}","合計 / 平衡檢查",Ff(10,True),al=LEF,fl=SUB)
for col in ["E","F","G","H"]:
    W(tb,f"{col}{rN}",f"=SUM({col}{r0}:{col}{rN-1})",Ff(10,True),USD,fl=SUB)
W(tb,f"J{rN}",f"=SUM(J{r0}:J{rN-1})",Ff(10,True),USD,fl=OUT); W(tb,f"K{rN}",f"=SUM(K{r0}:K{rN-1})",Ff(10,True),USD,fl=OUT)
W(tb,f"B{rN+1}","借貸差額 Dr−Cr (post)",Ff(10,True),al=LEF)
W(tb,f"J{rN+1}",f"=J{rN}-K{rN}",Ff(10,True),USD,fl=OK)
W(tb,f"K{rN+1}","← 應為 0 ✓",Ff(9),al=LEF)
# net income
W(tb,f"B{rN+3}","YTD 淨利(post) = 收入 − 費用",Ff(10,True),al=LEF)
W(tb,f"H{rN+3}",f"=(H{r0+9}+H{r0+10})-SUM(H{r0+11}:H{r0+17})",Ff(10,True),USD,fl=OUT)
W(tb,f"B{rN+4}","（pre-close 淨利對照）",Ff(9),al=LEF)
W(tb,f"H{rN+4}",f"=(F{r0+9}+F{r0+10})-SUM(F{r0+11}:F{r0+17})",Ff(9),USD)
tb.column_dimensions["A"].width=7; tb.column_dimensions["B"].width=16; tb.column_dimensions["C"].width=7; tb.column_dimensions["D"].width=6
for c in "EFGHJK": tb.column_dimensions[c].width=13

# ===================== Accruals & JEs =====================
ac=sheet("Accruals & JEs",0)
ac.merge_cells("A1:F1"); ac["A1"]="應計排程與草稿分錄 Accruals & Draft JEs · 2026-04（待 controller 核准，勿過帳）"
ac["A1"].font=Fh(12); ac["A1"].fill=fill(HDR); ac["A1"].alignment=Alignment(horizontal="left",vertical="center")
for i,h in enumerate(["#","項目","基礎 Basis","費用科目","本期估列","支援 Support"]):
    W(ac,f"{get_column_letter(1+i)}3",h,Fh(9),al=CEN if i in(0,4) else LEF,fl=HDR)
acc=[("A1","IT 服務(4月)","月帳單 ~75,000","6300 資訊技術",75000,"發票 2026-05-03,屬4月;發票入帳時沖回"),
 ("A2","年度分紅","獎金池 1,200,000 ÷ 12","6000 薪酬",100000,"薪酬計畫(按月攤)"),
 ("A3","FY2026 外部審計","全年 300,000,~1/3","6200 專業服務費",100000,"審計委任書(累計估列,已入帳0)")]
for i,(n,it,bs,exa,amt,sup) in enumerate(acc):
    r=4+i
    W(ac,f"A{r}",n,Ff(9),al=CEN); W(ac,f"B{r}",it,Ff(9),al=LEF); W(ac,f"C{r}",bs,Ff(9),al=LEF)
    W(ac,f"D{r}",exa,Ff(9),al=LEF); W(ac,f"E{r}",amt,Fi(9),USD); W(ac,f"F{r}",sup,Ff(8),al=LEF)
W(ac,"D7","合計",Ff(10,True),al=LEF,fl=SUB); W(ac,"E7","=SUM(E4:E6)",Ff(10,True),USD,fl=SUB)
# JE block
hdr(ac,"A9:F9","草稿分錄 Draft Journal Entries")
je=[("JE-A1","Dr 6300 資訊技術",75000,"Cr 2100 應付費用",75000,"IT 服務 2026-04 估列(自動沖回)"),
 ("JE-A2","Dr 6000 薪酬",100000,"Cr 2100 應付費用",100000,"分紅 2026-04 按月估列"),
 ("JE-A3","Dr 6200 專業服務費",100000,"Cr 2100 應付費用",100000,"FY2026 審計累計估列")]
W(ac,"A10","JE",Fh(9),al=CEN,fl=HDR); W(ac,"B10","借方 Dr",Fh(9),al=LEF,fl=HDR); W(ac,"C10","金額",Fh(9),al=CEN,fl=HDR)
W(ac,"D10","貸方 Cr",Fh(9),al=LEF,fl=HDR); W(ac,"E10","金額",Fh(9),al=CEN,fl=HDR); W(ac,"F10","Memo",Fh(9),al=LEF,fl=HDR)
for i,(j,dr,da,cr,ca,memo) in enumerate(je):
    r=11+i
    W(ac,f"A{r}",j,Ff(9),al=CEN); W(ac,f"B{r}",dr,Ff(9),al=LEF); W(ac,f"C{r}",da,Fi(9),USD)
    W(ac,f"D{r}",cr,Ff(9),al=LEF); W(ac,f"E{r}",ca,Fi(9),USD); W(ac,f"F{r}",memo,Ff(8),al=LEF)
W(ac,"B14","JE 合計(Dr=Cr 檢查)",Ff(10,True),al=LEF,fl=SUB)
W(ac,"C14","=SUM(C11:C13)",Ff(10,True),USD,fl=OUT); W(ac,"E14","=SUM(E11:E13)",Ff(10,True),USD,fl=OUT)
W(ac,"F14","← Dr 應 = Cr ✓",Ff(9),al=LEF)
ac.column_dimensions["A"].width=7; ac.column_dimensions["B"].width=20; ac.column_dimensions["C"].width=12
ac.column_dimensions["D"].width=20; ac.column_dimensions["E"].width=12; ac.column_dimensions["F"].width=34

# ===================== Roll-forward 2100 =====================
rf=sheet("Roll-forward 2100")
rf.merge_cells("A1:C1"); rf["A1"]="應付費用 (2100) 結轉 Roll-forward · 2026-04"
rf["A1"].font=Fh(12); rf["A1"].fill=fill(HDR); rf["A1"].alignment=Alignment(horizontal="left",vertical="center")
data=[("期初餘額(2026-03 月底)",1200000,"input"),("減：4月支付/沖回",-200000,"input"),
      ("＝ 結帳前餘額",None,"=C3+C4"),("加：新估列(A1+A2+A3)",None,"='Accruals & JEs'!E7"),
      ("＝ 期末餘額(post-close)",None,"=C5+C6")]
r=3
for lab,val,how in data:
    W(rf,f"B{r}",lab,Ff(10,how!="input"),al=LEF,fl=(SUB if how!="input" and "＝" in lab else None))
    if how=="input": W(rf,f"C{r}",val,Fi(10),USD)
    elif how.startswith("="): W(rf,f"C{r}",how,Ff(10,True),USD,fl=(OUT if "期末" in lab else None))
    r+=1
W(rf,"B10","Foot check：期初+變動=期末",Ff(10,True),al=LEF)
W(rf,"C10","=IF(C7=C3+C4+C6,\"✓ 對得上\",\"✗\")",Ff(10,True),al=CEN,fl=OK)
W(rf,"B11","對試算表 2100 post-close 一致？",Ff(9),al=LEF)
W(rf,"C11","='Trial Balance'!H9",Fl(10),USD)
rf.column_dimensions["A"].width=2; rf.column_dimensions["B"].width=28; rf.column_dimensions["C"].width=16

# ===================== Variance =====================
va=sheet("Variance")
va.merge_cells("A1:H1"); va["A1"]="差異說明 Variance（當月 vs 前3月跑率；門檻 >5%）· post-close"
va["A1"].font=Fh(12); va["A1"].fill=fill(HDR); va["A1"].alignment=Alignment(horizontal="left",vertical="center")
for i,h in enumerate(["科目","名稱","前YTD","後YTD","當月","月跑率","當月vs跑率","判讀"]):
    W(va,f"{get_column_letter(1+i)}3",h,Fh(9),al=CEN if i>1 else LEF,fl=HDR)
# reference TB rows for expense accounts (TB rows: 6000=r15.. mapping)
exp=[("6000","薪酬",15,"分紅估列 A2 驅動;跑率持平。已解釋"),
 ("6100","租金",16,"純 YTD 機械增加,符合跑率。正常"),
 ("6200","專業服務費",17,"⚠ 審計估列+其餘~$350k 超跑率,來源不明→需查"),
 ("6300","資訊技術",18,"多為 A1 估列;其餘符合跑率。已解釋"),
 ("6400","差旅",19,"⚠ 當月為跑率4倍,來源不明→需查"),
 ("6500","其他營運",20,"符合跑率。正常"),
 ("6600","折舊",21,"⚠ 1–3月為0、4月才提,時點/補提疑慮→需查")]
for i,(acn,nm,tbr,note) in enumerate(exp):
    r=4+i; warn="⚠" in note
    W(va,f"A{r}",acn,Ff(9),al=CEN); W(va,f"B{r}",nm,Ff(9),al=LEF)
    W(va,f"C{r}",f"='Trial Balance'!E{tbr}",Fl(9),USD)
    W(va,f"D{r}",f"='Trial Balance'!H{tbr}",Fl(9),USD)
    W(va,f"E{r}",f"=D{r}-C{r}",Ff(9),USD)            # 當月 = postYTD - priorYTD
    W(va,f"F{r}",f"=C{r}/3",Ff(9),USD)               # 月跑率
    W(va,f"G{r}",f"=IF(F{r}=0,\"n/a\",E{r}/F{r}-1)",Ff(9,True),PCT)
    W(va,f"H{r}",note,Ff(8,warn),al=LEF,fl=(WARN if warn else None))
va.column_dimensions["A"].width=7; va.column_dimensions["B"].width=12
for c in "CDEFG": va.column_dimensions[c].width=12
va.column_dimensions["H"].width=40

if "Sheet" in wb.sheetnames: wb.remove(wb["Sheet"])
wb.save("./out/Close_Package_2026-04.xlsx")
print("SAVED ./out/Close_Package_2026-04.xlsx  sheets:", wb.sheetnames)
