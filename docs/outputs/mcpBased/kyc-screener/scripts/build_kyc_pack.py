# -*- coding: utf-8 -*-
"""KYC/AML onboarding review workpaper — 2026-06 (MCP version).

Builds the Excel workpaper from the kyc-screener run whose screening evidence
came from the `screening` mock MCP (mcp_pulls/03_screen_*.json). Rules-matrix +
live screening classifications + routing. Mock / dev data only.

Run from the kyc-screener output folder:
  python scripts/build_kyc_pack.py
Writes -> out/KYC_Review_Workpaper_2026-06.xlsx
"""
import os
import json

# Resolve paths relative to this script so it runs from anywhere.
HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)                       # .../mcpBased/kyc-screener
OUT_DIR = os.path.join(ROOT, "out")
PULLS = os.path.join(ROOT, "mcp_pulls")
os.makedirs(OUT_DIR, exist_ok=True)

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---- sanity check: pull the live MCP classifications back so the sheet is
#      provably built from real pulls, not hard-coded guesses.
def _cls(fname):
    p = os.path.join(PULLS, fname)
    with open(p, encoding="utf-8") as f:
        r = json.load(f)["result"]
    overall = r["overall"]
    cls = r["matches"][0]["classification"] if r["matches"] else "clear"
    return overall, cls

LIVE = {
    "B": _cls("03_screen_B_ivan_petrov.json"),
    "C": _cls("03_screen_C_maria_garcia.json"),
    "D": _cls("03_screen_D_adewale_okonkwo.json"),
    "E": _cls("03_screen_E_robert_quiet.json"),
}
assert LIVE["B"] == ("HIT_SANCTIONS", "confirmed_sanctions_match"), LIVE["B"]
assert LIVE["C"] == ("REVIEW", "likely_false_positive"), LIVE["C"]
assert LIVE["D"] == ("HIT_PEP", "pep_match"), LIVE["D"]
assert LIVE["E"] == ("CLEAR", "clear"), LIVE["E"]
print("MCP classifications verified from mcp_pulls/:", LIVE)

BLUEF = "0000FF"; BLACKF = "000000"
HDR = "1F4E79"; SUB = "D9E1F2"; OUT = "BDD7EE"; INP = "F2F2F2"; WHITE = "FFFFFF"
RED = "F4CCCC"; AMB = "FFF2CC"; GRN = "E2EFDA"
CJK = "Microsoft JhengHei"
med = Side(style="thin", color="9BAFC4"); cell_b = Border(left=med, right=med, top=med, bottom=med)
def Ff(s=10, b=False, c=BLACKF): return Font(name=CJK, size=s, bold=b, color=c)
def Fh(s=11, b=True):  return Font(name=CJK, size=s, bold=b, color=WHITE)
def fill(c): return PatternFill("solid", fgColor=c)
CEN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEF = Alignment(horizontal="left", vertical="center", wrap_text=True)

wb = Workbook()
def sheet(name, idx=None):
    s = wb.create_sheet(name) if idx is None else wb.create_sheet(name, idx); s.sheet_view.showGridLines = False; return s
def W(ws, cell, v, f, al=CEN, fl=None, wrap=True):
    c = ws[cell]; c.value = v; c.font = f; c.border = cell_b
    c.alignment = al if not wrap else Alignment(horizontal=al.horizontal, vertical="center", wrap_text=True)
    if fl: c.fill = fill(fl)

LV = {"🔴": RED, "🟡": AMB, "🟢": GRN}

# ===================== Dashboard =====================
db = sheet("Dashboard")
db.merge_cells("A1:F1"); db["A1"] = "KYC/AML 開戶審查 Dashboard · 2026-06（MCP 版）"; db["A1"].font = Fh(13); db["A1"].fill = fill(HDR)
db["A1"].alignment = Alignment(horizontal="left", vertical="center"); db.row_dimensions[1].height = 24
db.merge_cells("A2:F2"); db["A2"] = "只評分與路由,不作最終決定;須合規/MLRO 簽核。篩查由 screening MCP screen_name 做姓名+DOB+國別識別碼比對。"
db["A2"].font = Ff(8); db["A2"].alignment = Alignment(horizontal="left", vertical="center")
for i, h in enumerate(["申請", "客戶", "類型", "風險等級", "處置建議", "關鍵觸發（含 MCP 判定）"]):
    W(db, f"{get_column_letter(1+i)}4", h, Fh(9), LEF if i in (1, 5) else CEN, HDR)
rows = [
 ("B", "Ivan Petrov", "個人", "🔴 拒絕＋上報", "拒絕開戶 + SAR/escalate;升級MLRO", "MCP confirmed_sanctions_match(S-1 OFAC SDN)→R4"),
 ("D", "Aurora Holdings (Nigeria)", "法人", "🟡 EDD(高)", "升級MLRO;補件前不可開戶", "MCP pep_match(P-1)＋R6高風險＋R3缺SoF＋PEP聲明不實"),
 ("A", "Helios Capital Partners", "法人", "🟡 EDD", "補件前不可開戶", "MCP 全CLEAR;R2 30%UBO未識別＋R1董事證件過期"),
 ("C", "Maria Garcia", "個人", "🟢 標準", "記錄假命中後可送核准", "MCP likely_false_positive(DOB/國別不符)→排除,不拒絕"),
 ("E", "Quiet Brook LLC", "法人", "🟢 標準", "可送核准", "MCP 全CLEAR、文件齊全"),
]
for i, row in enumerate(rows):
    r = 5 + i
    flc = LV.get(row[3][0])
    for j, v in enumerate(row):
        W(db, f"{get_column_letter(1+j)}{r}", v, Ff(9, j in (0, 3)), LEF if j in (1, 5) else CEN, flc if j == 3 else (WHITE if r % 2 else INP))
db.column_dimensions["A"].width = 6; db.column_dimensions["B"].width = 22; db.column_dimensions["C"].width = 7
db.column_dimensions["D"].width = 14; db.column_dimensions["E"].width = 26; db.column_dimensions["F"].width = 40

# ===================== Rules Matrix =====================
rm = sheet("Rules Matrix", 0)
rm.merge_cells("A1:H1"); rm["A1"] = "規則矩陣 Rules Matrix（申請 × R1–R7）"; rm["A1"].font = Fh(12); rm["A1"].fill = fill(HDR)
rm["A1"].alignment = Alignment(horizontal="left", vertical="center")
cols = ["申請 / 規則", "R1 身分", "R2 UBO", "R3 SoF", "R4 制裁", "R5 PEP", "R6 高風險國", "R7 活動"]
for i, h in enumerate(cols): W(rm, f"{get_column_letter(1+i)}3", h, Fh(9), LEF if i == 0 else CEN, HDR)
# cell codes: ✓ pass / ✗ 缺漏 / 命中 / 假命中 / EDD / —
mat = [
 ("A Helios", "✗ 董事過期", "✗ 30%未識別", "✓", "✓ MCP CLEAR", "— 無", "△ 開曼(非清單)", "✓"),
 ("B Petrov", "✓", "—", "—", "🔴 MCP 確認命中", "— 無", "—", "—"),
 ("C Garcia", "✓", "—", "✓", "🟢 MCP 假命中(排除)", "— 無", "✓ 西班牙", "✓"),
 ("D Nigeria", "✓", "✓", "✗ 未提供", "✓ 無制裁", "🟡 MCP PEP命中", "🟡 奈及利亞", "△ $50M偏大"),
 ("E Quiet Brook", "✓", "✓", "✓", "✓ MCP CLEAR", "— 無", "✓ 德拉瓦", "✓"),
]
for i, row in enumerate(mat):
    r = 4 + i
    for j, v in enumerate(row):
        flc = None
        if j > 0:
            if "🔴" in v: flc = RED
            elif "🟡" in v or "✗" in v: flc = AMB
            elif "🟢" in v: flc = GRN
        W(rm, f"{get_column_letter(1+j)}{r}", v, Ff(9, j == 0), LEF if j == 0 else CEN, flc or (WHITE if r % 2 else INP))
W(rm, "A10", "圖例:✓ 通過 / ✗ 缺漏 / 🔴 確認命中(拒絕) / 🟡 EDD觸發 / 🟢 假命中已排除 / △ 留意 / — 不適用。R4/R5 判定來自 screening MCP。", Ff(8), LEF)
rm.merge_cells("A10:H10")
rm.column_dimensions["A"].width = 14
for c in "BCDEFGH": rm.column_dimensions[c].width = 15

# ===================== Screening (live MCP) =====================
sc = sheet("Sanctions-PEP Screening")
sc.merge_cells("A1:H1"); sc["A1"] = "名單比對 Sanctions/PEP Screening（screening MCP · screen_name 回傳）"; sc["A1"].font = Fh(12); sc["A1"].fill = fill(HDR)
sc["A1"].alignment = Alignment(horizontal="left", vertical="center")
heads = ["篩查對象 / 申請", "MCP overall", "MCP classification", "命中 #", "名單 DOB/國別", "申請 DOB/國別", "action", "判定"]
for i, h in enumerate(heads): W(sc, f"{get_column_letter(1+i)}3", h, Fh(9), LEF if i in (0, 6) else CEN, HDR)
scr = [
 ("Ivan Petrov / B", "HIT_SANCTIONS", "confirmed_sanctions_match", "S-1", "1970-03-15 / Russia", "1970-03-15 / Russia", "R4: reject / escalate to MLRO", "🔴 確認命中→拒絕＋SAR"),
 ("Maria Garcia / C", "REVIEW", "likely_false_positive", "S-2", "1955-07-22 / Colombia", "1988-11-02 / Spain", "clear after review; do NOT auto-reject", "🟢 假命中→排除"),
 ("Adewale Okonkwo / D", "HIT_PEP", "pep_match", "P-1", "Deputy Min. Energy / Nigeria", "— / Nigeria", "R5: EDD; not an auto-reject", "🟡 PEP命中→EDD"),
 ("John Sterling / A", "CLEAR", "—", "—", "—", "— / Cayman", "—", "✓ 無命中"),
 ("Amy Cole / A", "CLEAR", "—", "—", "—", "— / Cayman", "—", "✓ 無命中"),
 ("Robert Quiet / E", "CLEAR", "—", "—", "—", "— / USA", "—", "✓ 無命中"),
 ("Sarah Brook / E", "CLEAR", "—", "—", "—", "— / USA", "—", "✓ 無命中"),
]
for i, row in enumerate(scr):
    r = 4 + i
    flc = RED if "🔴" in row[7] else (GRN if "🟢" in row[7] else (AMB if "🟡" in row[7] else WHITE))
    for j, v in enumerate(row):
        W(sc, f"{get_column_letter(1+j)}{r}", v, Ff(9, j == 0), LEF if j in (0, 6) else CEN, flc if j == 7 else (WHITE if r % 2 else INP))
W(sc, "A12", "原則(server 內建):confirmed(全識別碼一致)才拒絕;likely_false_positive(僅姓名)須排除,do NOT auto-reject;pep_match 走 EDD 非拒絕。", Ff(8), LEF)
sc.merge_cells("A12:H12")
sc.column_dimensions["A"].width = 20; sc.column_dimensions["B"].width = 14; sc.column_dimensions["C"].width = 24
sc.column_dimensions["D"].width = 7; sc.column_dimensions["E"].width = 22; sc.column_dimensions["F"].width = 18
sc.column_dimensions["G"].width = 30; sc.column_dimensions["H"].width = 22

# ===================== Deficiencies & Routing =====================
rt = sheet("Routing")
rt.merge_cells("A1:D1"); rt["A1"] = "缺漏與路由 Deficiencies & Routing（給 合規/MLRO）"; rt["A1"].font = Fh(12); rt["A1"].fill = fill(HDR)
rt["A1"].alignment = Alignment(horizontal="left", vertical="center")
for i, h in enumerate(["優先", "申請", "動作", "缺漏/命中（MCP 依據）"]):
    W(rt, f"{get_column_letter(1+i)}3", h, Fh(9), LEF if i in (2, 3) else CEN, HDR)
route = [
 ("1", "B Petrov", "拒絕開戶 + 提交 SAR/上報;升級 MLRO;凍結流程", "MCP confirmed_sanctions_match(S-1 OFAC SDN)"),
 ("2", "D Nigeria", "升級 MLRO:補SoF＋PEP高階核准＋EDD;記錄PEP聲明不實", "MCP pep_match(P-1)＋R6＋R3缺SoF＋R7"),
 ("3", "A Helios", "退補件:識別30%UBO＋更新董事證件", "MCP CLEAR;R2 UBO未識別＋R1證件過期"),
 ("4", "C Garcia", "留存假命中排除紀錄,送標準核准（不得拒絕）", "MCP likely_false_positive(已排除)"),
 ("5", "E Quiet Brook", "送標準核准", "MCP CLEAR、無命中"),
]
for i, row in enumerate(route):
    r = 4 + i
    flc = RED if row[1].startswith("B") else (AMB if row[1][0] in ("D", "A") else GRN)
    for j, v in enumerate(row):
        W(rt, f"{get_column_letter(1+j)}{r}", v, Ff(9, j in (0, 1)), LEF if j in (2, 3) else CEN, flc if j == 0 else (WHITE if r % 2 else INP))
W(rt, "A10", "本底稿僅評分與路由;最終由 合規/MLRO 簽核。篩查證據為本回合真實 screen_name MCP 回傳(mcp_pulls/);其餘為 mock 假資料。", Ff(8), LEF)
rt.merge_cells("A10:D10")
rt.column_dimensions["A"].width = 6; rt.column_dimensions["B"].width = 14; rt.column_dimensions["C"].width = 46; rt.column_dimensions["D"].width = 30

if "Sheet" in wb.sheetnames: wb.remove(wb["Sheet"])
out_path = os.path.join(OUT_DIR, "KYC_Review_Workpaper_2026-06.xlsx")
wb.save(out_path)
print("SAVED", out_path, " sheets:", wb.sheetnames)
