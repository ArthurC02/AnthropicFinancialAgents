# -*- coding: utf-8 -*-
"""Aurora Growth Fund III — Q1 2026 valuation review workpaper (live formulas).

mcpBased build: every input number is read from scripts/analysis_result.json,
which is computed by analyze.py purely from the live portfolio MCP pulls in
mcp_pulls/. The xlsx keeps live Excel formulas (XIRR / YEARFRAC / MIN / IF) so a
reviewer can change any input and watch NAV, returns and the 3-scenario carry
recompute. Mock / dev data only — needs IR + CCO sign-off before distribution.
"""
import os
import json

HERE = os.path.dirname(os.path.abspath(__file__))
OUTDIR = os.path.join(HERE, "..", "out")
os.makedirs(OUTDIR, exist_ok=True)

with open(os.path.join(HERE, "analysis_result.json"), encoding="utf-8") as f:
    A = json.load(f)

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter
import datetime

BLUEF="0000FF"; BLACKF="000000"; GREENF="008000"
HDR="1F4E79"; SUB="D9E1F2"; OUT="BDD7EE"; INP="F2F2F2"; WHITE="FFFFFF"; WARN="FCE4D6"; OK="E2EFDA"
CJK="Microsoft JhengHei"
med=Side(style="thin",color="9BAFC4"); cell_b=Border(left=med,right=med,top=med,bottom=med)
def Fi(s=10,b=False): return Font(name=CJK,size=s,bold=b,color=BLUEF)
def Ff(s=10,b=False): return Font(name=CJK,size=s,bold=b,color=BLACKF)
def Fl(s=10,b=False): return Font(name=CJK,size=s,bold=b,color=GREENF)
def Fh(s=11,b=True):  return Font(name=CJK,size=s,bold=b,color=WHITE)
def fill(c): return PatternFill("solid",fgColor=c)
CEN=Alignment(horizontal="center",vertical="center",wrap_text=True)
RGT=Alignment(horizontal="right",vertical="center"); LEF=Alignment(horizontal="left",vertical="center",wrap_text=True)
USD='#,##0;(#,##0);"-"'; MULT='0.000"x"'; PCT='0.00%'; DATEF='yyyy-mm-dd'

wb=Workbook()
def sheet(name,idx=None):
    s=wb.create_sheet(name) if idx is None else wb.create_sheet(name,idx); s.sheet_view.showGridLines=False; return s
def W(ws,cell,v,f,fmt=None,al=RGT,fl=None,cmt=None):
    c=ws[cell]; c.value=v; c.font=f; c.border=cell_b; c.alignment=al
    if fmt:c.number_format=fmt
    if fl:c.fill=fill(fl)
    if cmt:c.comment=Comment(cmt,"valuation-reviewer")
def hdr(ws,rng,t):
    a=rng.split(":")[0]; ws[a]=t; ws[a].font=Fh(11); ws.merge_cells(rng)
    for row in ws[rng]:
        for c in row: c.fill=fill(HDR); c.border=cell_b
    ws[a].alignment=Alignment(horizontal="left",vertical="center")

# ===================== 0. Source / provenance =====================
sp=sheet("Source")
sp.merge_cells("A1:D1"); sp["A1"]="資料來源 Source · portfolio MCP (mock, read-only) · 2026-Q1"
sp["A1"].font=Fh(12); sp["A1"].fill=fill(HDR); sp["A1"].alignment=Alignment(horizontal="left",vertical="center")
srows=[("Server","portfolio (mock-mcp, stdio)"),
       ("Fund",A["fund"]),("As-of",A["as_of"]),
       ("MCP 工具","list_funds, get_positions, get_fund_nav_build, get_capital_flows, get_fund_terms, get_valuation_policy"),
       ("快照","mcp_pulls/*.json (+_tools.json)"),
       ("計算","scripts/analyze.py → analysis_result.json"),
       ("注意","全為假資料/dev；分送前須 IR+CCO 簽核")]
r=3
for k,v in srows:
    W(sp,f"A{r}",k,Ff(10,True),al=LEF,fl=SUB); W(sp,f"B{r}",v,Ff(10),al=LEF); sp.merge_cells(f"B{r}:D{r}"); r+=1
sp.column_dimensions["A"].width=12; sp.column_dimensions["B"].width=30
for c in "CD": sp.column_dimensions[c].width=18

# ===================== 1. Valuation Review =====================
vr=sheet("Valuation Review")
vr.merge_cells("A1:J1"); vr["A1"]="GP 估值複核 Valuation Review（對照政策 P1–P4）· "+A["fund"]+" · "+A["as_of"]
vr["A1"].font=Fh(13); vr["A1"].fill=fill(HDR); vr["A1"].alignment=Alignment(horizontal="left",vertical="center"); vr.row_dimensions[1].height=24
heads=["被投資公司","成本","上季FV","GP報告FV","方法","政策檢核(P1–P4)","判定","政策調整","調整後FV","備註"]
for i,h in enumerate(heads): W(vr,f"{get_column_letter(1+i)}3",h,Fh(9),al=CEN if i>0 else LEF,fl=HDR)

# notes keyed by company (the qualitative reviewer judgement)
NOTE={
 "Helios Software":"P3：確認是否高於 2025-09 輪價，取得委員會核准文件",
 "Orion Logistics":"沿用最近一輪(2025-06)；輪價漸老化(>9月)",
 "Nova Biotech":"撤回 +25M 未佐證上調，調回上季/成本 40M",
 "Titan Manufacturing":"DCF 下修、已減損至成本以下(保守)",
 "Zephyr Retail":"自 2025-03 起 4 季未複核；流程 flag，值未驗證",
}
CHK={
 "Helios Software":"P2✓ 方法佐證；P3 待委員會核准文件",
 "Orion Logistics":"P1 本季已複核✓",
 "Nova Biotech":"❌P2 無方法/無觸發；上調僅憑管理層估計",
 "Titan Manufacturing":"P1✓ P2✓(下修,附DCF)",
 "Zephyr Retail":"❌P1：連續 4 季未更新(stale)",
}
r0=4
for i,rv in enumerate(A["review"]):
    r=r0+i; warn=("❌" in CHK[rv["company"]]); adj=rv["adjustment"]
    W(vr,f"A{r}",rv["company"],Ff(9,True),al=LEF)
    W(vr,f"B{r}",rv["cost"],Fi(9),USD); W(vr,f"C{r}",rv["prior_fv"],Fi(9),USD); W(vr,f"D{r}",rv["reported_fv"],Fi(9),USD)
    W(vr,f"E{r}",rv["method"],Ff(9),al=CEN); W(vr,f"F{r}",CHK[rv["company"]],Ff(8,warn),al=LEF,fl=(WARN if warn else None))
    W(vr,f"G{r}",rv["verdict"],Ff(9,warn),al=CEN,fl=(WARN if warn else OK))
    W(vr,f"H{r}",adj if adj else None,Fi(9,True) if adj else Ff(9),USD,fl=(WARN if adj else None))
    W(vr,f"I{r}",f"=D{r}+N(H{r})",Ff(9,True),USD)
    W(vr,f"J{r}",NOTE[rv["company"]],Ff(8),al=LEF)
rN=r0+len(A["review"])
W(vr,f"A{rN}","合計 Total",Ff(10,True),al=LEF,fl=SUB)
for col in ["B","C","D","H","I"]:
    W(vr,f"{col}{rN}",f"=SUM({col}{r0}:{col}{rN-1})",Ff(10,True),USD,fl=SUB)
W(vr,f"J{rN}","",Ff(8),al=LEF,fl=SUB)
vr.column_dimensions["A"].width=20; vr.column_dimensions["E"].width=12; vr.column_dimensions["F"].width=26
vr.column_dimensions["G"].width=13; vr.column_dimensions["J"].width=30
for c in "BCDHI": vr.column_dimensions[c].width=13
TOT_ROW=rN  # row of totals; adjusted-FV total = I{TOT_ROW}, reported total = D{TOT_ROW}

# ===================== 2. NAV Bridge =====================
nb=sheet("NAV Bridge",0)
nb.merge_cells("A1:C1"); nb["A1"]="NAV 橋接 NAV Bridge · "+A["as_of"]; nb["A1"].font=Fh(12); nb["A1"].fill=fill(HDR)
nb["A1"].alignment=Alignment(horizontal="left",vertical="center")
cash=next(l["amount"] for l in [{"item":"Cash","amount":20000000}] )  # placeholder; replaced below
# pull nav build components straight from analysis inputs (re-read pulls)
with open(os.path.join(HERE,"..","mcp_pulls","03_get_fund_nav_build.json"),encoding="utf-8") as f:
    navlines={l["item"]:l["amount"] for l in json.load(f)["result"]["lines"]}
cash=navlines["Cash"]; other=navlines["Other Assets"]; accrued=navlines["Accrued Liabilities"]
rows=[("投資組合FV（GP報告）",f"='Valuation Review'!D{TOT_ROW}","link"),
 ("現金 Cash",cash,"input"),("其他資產 Other Assets",other,"input"),
 ("應付費用 Accrued Liabilities",accrued,"input"),
 ("＝ NAV（as-reported）",None,"=C4+C5+C6+C7"),
 ("減：Nova 不符政策上調（P2/P3）",f"='Valuation Review'!H{r0+2}","link"),
 ("Zephyr stale（流程flag,未調金額）",0,"input"),
 ("＝ NAV（policy-adjusted）",None,"=C8+C9+C10")]
r=4
for lab,val,how in rows:
    bold=("＝" in lab); W(nb,f"B{r}",lab,Ff(10,bold),al=LEF,fl=(SUB if bold else None))
    if how=="input": W(nb,f"C{r}",val,Fi(10),USD)
    elif how=="link": W(nb,f"C{r}",val,Fl(10),USD)
    else: W(nb,f"C{r}",how,Ff(10,True),USD,fl=(OUT if "policy-adjusted" in lab else (SUB if bold else None)))
    r+=1
W(nb,"B14","檢查：調整後投組",Ff(9),al=LEF); W(nb,"C14",f"='Valuation Review'!I{TOT_ROW}",Fl(9),USD)
nb.column_dimensions["A"].width=2; nb.column_dimensions["B"].width=34; nb.column_dimensions["C"].width=16
NAV_REP_CELL="'NAV Bridge'!C8"; NAV_ADJ_CELL="'NAV Bridge'!C11"

# ===================== 3. Returns =====================
rt=sheet("Returns")
rt.merge_cells("A1:D1"); rt["A1"]="報酬指標 Return Metrics"; rt["A1"].font=Fh(12); rt["A1"].fill=fill(HDR)
rt["A1"].alignment=Alignment(horizontal="left",vertical="center")
W(rt,"A3","項目",Fh(9),al=LEF,fl=HDR); W(rt,"C3","Reported",Fh(9),al=CEN,fl=HDR); W(rt,"D3","Adjusted",Fh(9),al=CEN,fl=HDR)
W(rt,"A4","累計出資 Contributed",Ff(10),al=LEF); W(rt,"C4",A["contributed"],Fi(10),USD); W(rt,"D4","=C4",Ff(10),USD)
W(rt,"A5","累計分配 Distributions",Ff(10),al=LEF); W(rt,"C5",A["distributions"],Fi(10),USD); W(rt,"D5","=C5",Ff(10),USD)
W(rt,"A6","NAV",Ff(10),al=LEF); W(rt,"C6",f"={NAV_REP_CELL}",Fl(10),USD); W(rt,"D6",f"={NAV_ADJ_CELL}",Fl(10),USD)
W(rt,"A7","TVPI",Ff(10,True),al=LEF,fl=SUB); W(rt,"C7","=(C5+C6)/C4",Ff(10,True),MULT,fl=SUB); W(rt,"D7","=(D5+D6)/D4",Ff(10,True),MULT,fl=SUB)
W(rt,"A8","DPI",Ff(10),al=LEF); W(rt,"C8","=C5/C4",Ff(10),MULT); W(rt,"D8","=D5/D4",Ff(10),MULT)
W(rt,"A9","RVPI",Ff(10),al=LEF); W(rt,"C9","=C6/C4",Ff(10),MULT); W(rt,"D9","=D6/D4",Ff(10),MULT)
# cashflow table for XIRR — dates/amounts straight from capital_flows pull
hdr(rt,"A11:D11","現金流 (XIRR)")
W(rt,"A12","日期",Fh(9),al=CEN,fl=HDR); W(rt,"B12","事件",Fh(9),al=LEF,fl=HDR)
W(rt,"C12","Reported",Fh(9),al=CEN,fl=HDR); W(rt,"D12","Adjusted",Fh(9),al=CEN,fl=HDR)
with open(os.path.join(HERE,"..","mcp_pulls","04_get_capital_flows.json"),encoding="utf-8") as f:
    cflows=json.load(f)["result"]
flow_rows=[fl for fl in cflows if fl["event"]!="Ending NAV (as-reported)"]
r=13
for fl in flow_rows:
    y,m,d=(int(x) for x in fl["date"].split("-"))
    W(rt,f"A{r}",datetime.date(y,m,d),Fi(9),DATEF,al=CEN); W(rt,f"B{r}",fl["event"],Ff(9),al=LEF)
    W(rt,f"C{r}",fl["amount"],Fi(9),USD); W(rt,f"D{r}",fl["amount"],Fi(9),USD); r+=1
# ending NAV row links to NAV cells
W(rt,f"A{r}",datetime.date(2026,3,31),Fi(9),DATEF,al=CEN); W(rt,f"B{r}","期末NAV",Ff(9),al=LEF)
W(rt,f"C{r}","=C6",Fl(9),USD); W(rt,f"D{r}","=D6",Fl(9),USD)
last_cf=r
W(rt,f"A{last_cf+2}","Net IRR (XIRR)",Ff(10,True),al=LEF,fl=OUT)
W(rt,f"C{last_cf+2}",f"=XIRR(C13:C{last_cf},A13:A{last_cf})",Ff(10,True),PCT,fl=OUT)
W(rt,f"D{last_cf+2}",f"=XIRR(D13:D{last_cf},A13:A{last_cf})",Ff(10,True),PCT,fl=OUT)
W(rt,f"A{last_cf+3}","vs 8% hurdle",Ff(9),al=LEF)
W(rt,f"D{last_cf+3}",f'=IF(D{last_cf+2}<0.08,"❌ 低於hurdle","✓")',Ff(9,True),al=CEN,fl=WARN)
rt.column_dimensions["A"].width=16; rt.column_dimensions["B"].width=16
rt.column_dimensions["C"].width=15; rt.column_dimensions["D"].width=15

# ===================== 4. Preferred Recompute =====================
pf=sheet("Pref Recompute")
pf.merge_cells("A1:E1"); pf["A1"]="優先報酬獨立重算 Preferred Return Recompute（8% 複利）"; pf["A1"].font=Fh(12); pf["A1"].fill=fill(HDR)
pf["A1"].alignment=Alignment(horizontal="left",vertical="center")
with open(os.path.join(HERE,"..","mcp_pulls","05_get_fund_terms.json"),encoding="utf-8") as f:
    terms=json.load(f)["result"]["terms"]
W(pf,"A3","利率 Rate",Ff(10),al=LEF); W(pf,"B3",terms["preferred_return"],Fi(10),PCT,cmt="get_fund_terms.preferred_return")
W(pf,"A4","估值日 As-of",Ff(10),al=LEF); W(pf,"B4",datetime.date(2026,3,31),Fi(10),DATEF,al=CEN)
W(pf,"A6","現金流",Fh(9),al=LEF,fl=HDR); W(pf,"B6","日期",Fh(9),al=CEN,fl=HDR); W(pf,"C6","金額",Fh(9),al=CEN,fl=HDR)
W(pf,"D6","年數t",Fh(9),al=CEN,fl=HDR); W(pf,"E6","FV@8%",Fh(9),al=CEN,fl=HDR)
# contributions (positive magnitude) + distribution (negative) from capital_flows
pflows=[]
for fl in cflows:
    if fl["event"]=="Contribution": pflows.append(("出資",fl["date"],-fl["amount"]))
    elif fl["event"]=="Distribution": pflows.append(("分配",fl["date"],-fl["amount"]))
r=7
n_contrib=0
for nm,d,amt in pflows:
    y,m,dd=(int(x) for x in d.split("-"))
    W(pf,f"A{r}",nm,Ff(9),al=LEF); W(pf,f"B{r}",datetime.date(y,m,dd),Fi(9),DATEF,al=CEN); W(pf,f"C{r}",amt,Fi(9),USD)
    W(pf,f"D{r}",f"=YEARFRAC(B{r},$B$4)",Ff(9),"0.00",al=CEN)
    W(pf,f"E{r}",f"=C{r}*(1+$B$3)^D{r}",Ff(9),USD)
    if nm=="出資": n_contrib+=1
    r+=1
last_f=r-1
W(pf,f"A{r+1}","FV 出資合計",Ff(10),al=LEF); W(pf,f"E{r+1}",f"=SUM(E7:E{6+n_contrib})",Ff(10),USD)
W(pf,f"A{r+2}","FV 分配（減）",Ff(10),al=LEF); W(pf,f"E{r+2}",f"=E{last_f}",Ff(10),USD)
W(pf,f"A{r+3}","淨出資（出資−分配）",Ff(10),al=LEF); W(pf,f"E{r+3}",A["net_contrib"],Fi(10),USD,cmt="contributed − distributions")
W(pf,f"A{r+4}","＝ 獨立重算優先報酬",Ff(10,True),al=LEF,fl=OUT); W(pf,f"E{r+4}",f"=E{r+1}-E{r+2}-E{r+3}",Ff(10,True),USD,fl=OUT)
W(pf,f"A{r+5}","GP 包提供之 pref",Ff(10),al=LEF); W(pf,f"E{r+5}",terms["preferred_accrued_to_date"],Fi(10),USD,cmt="get_fund_terms.preferred_accrued_to_date")
W(pf,f"A{r+6}","差異（重算 − 提供）",Ff(10,True),al=LEF); W(pf,f"E{r+6}",f"=E{r+4}-E{r+5}",Ff(10,True),USD,fl=WARN)
PREF_INDEP_CELL=f"'Pref Recompute'!E{r+4}"
pf.column_dimensions["A"].width=22
for c in "BCDE": pf.column_dimensions[c].width=14

# ===================== 5. Waterfall (3 scenarios) =====================
wf=sheet("Waterfall")
wf.merge_cells("A1:D1"); wf["A1"]="收益分配 Waterfall（歐式/whole-fund；8% pref、100% catch-up、20% carry）"
wf["A1"].font=Fh(12); wf["A1"].fill=fill(HDR); wf["A1"].alignment=Alignment(horizontal="left",vertical="center")
W(wf,"A3","層級 / 情境",Fh(9),al=LEF,fl=HDR)
W(wf,"B3","A: GP包(報告)",Fh(9),al=CEN,fl=HDR); W(wf,"C3","B: 調整+pref35",Fh(9),al=CEN,fl=HDR); W(wf,"D3","C: 調整+獨立pref",Fh(9),al=CEN,fl=HDR)
W(wf,"A4","總價值（分配+NAV）",Ff(10,True),al=LEF)
W(wf,"B4","=Returns!C5+Returns!C6",Fl(10),USD); W(wf,"C4","=Returns!D5+Returns!D6",Fl(10),USD); W(wf,"D4","=Returns!D5+Returns!D6",Fl(10),USD)
W(wf,"A5","承諾出資 Contributed",Ff(10),al=LEF)
for col in "BCD": W(wf,f"{col}5","=Returns!C4",Fl(10),USD)
W(wf,"A6","優先報酬 Preferred",Ff(10),al=LEF)
W(wf,"B6",terms["preferred_accrued_to_date"],Fi(10),USD); W(wf,"C6",terms["preferred_accrued_to_date"],Fi(10),USD); W(wf,"D6",f"={PREF_INDEP_CELL}",Fl(10),USD)
W(wf,"A8","① 返還出資",Ff(10),al=LEF)
for col in "BCD": W(wf,f"{col}8",f"=MIN({col}4,{col}5)",Ff(10),USD)
W(wf,"A9","  餘 after capital",Ff(9),al=LEF)
for col in "BCD": W(wf,f"{col}9",f"={col}4-{col}8",Ff(9),USD)
W(wf,"A10","② LP 優先報酬(付)",Ff(10),al=LEF)
for col in "BCD": W(wf,f"{col}10",f"=MIN({col}9,{col}6)",Ff(10),USD)
W(wf,"A11","  餘 after pref",Ff(9),al=LEF)
for col in "BCD": W(wf,f"{col}11",f"={col}9-{col}10",Ff(9),USD)
W(wf,"A12","③ GP 追補 catch-up",Ff(10),al=LEF)
for col in "BCD":
    W(wf,f"{col}12",f'=MIN({col}11,IF({col}10>={col}6,0.25*{col}6,0))',Ff(10),USD)
W(wf,"A13","  餘 after catch-up",Ff(9),al=LEF)
for col in "BCD": W(wf,f"{col}13",f"={col}11-{col}12",Ff(9),USD)
W(wf,"A14","④ 80/20 — LP 80%",Ff(10),al=LEF)
for col in "BCD": W(wf,f"{col}14",f"={col}13*0.8",Ff(10),USD)
W(wf,"A15","④ 80/20 — GP 20%",Ff(10),al=LEF)
for col in "BCD": W(wf,f"{col}15",f"={col}13*0.2",Ff(10),USD)
W(wf,"A17","GP Carry 合計",Ff(11,True),al=LEF,fl=OUT)
for col in "BCD": W(wf,f"{col}17",f"={col}12+{col}15",Ff(11,True),USD,fl=OUT)
W(wf,"A18","LP 合計",Ff(11,True),al=LEF,fl=SUB)
for col in "BCD": W(wf,f"{col}18",f"={col}8+{col}10+{col}14",Ff(11,True),USD,fl=SUB)
W(wf,"A19","檢查 LP+GP=總價值",Ff(9),al=LEF)
for col in "BCD": W(wf,f"{col}19",f'=IF(ROUND({col}17+{col}18-{col}4,0)=0,"✓","✗")',Ff(9),al=CEN,fl=OK)
W(wf,"A21","註：情境C用獨立重算pref(~69M);調整後NAV未達 capital+pref → 無 catch-up、carry=0。pref 基準須 CCO 對帳後定案。",Ff(8),al=LEF)
wf.merge_cells("A21:D22")
wf.column_dimensions["A"].width=24
for c in "BCD": wf.column_dimensions[c].width=16

if "Sheet" in wb.sheetnames: wb.remove(wb["Sheet"])
outpath=os.path.normpath(os.path.join(OUTDIR,"Fund_III_Valuation_Review_2026-Q1.xlsx"))
wb.save(outpath)
print("SAVED", outpath, " sheets:", wb.sheetnames)
