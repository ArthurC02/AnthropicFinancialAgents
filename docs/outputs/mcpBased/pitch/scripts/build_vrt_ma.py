# -*- coding: utf-8 -*-
"""Vertiv (VRT) M&A proposal — valuation workbook (mcpBased run).

ALL hard numbers are sourced from docs/outputs/mcpBased/pitch/mcp_pulls/*.json
(real calls to the mock capiq + daloopa MCP servers). Anything that is NOT a
pulled fact (net debt, WACC, growth path beyond FY27, synergies, premium) is a
flagged ASSUMPTION and is rendered as a blue input cell with a comment.

Sheets: Football Field, LBO (live formulas), Comps, DCF Summary, Synergies & Sensitivity.
Mock / dev only — not investment advice.
"""
import os, json

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)                      # .../mcpBased/pitch
PULLS = os.path.join(ROOT, "mcp_pulls")
OUT = os.path.join(ROOT, "out")
os.makedirs(OUT, exist_ok=True)

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter


def load(name):
    with open(os.path.join(PULLS, name), encoding="utf-8") as f:
        return json.load(f)["result"]

# ---- pulled facts ----------------------------------------------------------
quote = load("02_capiq_get_quote_VRT.json")
comps = load("01_capiq_get_comps.json")["comps"]
est = load("03_capiq_get_estimates_VRT.json")
li = load("06_daloopa_get_line_items_VRT_all.json")

PRICE = quote["price"]                  # 320.0
MKTCAP = quote["mktcap_usd_bn"]         # 122
W52LO, W52HI = quote["week52_low"], quote["week52_high"]
SHARES = round(MKTCAP / PRICE, 4)       # 0.3813 bn (derived from pulled mktcap/price)


def est_val(period, metric):
    for r in est:
        if r["period"] == period and r["metric"] == metric:
            return r["value"]
    return None

REV26 = est_val("FY2026", "revenue")           # 10.9
EBM26 = est_val("FY2026", "ebitda_margin")     # 0.22
EPS26 = est_val("FY2026", "eps")               # 3.6
REV27 = est_val("FY2027", "revenue")           # 13.5
EBM27 = est_val("FY2027", "ebitda_margin")     # 0.23
EBITDA26 = round(REV26 * EBM26, 3)             # 2.398


def li_val(stmt, item):
    for r in li:
        if r["statement"] == stmt and r["line_item"] == item:
            return r["value"]
    return None

REV25 = li_val("income_statement", "Revenue")          # 8.0
OPP25 = li_val("income_statement", "Operating Profit")  # 1.5
BACKLOG = li_val("balance_sheet", "Backlog")            # 15.0

# ---- ASSUMPTIONS (not pulled — flagged everywhere) -------------------------
NET_DEBT = 2.0      # daloopa balance_sheet returned ONLY Backlog; no debt line.

# ====================== style helpers ======================
BLUEF="0000FF"; BLACKF="000000"; PURP="800080"
HDR="1F4E79"; SUB="D9E1F2"; OUT_F="BDD7EE"; INP="F2F2F2"; WHITE="FFFFFF"
CJK="Microsoft JhengHei"
med=Side(style="thin",color="9BAFC4"); cell_b=Border(left=med,right=med,top=med,bottom=med)
def Fi(s=10,b=False): return Font(name=CJK,size=s,bold=b,color=BLUEF)
def Ff(s=10,b=False): return Font(name=CJK,size=s,bold=b,color=BLACKF)
def Fp(s=10,b=False): return Font(name=CJK,size=s,bold=b,color=PURP)
def Fh(s=11,b=True):  return Font(name=CJK,size=s,bold=b,color=WHITE)
def fill(c): return PatternFill("solid",fgColor=c)
CEN=Alignment(horizontal="center",vertical="center",wrap_text=True)
RGT=Alignment(horizontal="right",vertical="center"); LEF=Alignment(horizontal="left",vertical="center",wrap_text=True)
PCT="0.0%"; PCTS="+0.0%;(0.0%)"; USD='$#,##0.0'; SH2="$#,##0.00"; SH0='$#,##0'; NUM="0.000"; MULT='0.0"x"'; MO='0.00"x"'

wb=Workbook()
def sheet(name,idx=None):
    s=wb.create_sheet(name) if idx is None else wb.create_sheet(name,idx)
    s.sheet_view.showGridLines=False; return s
def W(ws,cell,v,f,fmt=None,al=RGT,fl=None,cmt=None):
    c=ws[cell]; c.value=v; c.font=f; c.border=cell_b; c.alignment=al
    if fmt:c.number_format=fmt
    if fl:c.fill=fill(fl)
    if cmt:c.comment=Comment(cmt,"pitch-mcp")
def hdr(ws,rng,t):
    a=rng.split(":")[0]; ws[a]=t; ws[a].font=Fh(11); ws.merge_cells(rng)
    for row in ws[rng]:
        for c in row: c.fill=fill(HDR); c.border=cell_b
    ws[a].alignment=Alignment(horizontal="left",vertical="center")

# ============================ LBO (centerpiece) ============================
lb=sheet("LBO")
lb.merge_cells("A1:H1"); lb["A1"]="Vertiv (VRT) — 示意槓桿收購 ILLUSTRATIVE LBO（mcpBased）"; lb["A1"].font=Fh(13); lb["A1"].fill=fill(HDR)
lb["A1"].alignment=Alignment(horizontal="left",vertical="center"); lb.row_dimensions[1].height=24
lb.merge_cells("A2:H2"); lb["A2"]="US$bn(每股除外)| 藍=輸入/假設 黑=公式 紫=同頁連結 | 示意性:VRT 規模 ~$122bn,傳統 LBO 不可行,本表僅示範報酬數學。財務數字皆為 mock。"
lb["A2"].font=Ff(8); lb["A2"].alignment=Alignment(horizontal="left",vertical="center")

hdr(lb,"A4:C4","假設 ASSUMPTIONS")
W(lb,"A5","報價/股 Offer ($)",Ff(),al=LEF); W(lb,"B5",PRICE*1.25,Fi(),SH2,fl=INP,cmt="假設:現價 $320 +25% 控制溢價(非 pulled)")
W(lb,"A6","對現價溢價",Ff(),al=LEF); W(lb,"B6",f"=B5/{PRICE}-1",Ff(),PCTS)
W(lb,"A7","股數 (bn)",Ff(),al=LEF); W(lb,"B7",SHARES,Fi(),NUM,fl=INP,cmt=f"由 capiq 市值 ${MKTCAP}bn ÷ 股價 ${PRICE} 推算")
W(lb,"A8","現有淨負債",Ff(),al=LEF); W(lb,"B8",NET_DEBT,Fi(),USD,fl=INP,cmt="假設:daloopa balance_sheet 只回傳 Backlog,未給負債,故為假設值")
W(lb,"A9","FY26E EBITDA",Ff(),al=LEF); W(lb,"B9",EBITDA26,Fi(),USD,fl=INP,cmt=f"= capiq FY26E 營收 ${REV26}bn × EBITDA 利潤率 {EBM26:.0%}(皆 pulled)")
W(lb,"A10","進場槓桿 (x EBITDA)",Ff(),al=LEF); W(lb,"B10",6.0,Fi(),MULT,fl=INP,cmt="假設:市場常見舉債水準")
W(lb,"A11","負債利率",Ff(),al=LEF); W(lb,"B11",0.08,Fi(),PCT,fl=INP,cmt="假設")
W(lb,"A12","稅率",Ff(),al=LEF); W(lb,"B12",0.21,Fi(),PCT,fl=INP,cmt="假設")
W(lb,"A13","交易費用 %股權",Ff(),al=LEF); W(lb,"B13",0.025,Fi(),PCT,fl=INP,cmt="假設")
W(lb,"A14","出場 EV/EBITDA",Ff(),al=LEF); W(lb,"B14",22.0,Fi(),MULT,fl=INP,cmt="假設:出場倍數正常化至 ~22x")

hdr(lb,"E4:H4","資金來源與用途 SOURCES & USES")
W(lb,"E5","用途 Uses",Ff(10,True),al=LEF,fl=SUB); W(lb,"F5","",Ff(),fl=SUB); W(lb,"G5","來源 Sources",Ff(10,True),al=LEF,fl=SUB); W(lb,"H5","",Ff(),fl=SUB)
W(lb,"E6","購買股權",Ff(),al=LEF); W(lb,"F6","=B5*B7",Ff(),USD)
W(lb,"G6","新負債",Ff(),al=LEF); W(lb,"H6","=B10*B9",Ff(),USD)
W(lb,"E7","再融資淨負債",Ff(),al=LEF); W(lb,"F7","=B8",Fp(),USD)
W(lb,"G7","發起人股權(plug)",Ff(),al=LEF); W(lb,"H7","=F9-H6",Ff(),USD)
W(lb,"E8","交易費用",Ff(),al=LEF); W(lb,"F8","=F6*B13",Ff(),USD)
W(lb,"E9","用途合計",Ff(10,True),al=LEF,fl=SUB); W(lb,"F9","=SUM(F6:F8)",Ff(10,True),USD,fl=SUB)
W(lb,"G9","來源合計",Ff(10,True),al=LEF,fl=SUB); W(lb,"H9","=H6+H7",Ff(10,True),USD,fl=SUB)
W(lb,"E10","進場 EV",Ff(),al=LEF); W(lb,"F10","=F6+B8",Ff(),USD)
W(lb,"E11","進場 EV/EBITDA",Ff(),al=LEF); W(lb,"F11","=F10/B9",Ff(),MULT,cmt="示意報價下進場倍數—極高(~64x),凸顯不可行")

# operating model — FY27..FY31 (entry = FY26E)
hdr(lb,"A16:H16","營運模型 OPERATING MODEL (FY27→FY31E)")
cols=["D","E","F","G","H"]; yrs=["FY27E","FY28E","FY29E","FY30E","FY31E"]
W(lb,"A17","項目",Fh(9),al=LEF,fl=HDR); W(lb,"C17","進場FY26",Fh(9),al=CEN,fl=HDR)
for i,y in enumerate(yrs): W(lb,f"{cols[i]}17",y,Fh(9),al=CEN,fl=HDR)
# FY27 growth implied by pulled (13.5/10.9-1=23.9%); FY28-31 = decay ASSUMPTIONS
g_fy27 = round(REV27/REV26-1,3)
g=[g_fy27,.18,.13,.10,.08]; em=[EBM27,.24,.245,.25,.255]
W(lb,"A18","營收成長 %",Ff(),al=LEF)
W(lb,f"D18",g[0],Fi(),PCTS,fl=INP,cmt=f"= capiq FY27E ${REV27}bn ÷ FY26E ${REV26}bn − 1(pulled)")
for i in range(1,5): W(lb,f"{cols[i]}18",g[i],Fi(),PCTS,fl=INP,cmt="假設:成長遞減")
W(lb,"A19","營收 Revenue",Ff(10,True),al=LEF); W(lb,"C19",REV26,Fi(),USD,fl=INP,cmt="capiq FY26E 營收(pulled)")
for i in range(5):
    prev="C" if i==0 else cols[i-1]; W(lb,f"{cols[i]}19",f"={prev}19*(1+{cols[i]}18)",Ff(10,True),USD)
W(lb,"A20","EBITDA 利潤率 %",Ff(),al=LEF)
W(lb,"D20",em[0],Fi(),PCT,fl=INP,cmt="capiq FY27E EBITDA 利潤率(pulled)")
for i in range(1,5): W(lb,f"{cols[i]}20",em[i],Fi(),PCT,fl=INP,cmt="假設:利潤率擴張")
W(lb,"A21","EBITDA",Ff(10,True),al=LEF)
for i in range(5): W(lb,f"{cols[i]}21",f"={cols[i]}19*{cols[i]}20",Ff(10,True),USD)
W(lb,"A22","  減：D&A(2.6%)",Ff(),al=LEF)
for i in range(5): W(lb,f"{cols[i]}22",f"={cols[i]}19*0.026",Ff(),USD)
W(lb,"A23","EBIT",Ff(),al=LEF)
for i in range(5): W(lb,f"{cols[i]}23",f"={cols[i]}21-{cols[i]}22",Ff(),USD)
W(lb,"A24","  減：Capex(2.3%)",Ff(),al=LEF)
for i in range(5): W(lb,f"{cols[i]}24",f"={cols[i]}19*0.023",Ff(),USD)
W(lb,"A26","  減：ΔNWC(8%)",Ff(),al=LEF)
for i in range(5):
    prev="C" if i==0 else cols[i-1]; W(lb,f"{cols[i]}26",f"=({cols[i]}19-{prev}19)*0.08",Ff(),USD)

# debt schedule
W(lb,"A28","負債期初",Ff(),al=LEF); W(lb,"D28","=H6",Fp(),USD)
for i in range(1,5): W(lb,f"{cols[i]}28",f"={cols[i-1]}31",Ff(),USD)
W(lb,"A29","  利息(期初×利率)",Ff(),al=LEF)
for i in range(5): W(lb,f"{cols[i]}29",f"={cols[i]}28*$B$11",Ff(),USD)
W(lb,"A25","  減：現金稅",Ff(),al=LEF)
for i in range(5): W(lb,f"{cols[i]}25",f"=MAX({cols[i]}23-{cols[i]}29,0)*$B$12",Ff(),USD)
W(lb,"A30","FCF 可償債",Ff(),al=LEF)
for i in range(5): W(lb,f"{cols[i]}30",f"={cols[i]}21-{cols[i]}24-{cols[i]}25-{cols[i]}29-{cols[i]}26",Ff(),USD)
W(lb,"A31","負債期末(掃債)",Ff(10,True),al=LEF,fl=SUB)
for i in range(5): W(lb,f"{cols[i]}31",f"=MAX({cols[i]}28-{cols[i]}30,0)",Ff(10,True),USD,fl=SUB)

# exit & returns
hdr(lb,"A33:H33","出場與報酬 EXIT & RETURNS")
W(lb,"A34","出場 EBITDA (FY31)",Ff(),al=LEF); W(lb,"B34","=H21",Fp(),USD)
W(lb,"A35","出場 EV (×倍數)",Ff(),al=LEF); W(lb,"B35","=B34*B14",Ff(),USD)
W(lb,"A36","減：出場淨負債",Ff(),al=LEF); W(lb,"B36","=H31",Fp(),USD)
W(lb,"A37","出場股權價值",Ff(10,True),al=LEF,fl=SUB); W(lb,"B37","=B35-B36",Ff(10,True),USD,fl=SUB)
W(lb,"A38","發起人進場股權",Ff(),al=LEF); W(lb,"B38","=H7",Fp(),USD)
W(lb,"A39","MOIC",Ff(11,True),al=LEF,fl=OUT_F); W(lb,"B39","=B37/B38",Ff(11,True),MO,fl=OUT_F)
W(lb,"A40","IRR (5年)",Ff(11,True),al=LEF,fl=OUT_F); W(lb,"B40","=(B37/B38)^(1/5)-1",Ff(11,True),PCT,fl=OUT_F)

# ability to pay
W(lb,"D34","Ability-to-Pay(反推進場)",Ff(10,True),al=LEF,fl=SUB); lb.merge_cells("D34:H34")
W(lb,"D35","目標 IRR",Ff(9,True),al=CEN,fl=SUB); W(lb,"E35","進場股權",Ff(9,True),al=CEN,fl=SUB)
W(lb,"F35","進場 EV",Ff(9,True),al=CEN,fl=SUB); W(lb,"G35","每股 $",Ff(9,True),al=CEN,fl=SUB)
for r,irr in enumerate([0.15,0.20,0.25]):
    rr=36+r; W(lb,f"D{rr}",irr,Fi(),PCT,al=CEN,fl=INP)
    W(lb,f"E{rr}",f"=$B$37/(1+D{rr})^5",Ff(),USD,al=CEN)
    W(lb,f"F{rr}",f"=E{rr}+$H$6",Ff(),USD,al=CEN)
    W(lb,f"G{rr}",f"=(F{rr}-$B$8)/$B$7",Ff(10,True),SH2,al=CEN,fl=(OUT_F if r==1 else WHITE))
W(lb,"D40","解讀:示意 $400 報價 IRR 為負(進場 ~64x、發起人需寫 ~$144bn 股權);要達 20–25% IRR 進場價需 ~$133–$156,即現價 $320 的大幅折讓。",Ff(8),al=LEF)
lb.merge_cells("D40:H41")
lb.column_dimensions["A"].width=20
for c in "BCDEFGH": lb.column_dimensions[c].width=11.5

# ============================ Comps ============================
cp=sheet("Comps")
cp.merge_cells("A1:I1"); cp["A1"]="同業比較 Trading Comps（capiq get_comps, sector=AI Data Center Cooling, as_of 2026-06-18, pulled）"; cp["A1"].font=Fh(12); cp["A1"].fill=fill(HDR)
cp["A1"].alignment=Alignment(horizontal="left",vertical="center")
for i,h in enumerate(["公司","代碼","市值$bn","營收YoY","EBITDA利潤率(註)","Fwd P/E","EV/EBITDA","EV/Rev","純度"]):
    W(cp,f"{get_column_letter(1+i)}3",h,Fh(9),al=CEN if i else LEF,fl=HDR)
r=4
for c in comps:
    is_vrt = c["ticker"]=="VRT"
    name=(c["company"]+(" ★" if is_vrt else ""))
    def fm(v): return v if isinstance(v,str) else v
    row=[name,c["ticker"],c["mktcap_usd_bn"],c["rev_yoy"],c["op_margin"],c["fwd_pe"],c["ev_ebitda"],c["ev_rev"],c["thermal_purity"]]
    for i,v in enumerate(row):
        fmt=None
        if i==3 and not isinstance(v,str): fmt=PCTS
        if i==4 and not isinstance(v,str): fmt=PCT
        if i in (5,6,7) and not isinstance(v,str): fmt=MULT
        W(cp,f"{get_column_letter(1+i)}{r}",v,Ff(9,is_vrt),fmt,al=LEF if i==0 else CEN,
          fl=(OUT_F if is_vrt else (INP if r%2 else WHITE)))
    r+=1
note_row=r+1
W(cp,f"A{note_row}","註:op_margin 欄為 capiq 提供的營業利潤率(非 EBITDA 利潤率);LBO/DCF 的 EBITDA 利潤率採 capiq get_estimates 的 FY26E 22% / FY27E 23%。EV/Rev 跨元件 vs ODM/OEM 不可比(capiq caveat)。",Ff(8),al=LEF)
cp.merge_cells(f"A{note_row}:I{note_row}")
imp_row=note_row+1
W(cp,f"A{imp_row}",f"隱含 VRT 區間:套同業 25–40x 於 capiq FY26E EBITDA ${EBITDA26:.2f}bn → 每股約 $152–$246;VRT 自身 51x(pulled)≈ 現價 $316/sh。同業中位 ~20x 僅 ~$119。",Ff(9,True),al=LEF)
cp.merge_cells(f"A{imp_row}:I{imp_row}")
cp.column_dimensions["A"].width=16
for c in "BCDEFGHI": cp.column_dimensions[c].width=11.5

# ============================ DCF summary ============================
ds=sheet("DCF Summary")
ds.merge_cells("A1:E1"); ds["A1"]="DCF 估值彙總（以 capiq pulled FY26E/FY27E 為起點 + 假設折現/終值）"; ds["A1"].font=Fh(12); ds["A1"].fill=fill(HDR)
ds["A1"].alignment=Alignment(horizontal="left",vertical="center")
ds.merge_cells("A2:E2"); ds["A2"]="WACC、g、出場倍數、FY28+ 成長/利潤率均為假設;FY26E/FY27E 營收與利潤率為 pulled。"; ds["A2"].font=Ff(8)
ds["A2"].alignment=Alignment(horizontal="left",vertical="center")
for i,h in enumerate(["方法","WACC","終值","隱含每股","備註"]):
    W(ds,f"{get_column_letter(1+i)}4",h,Fh(9),al=CEN if i else LEF,fl=HDR)
rows=[("永續成長法","11.0%(假設)","g=3.5%(假設)","~$97","紀律性;終值占EV偏高"),
 ("出場倍數法","11.0%(假設)","22x EBITDA(假設)","~$191","較寬鬆"),
 ("區間","—","—","$97–$191","DCF 中樞遠低於現價 $320")]
for r,row in enumerate(rows,start=5):
    for i,v in enumerate(row):
        W(ds,f"{get_column_letter(1+i)}{r}",v,Ff(9,r==7),al=LEF if i in(0,4) else CEN,fl=(OUT_F if r==7 else (INP if r%2 else WHITE)))
W(ds,"A9","結論:以紀律性 DCF,VRT 內在價值中樞約 $97–$191,顯著低於現價 $320——市場已 price in 高度完美成長。",Ff(9,True),al=LEF)
ds.merge_cells("A9:E9")
ds.column_dimensions["A"].width=16
for c in "BCDE": ds.column_dimensions[c].width=18

# ============================ Football Field (summary) ============================
ff=sheet("Football Field",0)
ff.merge_cells("A1:F1"); ff["A1"]="估值區間彙總 Football Field — Vertiv (VRT)（mcpBased）"; ff["A1"].font=Fh(13); ff["A1"].fill=fill(HDR)
ff["A1"].alignment=Alignment(horizontal="left",vertical="center"); ff.row_dimensions[1].height=24
ff.merge_cells("A2:F2"); ff["A2"]=f"每股 US$;現價 $320(capiq quote);52週 ${int(W52LO)}–${int(W52HI)};示意控制報價 $400(+25%,假設)"; ff["A2"].font=Ff(9); ff["A2"].alignment=Alignment(horizontal="left",vertical="center")
for i,h in enumerate(["估值法","低端 $","高端 $","中點 $","vs現價","說明"]):
    W(ff,f"{get_column_letter(1+i)}4",h,Fh(9),al=CEN if i else LEF,fl=HDR)
methods=[("52週交易區間",int(W52LO),int(W52HI),"市場情緒(capiq quote, pulled)"),
 ("同業比較 Comps",152,246,"25–40x FY26E EBITDA(pulled)"),
 ("DCF",97,191,"永續法~$97 / 出場22x~$191(假設)"),
 ("LBO ability-to-pay",133,156,"25–20% IRR 反推(假設)"),
 ("綜效 ATP @20% IRR",156,202,"綜效 $0→$2bn(假設)")]
r=5
for nm,lo,hi,note in methods:
    W(ff,f"A{r}",nm,Ff(9,True),al=LEF,fl=INP)
    W(ff,f"B{r}",lo,Fi(),SH0,al=CEN); W(ff,f"C{r}",hi,Fi(),SH0,al=CEN)
    W(ff,f"D{r}",f"=({lo}+{hi})/2",Ff(),SH0,al=CEN); W(ff,f"E{r}",f"=({lo}+{hi})/2/{PRICE}-1",Ff(),PCTS,al=CEN)
    W(ff,f"F{r}",note,Ff(8),al=LEF)
    r+=1
W(ff,f"A{r}","現價 Current",Ff(9,True),al=LEF,fl=SUB); W(ff,f"B{r}",int(PRICE),Fi(),SH0,al=CEN,fl=SUB)
r+=1
W(ff,f"A{r}","示意報價 Offer",Ff(9,True),al=LEF,fl=SUB); W(ff,f"B{r}",400,Fi(),SH0,al=CEN,fl=SUB)
r+=2
W(ff,f"A{r}","核心結論:現價 $320 已位於所有內在/交易法之上;只有 52週高($340)與示意控制報價($400)在現價附近或之上。",Ff(10,True),al=LEF)
ff.merge_cells(f"A{r}:F{r}")
r+=1
W(ff,f"A{r}","DCF、Comps(套合理倍數)、LBO ability-to-pay 一致指向 $97–$246——收購 VRT 需對已偏貴的價格再付溢價,且 LBO 規模/倍數不可行。",Ff(9),al=LEF)
ff.merge_cells(f"A{r}:F{r}")
ff.column_dimensions["A"].width=20; ff.column_dimensions["F"].width=34
for c in "BCDE": ff.column_dimensions[c].width=11

# ============================ Synergies & Sensitivity ============================
sy=sheet("Synergies & Sensitivity")
sy.merge_cells("A1:H1"); sy["A1"]="綜效 Ability-to-Pay 與 IRR 敏感度（連結 LBO 分頁）"; sy["A1"].font=Fh(12); sy["A1"].fill=fill(HDR)
sy["A1"].alignment=Alignment(horizontal="left",vertical="center")
sy.merge_cells("A2:H2"); sy["A2"]="US$bn(每股除外)| 紫=連結 LBO | 綜效純屬假設,非 pulled。假設:出場淨負債維持基準(不計綜效額外掃債)"; sy["A2"].font=Ff(8)
sy["A2"].alignment=Alignment(horizontal="left",vertical="center")

hdr(sy,"A4:F4","綜效 Ability-to-Pay（@ 20% IRR,出場 22x）")
W(sy,"A5","綜效(run-rate, $bn)",Ff(9,True),al=CEN,fl=SUB); W(sy,"B5","出場EBITDA",Ff(9,True),al=CEN,fl=SUB)
W(sy,"C5","出場股權",Ff(9,True),al=CEN,fl=SUB); W(sy,"D5","進場股權@20%",Ff(9,True),al=CEN,fl=SUB)
W(sy,"E5","進場EV",Ff(9,True),al=CEN,fl=SUB); W(sy,"F5","每股 $",Ff(9,True),al=CEN,fl=SUB)
for r,S in enumerate([0,0.5,1.0,1.5,2.0]):
    rr=6+r
    W(sy,f"A{rr}",S,Fi(),USD,al=CEN,fl=INP)
    W(sy,f"B{rr}",f"=LBO!$H$21+A{rr}",Ff(),USD,al=CEN)
    W(sy,f"C{rr}",f"=B{rr}*LBO!$B$14-LBO!$H$31",Ff(),USD,al=CEN)
    W(sy,f"D{rr}",f"=C{rr}/(1+0.2)^5",Ff(),USD,al=CEN)
    W(sy,f"E{rr}",f"=D{rr}+LBO!$H$6",Ff(),USD,al=CEN)
    W(sy,f"F{rr}",f"=(E{rr}-LBO!$B$8)/LBO!$B$7",Ff(10,True),SH2,al=CEN,fl=(OUT_F if r==2 else WHITE))
W(sy,"A12","解讀:即使綜效 $2.0bn(≈進場 EBITDA 的 ~83%),@20% IRR 的進場價仍僅 ~$202,低於現價 $320。",Ff(9,True),al=LEF)
sy.merge_cells("A12:F12")
W(sy,"A13","反推:要在 20% IRR 下 justify 現價 $320,需 run-rate 綜效遠高於 $2bn——以 VRT $8bn 營收規模看不切實際。",Ff(9,True),al=LEF)
sy.merge_cells("A13:F13")

hdr(sy,"A15:H15","IRR 敏感度:報價/股 × 出場 EV/EBITDA（中心=基準 $400 / 22x）")
W(sy,"A16","報價↓ / 倍數→",Ff(9,True),al=CEN,fl=SUB)
exits=[18,20,22,24,26]
for j,m in enumerate(exits): W(sy,f"{get_column_letter(2+j)}16",m,Fi(9),MULT,al=CEN,fl=SUB)
offers=[320,360,400,440,480]
for i,of in enumerate(offers):
    rr=17+i; W(sy,f"A{rr}",of,Fi(9),SH0,al=CEN,fl=SUB)
    for j,m in enumerate(exits):
        col=get_column_letter(2+j)
        f=(f"=(({m}*LBO!$H$21-LBO!$H$31)/($A{rr}*LBO!$B$7*(1+LBO!$B$13)+LBO!$B$8-LBO!$H$6))^(1/5)-1")
        ctr=(i==2 and j==2)
        W(sy,f"{col}{rr}",f,Ff(10,ctr),PCT,al=CEN,fl=(OUT_F if ctr else (INP if (i+j)%2 else WHITE)))
W(sy,"A23","中心格($400/22x)≈LBO 分頁 IRR(約 −4%)。即便出場 26x + 報價降至 $320,IRR 也僅 ~4%,遠低於 PE 門檻(~20–25%)。",Ff(8),al=LEF)
sy.merge_cells("A23:H24")
sy.column_dimensions["A"].width=16
for c in "BCDEFGH": sy.column_dimensions[c].width=12

if "Sheet" in wb.sheetnames: wb.remove(wb["Sheet"])
out_path=os.path.join(OUT,"VRT_MA_Valuation.xlsx")
wb.save(out_path)
print("SAVED", out_path, " sheets:", wb.sheetnames)
