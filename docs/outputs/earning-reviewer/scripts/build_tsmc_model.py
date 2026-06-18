# -*- coding: utf-8 -*-
"""TSMC (2330.TW / TSM) Q1 2026 estimate model — new build (no prior model existed).
Anchored on Q1'26 ACTUAL + Q2'26 GUIDANCE; H2'26 = flagged estimates."""
import os
os.makedirs("./out", exist_ok=True)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter

BLUEF="0000FF"; BLACKF="000000"; GREENF="008000"
HDR="1F4E79"; SUB="D9E1F2"; OUT="BDD7EE"; INP="F2F2F2"; WHITE="FFFFFF"; GOOD="E2EFDA"; BAD="FCE4E4"
CJK="Microsoft JhengHei"
med=Side(style="thin",color="9BAFC4"); cell_b=Border(left=med,right=med,top=med,bottom=med)
def Fi(s=10,b=False): return Font(name=CJK,size=s,bold=b,color=BLUEF)
def Ff(s=10,b=False): return Font(name=CJK,size=s,bold=b,color=BLACKF)
def Fh(s=11,b=True):  return Font(name=CJK,size=s,bold=b,color=WHITE)
def fill(c): return PatternFill("solid",fgColor=c)
CEN=Alignment(horizontal="center",vertical="center",wrap_text=True)
RGT=Alignment(horizontal="right",vertical="center"); LEF=Alignment(horizontal="left",vertical="center",wrap_text=True)
PCT="0.0%"; PCTS="+0.0%;(0.0%)"; USD='$#,##0.0'; NTD='#,##0'; EPS='#,##0.00'; BPS='+0"bps";(0"bps")'; FX="0.00"

wb=Workbook()
# ===================== Sheet 1: Model =====================
ws=wb.active; ws.title="Model"; ws.sheet_view.showGridLines=False
def put(cell,v,f,fmt=None,al=RGT,fl=None,cmt=None):
    c=ws[cell]; c.value=v; c.font=f; c.border=cell_b; c.alignment=al
    if fmt:c.number_format=fmt
    if fl:c.fill=fill(fl)
    if cmt:c.comment=Comment(cmt,"earnings-reviewer")
    return c
def hdr(rng,t):
    a=rng.split(":")[0]; ws[a]=t; ws[a].font=Fh(11); ws.merge_cells(rng)
    for row in ws[rng]:
        for c in row: c.fill=fill(HDR); c.border=cell_b
    ws[a].alignment=Alignment(horizontal="left",vertical="center")

ws.merge_cells("A1:H1"); ws["A1"]="台積電 TSMC (2330.TW / TSM) — Q1 2026 季度估計模型"; ws["A1"].font=Fh(14); ws["A1"].fill=fill(HDR)
ws["A1"].alignment=Alignment(horizontal="left",vertical="center"); ws.row_dimensions[1].height=26
ws.merge_cells("A2:H2"); ws["A2"]="實際發布 2026-04-16 | 營收/利潤率以 US$ 計;EPS 以 NT$（含 ADR US$）| 藍=輸入(實際/指引/假設) 黑=公式 | H2'26 為估計"
ws["A2"].font=Ff(8); ws["A2"].alignment=Alignment(horizontal="left",vertical="center")

cols=["B","C","D","E","F"]   # Q1'26A, Q2'26E, Q3'26E, Q4'26E, FY2026E
labels=["Q1'26A","Q2'26E","Q3'26E","Q4'26E","FY2026E"]
hdr("A4:F4","損益預測 INCOME STATEMENT")
put("A5","項目",Fh(9),al=LEF,fl=HDR)
for i,l in enumerate(labels):
    put(f"{cols[i]}5",l,Fh(9),al=CEN,fl=HDR)

# Revenue US$bn
put("A6","營收 Revenue (US$bn)",Ff(10,True),al=LEF)
put("B6",35.90,Fi(10,True),USD,fl=INP,cmt="ACTUAL Q1'26: US$35.90bn (NT$1,134.10bn), +40.6% YoY, +6.4% QoQ. Source: TSMC 6-K 2026-04-16")
put("C6",39.60,Fi(10,True),USD,fl=INP,cmt="Q2'26 GUIDANCE midpoint of US$39.0-40.2bn. Source: TSMC Q1'26 call")
put("D6",42.77,Fi(10,True),USD,fl=INP,cmt="ESTIMATE: +8.0% QoQ (seasonal AI/HPC strength)")
put("E6",45.33,Fi(10,True),USD,fl=INP,cmt="ESTIMATE: +6.0% QoQ")
put("F6","=SUM(B6:E6)",Ff(10,True),USD)
# QoQ
put("A7","  QoQ %",Ff(),al=LEF)
put("B7",0.064,Fi(),PCTS,fl=INP,cmt="ACTUAL +6.4% QoQ")
for i,c in enumerate(["C","D","E"]):
    prev=["B","C","D"][i]; put(f"{c}7",f"={c}6/{prev}6-1",Ff(),PCTS)
put("F7","=F6/124.0-1",Ff(),PCTS,cmt="vs FY2025 base ~US$124bn (approx; verify) → guide '>30%'")
# YoY memo
put("A8","  YoY %",Ff(),al=LEF)
put("B8",0.406,Fi(),PCTS,fl=INP,cmt="ACTUAL +40.6% YoY (US$)")
put("F8","=F7",Ff(),PCTS)
# Gross margin
put("A9","毛利率 Gross Margin %",Ff(10,True),al=LEF)
put("B9",0.662,Fi(10,True),PCT,fl=INP,cmt="ACTUAL 66.2% (+390bps QoQ, +740bps YoY); beat top of 63-65% guide by ~120bps")
put("C9",0.665,Fi(10,True),PCT,fl=INP,cmt="Q2 GUIDANCE midpoint 65.5-67.5%")
put("D9",0.655,Fi(10,True),PCT,fl=INP,cmt="ESTIMATE: N2 ramp + overseas dilution begins H2'26")
put("E9",0.650,Fi(10,True),PCT,fl=INP,cmt="ESTIMATE: wider N2/overseas dilution")
put("F9","=SUMPRODUCT(B6:E6,B9:E9)/F6",Ff(10,True),PCT)
# Gross profit
put("A10","毛利 Gross Profit (US$bn)",Ff(),al=LEF)
for c in cols[:4]: put(f"{c}10",f"={c}6*{c}9",Ff(),USD)
put("F10","=F6*F9",Ff(),USD)
# Operating margin
put("A11","營業利益率 Op Margin %",Ff(10,True),al=LEF)
put("B11",0.581,Fi(10,True),PCT,fl=INP,cmt="ACTUAL 58.1%")
put("C11",0.575,Fi(10,True),PCT,fl=INP,cmt="Q2 GUIDANCE midpoint 56.5-58.5%")
put("D11",0.568,Fi(10,True),PCT,fl=INP,cmt="ESTIMATE")
put("E11",0.563,Fi(10,True),PCT,fl=INP,cmt="ESTIMATE")
put("F11","=SUMPRODUCT(B6:E6,B11:E11)/F6",Ff(10,True),PCT)
# FX
put("A12","匯率 USD/TWD",Ff(),al=LEF)
put("B12",31.6,Fi(),FX,fl=INP,cmt="ACTUAL implied (NT$1,134.10/US$35.90)")
for c in ["C","D","E"]: put(f"{c}12",31.7,Fi(),FX,fl=INP,cmt="Q2 guide FX assumption 31.7")
put("F12","=SUMPRODUCT(B6:E6,B12:E12)/F6",Ff(),FX)
# NT$ revenue
put("A13","營收 Revenue (NT$bn)",Ff(),al=LEF)
for c in cols[:4]: put(f"{c}13",f"={c}6*{c}12",Ff(),NTD)
put("F13","=SUM(B13:E13)",Ff(),NTD)
# Net margin
put("A14","淨利率 Net Margin %",Ff(10,True),al=LEF)
put("B14",0.505,Fi(10,True),PCT,fl=INP,cmt="ACTUAL 50.5%")
put("C14",0.495,Fi(10,True),PCT,fl=INP,cmt="ESTIMATE")
put("D14",0.490,Fi(10,True),PCT,fl=INP,cmt="ESTIMATE")
put("E14",0.485,Fi(10,True),PCT,fl=INP,cmt="ESTIMATE")
put("F14","=SUMPRODUCT(B13:E13,B14:E14)/F13",Ff(10,True),PCT)
# Net income NT$
put("A15","淨利 Net Income (NT$bn)",Ff(10,True),al=LEF,fl=SUB)
for c in cols[:4]: put(f"{c}15",f"={c}13*{c}14",Ff(10,True),NTD,fl=SUB)
put("F15","=SUM(B15:E15)",Ff(10,True),NTD,fl=SUB)
# Shares
put("A16","流通股數 Shares (bn)",Ff(),al=LEF)
for c in cols: put(f"{c}16",25.93,Fi(),"0.00",fl=INP,cmt="~25.93bn ordinary shares; 1 ADR = 5 shares" if c=="B" else None)
# EPS NT$
put("A17","EPS (NT$)",Ff(11,True),al=LEF,fl=OUT)
for c in cols[:4]: put(f"{c}17",f"={c}15/{c}16",Ff(11,True),EPS,fl=OUT)
put("F17","=F15/B16",Ff(11,True),EPS,fl=OUT)
# EPS ADR
put("A18","EPS / ADR (US$)",Ff(10,True),al=LEF)
for c in cols[:4]: put(f"{c}18",f"={c}17*5/{c}12",Ff(10,True),EPS)
put("F18","=F17*5/F12",Ff(10,True),EPS)

# Mix block
hdr("A20:F20","Q1'26 營收組合 REVENUE MIX (ACTUAL)")
put("A21","平台 By Platform",Ff(10,True),al=LEF,fl=SUB)
mixp=[("HPC",0.61),("智慧型手機 Smartphone",0.26),("IoT",0.06),("車用 Auto",0.04),("DCE",0.01),("其他 Other",0.02)]
r=22
for nm,v in mixp:
    put(f"A{r}",f"  {nm}",Ff(),al=LEF); put(f"B{r}",v,Fi(),PCT,fl=INP); r+=1
put("D21","技術節點 By Node (% wafer rev)",Ff(10,True),al=LEF,fl=SUB)
mixn=[("3nm (N3)",0.25),("5nm (N5)",0.36),("7nm",0.13),("先進(≤7nm) Advanced",0.74)]
r=22
for nm,v in mixn:
    put(f"D{r}",f"  {nm}",Ff(),al=LEF); put(f"E{r}",v,Fi(),PCT,fl=INP); r+=1

# widths
ws.column_dimensions["A"].width=26
for c in "BCDEF": ws.column_dimensions[c].width=12

# ===================== Sheet 2: Variance =====================
vs=wb.create_sheet("Actual vs Cons vs Prior"); vs.sheet_view.showGridLines=False
def vput(cell,v,f,fmt=None,al=RGT,fl=None,cmt=None):
    c=vs[cell]; c.value=v; c.font=f; c.border=cell_b; c.alignment=al
    if fmt:c.number_format=fmt
    if fl:c.fill=fill(fl)
    if cmt:c.comment=Comment(cmt,"earnings-reviewer")
    return c
def vhdr(rng,t):
    a=rng.split(":")[0]; vs[a]=t; vs[a].font=Fh(11); vs.merge_cells(rng)
    for row in vs[rng]:
        for c in row: c.fill=fill(HDR); c.border=cell_b
    vs[a].alignment=Alignment(horizontal="left",vertical="center")
vs.merge_cells("A1:F1"); vs["A1"]="Q1 2026:實際 vs. 市場預期 vs. 前次指引"; vs["A1"].font=Fh(13); vs["A1"].fill=fill(HDR)
vs["A1"].alignment=Alignment(horizontal="left",vertical="center"); vs.row_dimensions[1].height=24

vhdr("A3:F3","Q1'26 結果 RESULTS")
heads=["指標 Metric","前次指引\n(Jan'26)","市場預期\nConsensus","實際\nActual","實際 vs 指引","實際 vs 預期"]
for i,h in enumerate(heads):
    vput(f"{get_column_letter(1+i)}4",h,Fh(9),al=CEN if i else LEF,fl=HDR)
rows=[
 ("營收 Revenue (US$bn)","34.6–35.8","35.5","35.90","高於上緣","+1.1%"),
 ("毛利率 Gross Margin","63–65%","~64%","66.2%","+120bps(vs上緣)","+~220bps"),
 ("營業利益率 Op Margin","—","—","58.1%","高於指引","—"),
 ("EPS (NT$)","(無指引)","20.88","22.08","—","+5.7%"),
 ("EPS / ADR (US$)","—","3.26","3.49","—","+7.1%"),
 ("淨利 Net Income (NT$bn)","—","—","572.48","—","+58% YoY"),
]
r=5
for row in rows:
    for i,val in enumerate(row):
        fl=None
        if i==4 and ("高於" in str(val) or "+" in str(val)): fl=GOOD
        if i==5 and str(val).startswith("+"): fl=GOOD
        vput(f"{get_column_letter(1+i)}{r}",val,Ff(10, i==0),al=LEF if i==0 else CEN,fl=fl or (SUB if i==0 else WHITE))
    r+=1

vhdr(f"A{r+1}:F{r+1}","Q2'26 指引 vs. 市場預期 GUIDANCE vs CONSENSUS")
r+=2
q2=[("營收 Revenue (US$bn)","指引 39.0–40.2(中 39.6)","預期 ~38.1","+3.9% 高於預期 ▲"),
    ("毛利率 Gross Margin","指引 65.5–67.5%","—","強勁 / 高於"),
    ("營業利益率 Op Margin","指引 56.5–58.5%","—","—")]
vput(f"A{r}","指標",Fh(9),al=LEF,fl=HDR); vput(f"B{r}","台積電指引",Fh(9),al=CEN,fl=HDR)
vput(f"C{r}","市場預期",Fh(9),al=CEN,fl=HDR); vs.merge_cells(f"D{r}:F{r}"); vput(f"D{r}","評語",Fh(9),al=CEN,fl=HDR)
r+=1
for row in q2:
    vput(f"A{r}",row[0],Ff(10,True),al=LEF,fl=SUB); vput(f"B{r}",row[1],Ff(),al=CEN)
    vput(f"C{r}",row[2],Ff(),al=CEN); vs.merge_cells(f"D{r}:F{r}")
    vput(f"D{r}",row[3],Ff(10,True),al=CEN,fl=GOOD if "▲" in row[3] else WHITE); r+=1

vhdr(f"A{r+1}:F{r+1}","全年 2026 展望 FY2026 OUTLOOK")
r+=2
fy=[("營收成長(US$)","“> 30%”（AI 需求驅動,維持/上調）"),
    ("2026 資本支出","US$52–56bn,偏向高端(創高);~70–80% 先進製程"),
    ("長期毛利率底線","上調至 “56% 以上 through the cycle”"),
    ("長期 AI 加速器營收 CAGR","上調至 ~54–56% 至 2029")]
for nm,v in fy:
    vput(f"A{r}",nm,Ff(10,True),al=LEF,fl=SUB); vs.merge_cells(f"B{r}:F{r}"); vput(f"B{r}",v,Ff(),al=LEF); r+=1

vs.column_dimensions["A"].width=24
for c in "BCDEF": vs.column_dimensions[c].width=15

# ===================== Sheet 3: Notes =====================
ns=wb.create_sheet("Assumptions & Sources"); ns.sheet_view.showGridLines=False
ns.column_dimensions["A"].width=3; ns.column_dimensions["B"].width=115
ns.merge_cells("A1:B1"); ns["A1"]="假設、注意事項與來源"; ns["A1"].font=Fh(13); ns["A1"].fill=fill(HDR)
ns.row_dimensions[1].height=24
items=[("H","建模假設"),
 ("","Q1'26 全部為公司公布實際數;Q2'26 採法說會指引中點;Q3–Q4'26 為估計(營收 +8%/+6% QoQ、毛利率隨 N2 與海外廠稀釋下行)。"),
 ("","FY2025 營收基數 ~US$124bn 為近似值(用於 YoY 與『>30%』檢核),正式發布前請以 TSMC 官方 FY25 數字校正。"),
 ("","匯率:Q1 實際隱含 31.6;Q2–Q4 採台積電 Q2 指引假設 31.7。股數 ~25.93bn;1 ADR = 5 股。"),
 ("","H2'26 毛利率下行反映管理層明示:N2 量產 + 海外廠全年稀釋毛利率約 2–3%(海外早期 2–3%、後期擴大至 3–4%)。"),
 ("H","關鍵警示"),
 ("","本模型為新建(工作區先前無台積電模型)。『前次估計』以『前次官方指引(Jan'26 設定)』與市場 consensus 代表。"),
 ("","Consensus、股價反應(~-3.1%)、BofA 目標價(NT$2,360→2,530)等為彙整自第三方,發布前請以 FactSet/Bloomberg 校正。"),
 ("","地區別營收未揭露(官方 6-K 地理表 SEC 連結 403);N2 單獨佔比未拆分。"),
 ("H","主要來源"),
 ("","TSMC 6-K 新聞稿 & Q1'26 法說會 transcript(2026-04-16):Q1 實際、Q2 指引、FY 展望、資本支出、稀釋與 N2 進度。"),
 ("","TSMC Q4'25 結果(2026-01-15):Q1'26 前次指引 US$34.6–35.8bn / GM 63–65%。"),
 ("","Investing.com transcript;Manufacturing Dive;DataCenterDynamics;Tickeron;TipRanks(consensus 與股價反應)。"),
 ("H","免責"),
 ("","彙整自公開資訊與第三方,僅供內部研究參考,不構成投資建議。")]
r=3
for tag,t in items:
    if tag=="H": ns.cell(r,2,t).font=Font(name=CJK,size=12,bold=True,color="1F4E79"); r+=1
    else:
        c=ns.cell(r,2,"•  "+t); c.font=Ff(10); c.alignment=LEF; ns.row_dimensions[r].height=15; r+=1

wb.save("./out/TSMC_Q1-2026_Model.xlsx")
print("SAVED ./out/TSMC_Q1-2026_Model.xlsx  sheets:", wb.sheetnames)
