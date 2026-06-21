# -*- coding: utf-8 -*-
"""Vertiv (VRT) DCF — institutional build per model-builder dcf-model skill.
Two sheets (DCF + WACC), Bear/Base/Bull scenarios, mid-year convention,
perpetuity terminal value, one WACC x terminal-growth sensitivity table."""
import os
os.makedirs("./out", exist_ok=True)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter

# ---- xlsx skill palette: blue font=input, black=formula, green=cross-sheet link ----
BLUEF="0000FF"; BLACKF="000000"; GREENF="008000"
HDR="1F4E79"; SUB="D9E1F2"; OUT="BDD7EE"; INP="F2F2F2"; WHITE="FFFFFF"; CTR="BDD7EE"
CJK="Microsoft JhengHei"
med=Side(style="thin",color="9BAFC4"); thick=Side(style="medium",color="1F4E79")
box=Border(left=thick,right=thick,top=thick,bottom=thick)
cell_b=Border(left=med,right=med,top=med,bottom=med)
def Fi(sz=10,b=False): return Font(name=CJK,size=sz,bold=b,color=BLUEF)   # input
def Ff(sz=10,b=False): return Font(name=CJK,size=sz,bold=b,color=BLACKF)  # formula
def Fl(sz=10,b=False): return Font(name=CJK,size=sz,bold=b,color=GREENF)  # link
def Fh(sz=11,b=True):  return Font(name=CJK,size=sz,bold=b,color=WHITE)   # header
def fill(c): return PatternFill("solid",fgColor=c)
CEN=Alignment(horizontal="center",vertical="center",wrap_text=True)
RGT=Alignment(horizontal="right",vertical="center")
LEF=Alignment(horizontal="left",vertical="center",wrap_text=True)
PCT="0.0%"; PCTS="0.0%;(0.0%)"; BN='$#,##0.00;($#,##0.00);"-"'; SH2="$#,##0.00"; MULT='0.0"x"'; NUM="0.000"

wb=Workbook()
# ============================================================ WACC sheet
wc=wb.create_sheet("WACC"); wc.sheet_view.showGridLines=False
def wput(cell,val,font,fmt=None,al=RGT,bd=True,fl=None,cmt=None):
    c=wc[cell]; c.value=val; c.font=font
    if fmt:c.number_format=fmt
    c.alignment=al
    if bd:c.border=cell_b
    if fl:c.fill=fill(fl)
    if cmt:c.comment=Comment(cmt,"DCF skill")
    return c
def whdr(rng,text):
    a=rng.split(":")[0]; wc[a]=text; wc[a].font=Fh(11); wc.merge_cells(rng)
    for row in wc[rng]:
        for c in row: c.fill=fill(HDR); c.border=cell_b
wc.merge_cells("B1:D1"); wc["B1"]="Vertiv (VRT) — WACC（資金成本，CAPM）"; wc["B1"].font=Fh(13); wc["B1"].fill=fill(HDR)
wc["B1"].alignment=Alignment(horizontal="left",vertical="center"); wc.row_dimensions[1].height=24
whdr("B3:D3","權益成本 COST OF EQUITY")
rows=[("無風險利率 (10Y UST)",0.043,PCT,Fi(),"Source: US 10Y Treasury, 2026-06, ~4.3%"),
      ("Beta (5年月beta)",1.45,NUM,Fi(),"Source: 5-yr monthly beta vs S&P500, ~1.4-1.5 (high-beta AI infra name)"),
      ("股權風險溢酬 ERP",0.05,PCT,Fi(),"Source: market standard 5.0%")]
r=4
for lab,v,fmt,fn,cm in rows:
    wput(f"B{r}",lab,Ff(),al=LEF,fl=INP); wput(f"C{r}",v,fn,fmt,fl=INP,cmt=cm); r+=1
wput("B7","權益成本 Cost of Equity",Ff(10,True),al=LEF,fl=SUB); wput("C7","=C4+C5*C6",Ff(10,True),PCT,fl=SUB)
whdr("B9:D9","負債成本 COST OF DEBT")
wput("B10","稅前負債成本",Ff(),al=LEF,fl=INP); wput("C10",0.055,Fi(),PCT,fl=INP,cmt="Source: implied from interest exp/debt & BB+/BBB- credit, ~5.5%")
wput("B11","稅率 (連結 DCF)",Ff(),al=LEF); wput("C11","=DCF!$C$32",Fl(),PCT)
wput("B12","稅後負債成本",Ff(10,True),al=LEF,fl=SUB); wput("C12","=C10*(1-C11)",Ff(10,True),PCT,fl=SUB)
whdr("B14:D14","資本結構 CAPITAL STRUCTURE ($bn)")
wput("B15","股價 (連結 DCF)",Ff(),al=LEF); wput("C15","=DCF!$C$8",Fl(),SH2)
wput("B16","股數 bn (連結 DCF)",Ff(),al=LEF); wput("C16","=DCF!$C$9",Fl(),NUM)
wput("B17","市值 Market Cap",Ff(),al=LEF); wput("C17","=C15*C16",Ff(),BN)
wput("B18","總負債 Total Debt",Ff(),al=LEF,fl=INP); wput("C18",2.90,Fi(),BN,fl=INP,cmt="Source: VRT balance sheet FY25, total debt ~$2.9bn")
wput("B19","現金 Cash & Equiv",Ff(),al=LEF,fl=INP); wput("C19",1.00,Fi(),BN,fl=INP,cmt="Source: VRT balance sheet FY25, cash ~$1.0bn")
wput("B20","淨負債 Net Debt",Ff(10,True),al=LEF,fl=SUB); wput("C20","=C18-C19",Ff(10,True),BN,fl=SUB)
wput("B21","企業價值 EV",Ff(),al=LEF); wput("C21","=C17+C20",Ff(),BN)
wput("B23","WACC 計算",Fh(10),al=LEF,fl=HDR)
wput("C23","權重",Fh(10),al=CEN,fl=HDR); wput("D23","成本",Fh(10),al=CEN,fl=HDR); wput("E23","貢獻",Fh(10),al=CEN,fl=HDR)
wput("B24","權益 Equity",Ff(),al=LEF); wput("C24","=C17/C21",Ff(),PCT); wput("D24","=C7",Ff(),PCT); wput("E24","=C24*D24",Ff(),PCT)
wput("B25","負債 Debt",Ff(),al=LEF); wput("C25","=C20/C21",Ff(),PCT); wput("D25","=C12",Ff(),PCT); wput("E25","=C25*D25",Ff(),PCT)
wput("B27","WACC",Ff(11,True),al=LEF,fl=OUT); wc.merge_cells("C27:D27")
wput("C27","=E24+E25",Ff(11,True),PCT,al=CEN,fl=OUT)
wc.column_dimensions["A"].width=2; wc.column_dimensions["B"].width=26
for c in "CDE": wc.column_dimensions[c].width=12

# ============================================================ DCF sheet
ws=wb.create_sheet("DCF",0); ws.sheet_view.showGridLines=False
YRS=["FY25A","FY26E","FY27E","FY28E","FY29E","FY30E"]  # cols C..H
def put(cell,val,font,fmt=None,al=RGT,bd=True,fl=None,cmt=None):
    c=ws[cell]; c.value=val; c.font=font
    if fmt:c.number_format=fmt
    c.alignment=al
    if bd:c.border=cell_b
    if fl:c.fill=fill(fl)
    if cmt:c.comment=Comment(cmt,"DCF skill")
    return c
def hdr(rng,text):
    a=rng.split(":")[0]; ws[a]=text; ws[a].font=Fh(11)
    ws.merge_cells(rng)
    for row in ws[rng]:
        for c in row: c.fill=fill(HDR); c.border=cell_b
    ws[a].alignment=Alignment(horizontal="left",vertical="center")

# -- header --
ws.merge_cells("B1:H1"); ws["B1"]="Vertiv Holdings (VRT) — DCF 估值模型"; ws["B1"].font=Fh(14); ws["B1"].fill=fill(HDR)
ws["B1"].alignment=Alignment(horizontal="left",vertical="center"); ws.row_dimensions[1].height=26
ws.merge_cells("B2:H2"); ws["B2"]="Ticker: VRT | 日期: 2026-06-18 | 會計年底: 12月 | 單位: US$ bn（每股除外）| 期中折現慣例"
ws["B2"].font=Font(name=CJK,size=9,color=BLACKF); ws["B2"].alignment=Alignment(horizontal="left",vertical="center")
# -- case selector --
put("B4","情境選擇 (1熊 2基 3牛)",Ff(10,True),al=LEF,fl=SUB)
put("C4",2,Fi(10,True),"0",al=CEN,fl=INP,cmt="1=Bear, 2=Base, 3=Bull; 改此格切換情境")
put("B5","目前情境",Ff(),al=LEF); put("C5",'=IF($C$4=1,"熊 Bear",IF($C$4=2,"基 Base","牛 Bull"))',Ff(10,True),al=CEN)

# -- market data --
hdr("B7:H7","市場資料 MARKET DATA（與情境無關）")
put("B8","現價 Stock Price ($)",Ff(),al=LEF); put("C8",320.00,Fi(),SH2,fl=INP,cmt="Source: market data 2026-06, ~$320 (≈$122bn mcap / 0.382bn sh)")
put("B9","稀釋股數 Diluted Shares (bn)",Ff(),al=LEF); put("C9",0.382,Fi(),NUM,fl=INP,cmt="Source: VRT 10-K/Q FY25, diluted shares ~382m")
put("B10","市值 Market Cap",Ff(),al=LEF); put("C10","=C8*C9",Ff(),BN)
put("B11","淨負債 Net Debt (連結 WACC)",Ff(),al=LEF); put("C11","=WACC!$C$20",Fl(),BN)

# -- scenario assumption blocks (years in D..H) --
hdr("B13:H13","情境假設 SCENARIO ASSUMPTIONS（營收成長 / EBIT 利潤率）")
def yearhdr(rrow):
    put(f"B{rrow}","假設項目",Ff(9,True),al=LEF,fl=SUB)
    for i,y in enumerate(YRS[1:]):
        put(f"{get_column_letter(4+i)}{rrow}",y,Ff(9,True),al=CEN,fl=SUB)
# Bear
put("B14","熊 BEAR CASE",Ff(10,True),al=LEF,fl=SUB); ws.merge_cells("B14:H14")
yearhdr(15)
bear_g=[.22,.14,.09,.06,.04]; bear_m=[.185,.19,.195,.20,.20]
base_g=[.27,.20,.15,.11,.08]; base_m=[.195,.205,.215,.22,.225]
bull_g=[.32,.26,.20,.15,.11]; bull_m=[.20,.215,.23,.24,.245]
def arow(rrow,label,vals,cmt=None):
    put(f"B{rrow}",label,Ff(),al=LEF)
    for i,v in enumerate(vals):
        c=put(f"{get_column_letter(4+i)}{rrow}",v,Fi(),PCT,fl=INP)
        if cmt and i==0: c.comment=Comment(cmt,"DCF skill")
arow(16,"  營收成長 %",bear_g,"Bear: 雲端資本支出消化、競爭加劇情境")
arow(17,"  EBIT 利潤率 %",bear_m)
put("B18","基 BASE CASE",Ff(10,True),al=LEF,fl=SUB); ws.merge_cells("B18:H18")
yearhdr(19)
arow(20,"  營收成長 %",base_g,"Base: FY26 +27% 對齊管理層指引中點 (~$13.75bn)，後續遞減")
arow(21,"  EBIT 利潤率 %",base_m,"Base: EBIT 由 ~19% 擴張至 ~22.5% (營運槓桿+液冷組合)")
put("B22","牛 BULL CASE",Ff(10,True),al=LEF,fl=SUB); ws.merge_cells("B22:H22")
yearhdr(23)
arow(24,"  營收成長 %",bull_g,"Bull: 液冷滲透加速、份額擴張")
arow(25,"  EBIT 利潤率 %",bull_m)
# Selected (consolidation) via CHOOSE
put("B26","選定情境 SELECTED（=CHOOSE 由選擇器）",Ff(10,True),al=LEF,fl=OUT); ws.merge_cells("B26:H26")
for i in range(5):
    col=get_column_letter(4+i)
    put(f"{col}27",f"=CHOOSE($C$4,{col}16,{col}20,{col}24)",Ff(),PCT,fl=OUT)
    put(f"{col}28",f"=CHOOSE($C$4,{col}17,{col}21,{col}25)",Ff(),PCT,fl=OUT)
put("B27","  選定營收成長 %",Ff(),al=LEF,fl=OUT)
put("B28","  選定 EBIT 利潤率 %",Ff(),al=LEF,fl=OUT)

# -- global assumptions --
hdr("B30:H30","全域假設 GLOBAL ASSUMPTIONS（與情境無關）")
put("B31","稅率 Tax Rate",Ff(),al=LEF); put("C31",0.21,Fi(),PCT,fl=INP,cmt="Source: US statutory + state, ~21%")
put("B32","稅率 (供 WACC 連結)",Ff(),al=LEF); put("C32","=C31",Ff(),PCT)
put("B33","D&A % 營收",Ff(),al=LEF); put("C33",0.026,Fi(),PCT,fl=INP,cmt="Source: VRT historical D&A ~2.6% of revenue")
put("B34","Capex % 營收",Ff(),al=LEF); put("C34",0.023,Fi(),PCT,fl=INP,cmt="Source: VRT historical capex ~2.3% of revenue (asset-light)")
put("B35","ΔNWC % 增額營收",Ff(),al=LEF); put("C35",0.08,Fi(),PCT,fl=INP,cmt="Source: working-capital build ~8% of incremental revenue")
put("B36","終值成長 g",Ff(),al=LEF); put("C36",0.03,Fi(),PCT,fl=INP,cmt="Source: ~long-run GDP; constraint g<WACC")
put("B37","WACC (連結 WACC 分頁)",Ff(10,True),al=LEF,fl=SUB); put("C37","=WACC!$C$27",Fl(10,True),PCT,fl=SUB)

# -- projections C..H (C=FY25A) --
hdr("B39:H39","財務預測 PROJECTIONS (US$ bn)")
put("B40","項目 \\ 年度",Ff(9,True),al=LEF,fl=SUB)
for i,y in enumerate(YRS):
    put(f"{get_column_letter(3+i)}40",y,Ff(9,True),al=CEN,fl=SUB)
# Revenue
put("B41","營收 Revenue",Ff(10,True),al=LEF)
put("C41",10.84,Fi(10,True),BN,fl=INP,cmt="Source: VRT FY25 revenue ~$10.84bn (TTM)")
for i in range(5):
    col=get_column_letter(4+i); prev=get_column_letter(3+i)
    put(f"{col}41",f"={prev}41*(1+{col}27)",Ff(10,True),BN)
put("B42","  營收成長 %",Ff(),al=LEF)
for i in range(5):
    col=get_column_letter(4+i); prev=get_column_letter(3+i)
    put(f"{col}42",f"={col}41/{prev}41-1",Ff(),PCT)
put("B43","  EBIT 利潤率 %",Ff(),al=LEF)
put("C43",0.188,Fi(),PCT,fl=INP,cmt="Source: VRT FY25 adj. operating margin ~18.8%")
for i in range(5):
    col=get_column_letter(4+i); put(f"{col}43",f"={col}28",Ff(),PCT)
put("B44","EBIT",Ff(10,True),al=LEF)
for i in range(6):
    col=get_column_letter(3+i); put(f"{col}44",f"={col}41*{col}43",Ff(10,True),BN)
put("B45","  減：稅 Taxes",Ff(),al=LEF)
for i in range(6):
    col=get_column_letter(3+i); put(f"{col}45",f"={col}44*$C$31",Ff(),BN)
put("B46","NOPAT",Ff(10,True),al=LEF)
for i in range(6):
    col=get_column_letter(3+i); put(f"{col}46",f"={col}44-{col}45",Ff(10,True),BN)
put("B47","  加：D&A",Ff(),al=LEF)
for i in range(6):
    col=get_column_letter(3+i); put(f"{col}47",f"={col}41*$C$33",Ff(),BN)
put("B48","  減：Capex",Ff(),al=LEF)
for i in range(6):
    col=get_column_letter(3+i); put(f"{col}48",f"={col}41*$C$34",Ff(),BN)
put("B49","  減：ΔNWC",Ff(),al=LEF)
put("C49",0,Ff(),BN)
for i in range(5):
    col=get_column_letter(4+i); prev=get_column_letter(3+i)
    put(f"{col}49",f"=({col}41-{prev}41)*$C$35",Ff(),BN)
put("B50","無槓桿自由現金流 FCFF",Ff(10,True),al=LEF,fl=SUB)
for i in range(6):
    col=get_column_letter(3+i); put(f"{col}50",f"={col}46+{col}47-{col}48-{col}49",Ff(10,True),BN,fl=SUB)

# -- discounting (forecast D..H) --
hdr("B52:H52","折現與終值 DISCOUNTING & TERMINAL VALUE")
put("B53","期數 (期中) Period",Ff(),al=LEF)
periods=[0.5,1.5,2.5,3.5,4.5]
for i,p in enumerate(periods):
    put(f"{get_column_letter(4+i)}53",p,Fi(),"0.0",fl=INP,cmt="Mid-year convention" if i==0 else None)
put("B54","折現因子 Discount Factor",Ff(),al=LEF)
for i in range(5):
    col=get_column_letter(4+i); put(f"{col}54",f"=1/(1+$C$37)^{col}53",Ff(),"0.000")
put("B55","FCFF 現值 PV",Ff(10,True),al=LEF)
for i in range(5):
    col=get_column_letter(4+i); put(f"{col}55",f"={col}50*{col}54",Ff(10,True),BN)

# -- valuation summary (col C) --
hdr("B57:H57","估值總結 VALUATION SUMMARY (US$ bn)")
put("B58","終年 FCFF × (1+g) = 終值 FCFF",Ff(),al=LEF); put("C58","=H50*(1+$C$36)",Ff(),BN)
put("B59","終值 TV (永續成長法)",Ff(),al=LEF); put("C59","=C58/($C$37-$C$36)",Ff(),BN)
put("B60","終值現值 PV of TV",Ff(),al=LEF); put("C60","=C59/(1+$C$37)^H53",Ff(),BN)
put("B61","顯性期 PV 合計 ΣPV FCFF",Ff(),al=LEF); put("C61","=SUM(D55:H55)",Ff(),BN)
put("B62","企業價值 Enterprise Value",Ff(10,True),al=LEF,fl=SUB); put("C62","=C61+C60",Ff(10,True),BN,fl=SUB)
put("B63","  減：淨負債 Net Debt",Ff(),al=LEF); put("C63","=C11",Ff(),BN)
put("B64","股權價值 Equity Value",Ff(10,True),al=LEF,fl=SUB); put("C64","=C62-C63",Ff(10,True),BN,fl=SUB)
put("B65","  稀釋股數 (bn)",Ff(),al=LEF); put("C65","=C9",Ff(),NUM)
put("B66","隱含每股價值 Implied Price",Ff(11,True),al=LEF,fl=OUT); put("C66","=C64/C65",Ff(11,True),SH2,fl=OUT)
put("B67","現價 Current Price",Ff(),al=LEF); put("C67","=C8",Ff(),SH2)
put("B68","隱含上漲/(下跌)",Ff(11,True),al=LEF,fl=OUT); put("C68","=C66/C67-1",Ff(11,True),PCTS,fl=OUT)
# checks
put("B70","檢查：終值占 EV %",Ff(),al=LEF); put("C70","=C60/C62",Ff(),PCT,cmt="健康區間 50–70%；偏高代表高度仰賴終值")
put("B71","檢查：隱含 EV/EBITDA (FY26E)",Ff(),al=LEF); put("C71","=C62/(D44+D47)",Ff(),MULT)

# -- sensitivity: WACC x terminal growth (5x5, base-centered) --
hdr("B74:H74","敏感度分析 SENSITIVITY：隱含每股價值 — WACC × 終值成長 g")
put("B75","WACC↓ / g→",Ff(9,True),al=CEN,fl=SUB)
# g axis (cols D..H), centered on C36
g_off=[-0.01,-0.005,0,0.005,0.01]
for i,off in enumerate(g_off):
    col=get_column_letter(4+i)
    put(f"{col}75",f"=$C$36+({off})",Fi(9,True),PCT,al=CEN,fl=SUB)
# WACC axis (rows 76..80) in col C, centered on C37
w_off=[-0.01,-0.005,0,0.005,0.01]
for r,off in enumerate(w_off):
    rr=76+r
    put(f"C{rr}",f"=$C$37+({off})",Fi(9,True),PCT,al=CEN,fl=SUB)
    for cidx in range(5):
        col=get_column_letter(4+cidx)
        # full DCF recalc: PV explicit FCFF + PV perpetuity TV, minus net debt, / shares
        f=(f"=(SUMPRODUCT($D$50:$H$50,1/(1+$C{rr})^$D$53:$H$53)"
           f"+($H$50*(1+{col}$75)/($C{rr}-{col}$75))/(1+$C{rr})^$H$53"
           f"-$C$11)/$C$9")
        center = (r==2 and cidx==2)
        cc=put(f"{col}{rr}",f,Ff(10,center),SH2,al=CEN,fl=(CTR if center else (INP if (r+cidx)%2 else WHITE)))
put("B82","註：① 中心格(粗體高亮)=基準情境，應等於上方隱含每股價值 C66。② 表格採目前情境(C4)之 FCFF 流；切換情境後同步更新。③ g<WACC 為必要條件。",
    Ff(8,False),al=LEF); ws.merge_cells("B82:H83")

# widths
ws.column_dimensions["A"].width=2; ws.column_dimensions["B"].width=30
for c in "CDEFGH": ws.column_dimensions[c].width=11.5

if "Sheet" in wb.sheetnames: wb.remove(wb["Sheet"])
wb.save("./out/VRT_DCF_Model_2026-06-18.xlsx")
print("SAVED ./out/VRT_DCF_Model_2026-06-18.xlsx  sheets:", wb.sheetnames)
