# -*- coding: utf-8 -*-
"""TSMC (TSM/2330.TW) DCF — US$, net-cash, ADR, Bear/Base/Bull scenarios. Live formulas."""
import os
os.makedirs("./out", exist_ok=True)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter

BLUEF="0000FF"; BLACKF="000000"; GREENF="008000"
HDR="1F4E79"; SUB="D9E1F2"; OUT="BDD7EE"; INP="F2F2F2"; WHITE="FFFFFF"; CTR="BDD7EE"
CJK="Microsoft JhengHei"
med=Side(style="thin",color="9BAFC4"); cell_b=Border(left=med,right=med,top=med,bottom=med)
def Fi(s=10,b=False): return Font(name=CJK,size=s,bold=b,color=BLUEF)
def Ff(s=10,b=False): return Font(name=CJK,size=s,bold=b,color=BLACKF)
def Fl(s=10,b=False): return Font(name=CJK,size=s,bold=b,color=GREENF)
def Fh(s=11,b=True):  return Font(name=CJK,size=s,bold=b,color=WHITE)
def fill(c): return PatternFill("solid",fgColor=c)
CEN=Alignment(horizontal="center",vertical="center",wrap_text=True)
RGT=Alignment(horizontal="right",vertical="center"); LEF=Alignment(horizontal="left",vertical="center",wrap_text=True)
PCT="0.0%"; PCTS="+0.0%;(0.0%)"; USD='$#,##0.0'; SH2="$#,##0.00"; NUM="0.000"; MULT='0.0"x"'

wb=Workbook()
# ===================== WACC =====================
wc=wb.create_sheet("WACC"); wc.sheet_view.showGridLines=False
def wp(cell,v,f,fmt=None,al=RGT,fl=None,cmt=None):
    c=wc[cell]; c.value=v; c.font=f; c.border=cell_b; c.alignment=al
    if fmt:c.number_format=fmt
    if fl:c.fill=fill(fl)
    if cmt:c.comment=Comment(cmt,"valuation")
wc.merge_cells("B1:D1"); wc["B1"]="TSMC — WACC (CAPM, 淨現金)"; wc["B1"].font=Fh(12); wc["B1"].fill=fill(HDR)
wc["B1"].alignment=Alignment(horizontal="left",vertical="center")
for r,(lab,v,fmt,fn,cm) in enumerate([("無風險利率 10Y UST",0.043,PCT,Fi(),"US 10Y ~4.3%"),
        ("Beta",1.15,NUM,Fi(),"TSM ADR ~1.1-1.2"),("股權風險溢酬 ERP",0.05,PCT,Fi(),"5.0%")],start=3):
    wp(f"B{r}",lab,Ff(),al=LEF,fl=INP); wp(f"C{r}",v,fn,fmt,fl=INP,cmt=cm)
wp("B6","權益成本",Ff(10,True),al=LEF,fl=SUB); wp("C6","=C3+C4*C5",Ff(10,True),PCT,fl=SUB)
wp("B8","稅前負債成本",Ff(),al=LEF,fl=INP); wp("C8",0.025,Fi(),PCT,fl=INP)
wp("B9","稅率",Ff(),al=LEF,fl=INP); wp("C9",0.15,Fi(),PCT,fl=INP)
wp("B10","稅後負債成本",Ff(),al=LEF); wp("C10","=C8*(1-C9)",Ff(),PCT)
wp("B12","現價/股(US$,ADR/5)",Ff(),al=LEF,fl=INP); wp("C12",76.0,Fi(),SH2,fl=INP,cmt="ADR ~$380/5")
wp("B13","股數 (bn)",Ff(),al=LEF,fl=INP); wp("C13",25.93,Fi(),NUM,fl=INP)
wp("B14","市值",Ff(),al=LEF); wp("C14","=C12*C13",Ff(),USD)
wp("B15","總負債",Ff(),al=LEF,fl=INP); wp("C15",25.0,Fi(),USD,fl=INP)
wp("B16","現金",Ff(),al=LEF,fl=INP); wp("C16",90.0,Fi(),USD,fl=INP,cmt="net cash ~$65bn")
wp("B17","淨現金",Ff(10,True),al=LEF,fl=SUB); wp("C17","=C16-C15",Ff(10,True),USD,fl=SUB)
wp("B19","權益權重",Ff(),al=LEF); wp("C19","=C14/(C14+C15)",Ff(),PCT)
wp("B20","負債權重",Ff(),al=LEF); wp("C20","=C15/(C14+C15)",Ff(),PCT)
wp("B22","WACC",Ff(11,True),al=LEF,fl=OUT); wp("C22","=C6*C19+C10*C20",Ff(11,True),PCT,fl=OUT)
wc.column_dimensions["A"].width=2; wc.column_dimensions["B"].width=24; wc.column_dimensions["C"].width=13

# ===================== DCF =====================
ws=wb.create_sheet("DCF",0); ws.sheet_view.showGridLines=False
YRS=["FY26E","FY27E","FY28E","FY29E","FY30E","FY31E"]
def put(cell,v,f,fmt=None,al=RGT,fl=None,cmt=None):
    c=ws[cell]; c.value=v; c.font=f; c.border=cell_b; c.alignment=al
    if fmt:c.number_format=fmt
    if fl:c.fill=fill(fl)
    if cmt:c.comment=Comment(cmt,"valuation")
def hdr(rng,t):
    a=rng.split(":")[0]; ws[a]=t; ws[a].font=Fh(11); ws.merge_cells(rng)
    for row in ws[rng]:
        for c in row: c.fill=fill(HDR); c.border=cell_b
    ws[a].alignment=Alignment(horizontal="left",vertical="center")
ws.merge_cells("B1:H1"); ws["B1"]="台積電 TSMC (TSM/2330.TW) — DCF 估值（US$,熊/基/牛情境）"; ws["B1"].font=Fh(13); ws["B1"].fill=fill(HDR)
ws["B1"].alignment=Alignment(horizontal="left",vertical="center"); ws.row_dimensions[1].height=24
ws.merge_cells("B2:H2"); ws["B2"]="單位 US$bn(每股除外)| 藍=輸入 黑=公式 綠=連結 | FY26E 為基期 | 改 L11 切換情境"; ws["B2"].font=Ff(8)
ws["B2"].alignment=Alignment(horizontal="left",vertical="center")

# inputs panel
put("K1","輸入 INPUTS",Fh(10),al=LEF,fl=HDR); ws.merge_cells("K1:L1")
SEL="$L$11"
inp=[("WACC(連結)","=WACC!$C$22",PCT,Fl()),
     ("終值成長 g","=CHOOSE($L$11,$C$62,$D$62,$E$62)",PCT,Ff()),
     ("稅率",0.15,PCT,Fi()),("D&A %營收",0.19,PCT,Fi()),("ΔNWC %增額",0.03,PCT,Fi()),
     ("淨現金 $bn","=WACC!$C$17",USD,Fl()),("股數 bn","=WACC!$C$13",NUM,Fl()),
     ("ADR 比例(股/ADR)",5,"0",Fi()),("現價/ADR $",380.0,SH2,Fi())]
for r,(lab,v,fmt,fn) in enumerate(inp,start=2):
    put(f"K{r}",lab,Ff(9),al=LEF,fl=INP); put(f"L{r}",v,fn,fmt,fl=(INP if fn.color.rgb==BLUEF else WHITE))
put("K11","情境(1熊2基3牛)",Ff(9,True),al=LEF,fl=SUB); put("L11",2,Fi(10,True),"0",al=CEN,fl=INP,cmt="1=Bear 2=Base 3=Bull")
WACC="$L$2"; G="$L$3"; TAX="$L$4"; DA="$L$5"; NWC="$L$6"; NETC="$L$7"; SH="$L$8"; ADR="$L$9"; PX="$L$10"

hdr("B4:H4","現金流預測 PROJECTION (US$bn) — 目前情境")
put("B5","項目",Fh(9),al=LEF,fl=HDR)
for i,y in enumerate(YRS): put(f"{get_column_letter(3+i)}5",y,Fh(9),al=CEN,fl=HDR)
put("B6","營收 Revenue",Ff(10,True),al=LEF)
put("C6",163.6,Fi(10,True),USD,fl=INP,cmt="FY26E base from TSMC quarterly model")
put("B7","  成長 %",Ff(),al=LEF); put("C7",0.319,Fi(),PCTS,fl=INP)
for i in range(5):
    col=get_column_letter(4+i); prev=get_column_letter(3+i)
    put(f"{col}7",f"=CHOOSE({SEL},{col}48,{col}49,{col}50)",Ff(),PCTS)
    put(f"{col}6",f"={prev}6*(1+{col}7)",Ff(10,True),USD)
put("B8","營業利益率 %",Ff(),al=LEF)
for i in range(6):
    col=get_column_letter(3+i); put(f"{col}8",f"=CHOOSE({SEL},{col}53,{col}54,{col}55)",Ff(),PCT)
put("B9","EBIT",Ff(10,True),al=LEF)
for i in range(6): col=get_column_letter(3+i); put(f"{col}9",f"={col}6*{col}8",Ff(10,True),USD)
put("B10","  減：稅",Ff(),al=LEF)
for i in range(6): col=get_column_letter(3+i); put(f"{col}10",f"={col}9*{TAX}",Ff(),USD)
put("B11","NOPAT",Ff(10,True),al=LEF)
for i in range(6): col=get_column_letter(3+i); put(f"{col}11",f"={col}9-{col}10",Ff(10,True),USD)
put("B12","  加：D&A",Ff(),al=LEF)
for i in range(6): col=get_column_letter(3+i); put(f"{col}12",f"={col}6*{DA}",Ff(),USD)
put("B13","  減：Capex(%營收)",Ff(),al=LEF)
for i in range(6):
    col=get_column_letter(3+i); put(f"{col}13",f"=CHOOSE({SEL},{col}58,{col}59,{col}60)",Ff(),PCT)
put("B14","  Capex $",Ff(),al=LEF)
for i in range(6): col=get_column_letter(3+i); put(f"{col}14",f"={col}6*{col}13",Ff(),USD)
put("B15","  減：ΔNWC",Ff(),al=LEF); put("C15",0,Ff(),USD)
for i in range(5):
    col=get_column_letter(4+i); prev=get_column_letter(3+i); put(f"{col}15",f"=({col}6-{prev}6)*{NWC}",Ff(),USD)
put("B16","無槓桿 FCFF",Ff(10,True),al=LEF,fl=SUB)
for i in range(6): col=get_column_letter(3+i); put(f"{col}16",f"={col}11+{col}12-{col}14-{col}15",Ff(10,True),USD,fl=SUB)
put("B17","期數(期中)",Ff(),al=LEF)
for i in range(5): put(f"{get_column_letter(4+i)}17",0.5+i,Fi(),"0.0",fl=INP)
put("B18","折現因子",Ff(),al=LEF)
for i in range(5): col=get_column_letter(4+i); put(f"{col}18",f"=1/(1+{WACC})^{col}17",Ff(),"0.000")
put("B19","FCFF 現值",Ff(10,True),al=LEF)
for i in range(5): col=get_column_letter(4+i); put(f"{col}19",f"={col}16*{col}18",Ff(10,True),USD)

hdr("B21:H21","估值 VALUATION (US$bn)")
put("B22","終值 TV(永續)",Ff(),al=LEF); put("C22",f"=H16*(1+{G})/({WACC}-{G})",Ff(),USD)
put("B23","終值現值",Ff(),al=LEF); put("C23","=C22/(1+"+WACC+")^H17",Ff(),USD)
put("B24","顯性期 PV 合計",Ff(),al=LEF); put("C24","=SUM(D19:H19)",Ff(),USD)
put("B25","企業價值 EV",Ff(10,True),al=LEF,fl=SUB); put("C25","=C24+C23",Ff(10,True),USD,fl=SUB)
put("B26","加：淨現金",Ff(),al=LEF); put("C26","="+NETC,Ff(),USD)
put("B27","股權價值",Ff(10,True),al=LEF,fl=SUB); put("C27","=C25+C26",Ff(10,True),USD,fl=SUB)
put("B28","每股價值(US$)",Ff(),al=LEF); put("C28","=C27/"+SH,Ff(),SH2)
put("B29","每 ADR 價值(US$)",Ff(11,True),al=LEF,fl=OUT); put("C29","=C28*"+ADR,Ff(11,True),SH2,fl=OUT)
put("B30","現價/ADR(US$)",Ff(),al=LEF); put("C30","="+PX,Ff(),SH2)
put("B31","上漲/(下跌)",Ff(11,True),al=LEF,fl=OUT); put("C31","=C29/C30-1",Ff(11,True),PCTS,fl=OUT)
put("B32","檢查:TV 占 EV",Ff(),al=LEF); put("C32","=C23/C25",Ff(),PCT)

# sensitivity
hdr("B34:H34","敏感度:每 ADR 價值(US$)— WACC × 終值成長 g(目前情境)")
put("B35","WACC↓/g→",Ff(9,True),al=CEN,fl=SUB)
for i,off in enumerate([-0.01,-0.005,0,0.005,0.01]):
    put(f"{get_column_letter(4+i)}35",f"=$L$3+({off})",Fi(9,True),PCT,al=CEN,fl=SUB)
for r2,off in enumerate([-0.01,-0.005,0,0.005,0.01]):
    rr=36+r2; put(f"C{rr}",f"=$L$2+({off})",Fi(9,True),PCT,al=CEN,fl=SUB)
    for c2 in range(5):
        col=get_column_letter(4+c2)
        f=(f"=((SUMPRODUCT($D$16:$H$16,1/(1+$C{rr})^$D$17:$H$17)"
           f"+($H$16*(1+{col}$35)/($C{rr}-{col}$35))/(1+$C{rr})^$H$17"
           f"+{NETC})/{SH})*{ADR}")
        ctr=(r2==2 and c2==2)
        put(f"{col}{rr}",f,Ff(10,ctr),SH2,al=CEN,fl=(CTR if ctr else (INP if (r2+c2)%2 else WHITE)))

# scenario tables (rows 45-62)
hdr("B45:H45","情境假設 SCENARIO ASSUMPTIONS（可編輯）")
put("B47","營收成長 % (FY27→31)",Ff(9,True),al=LEF,fl=SUB)
for i,y in enumerate(YRS[1:]): put(f"{get_column_letter(4+i)}47",y,Ff(9,True),al=CEN,fl=SUB)
growth={"熊 Bear":[.14,.10,.08,.06,.05],"基 Base":[.20,.16,.13,.10,.08],"牛 Bull":[.26,.22,.17,.13,.10]}
for r,(nm,vals) in enumerate(growth.items(),start=48):
    put(f"B{r}",nm,Ff(9),al=LEF,fl=INP)
    for i,v in enumerate(vals): put(f"{get_column_letter(4+i)}{r}",v,Fi(9),PCTS,fl=INP)
put("B52","營業利益率 % (FY26→31)",Ff(9,True),al=LEF,fl=SUB)
for i,y in enumerate(YRS): put(f"{get_column_letter(3+i)}52",y,Ff(9,True),al=CEN,fl=SUB)
om={"熊 Bear":[.57,.54,.54,.545,.55,.55],"基 Base":[.57,.56,.565,.57,.57,.57],"牛 Bull":[.57,.57,.58,.585,.59,.59]}
for r,(nm,vals) in enumerate(om.items(),start=53):
    put(f"B{r}",nm,Ff(9),al=LEF,fl=INP)
    for i,v in enumerate(vals): put(f"{get_column_letter(3+i)}{r}",v,Fi(9),PCT,fl=INP)
put("B57","Capex % 營收 (FY26→31)",Ff(9,True),al=LEF,fl=SUB)
for i,y in enumerate(YRS): put(f"{get_column_letter(3+i)}57",y,Ff(9,True),al=CEN,fl=SUB)
capex={"熊 Bear":[.33,.33,.32,.31,.30,.30],"基 Base":[.33,.32,.30,.28,.27,.26],"牛 Bull":[.33,.31,.29,.27,.25,.24]}
for r,(nm,vals) in enumerate(capex.items(),start=58):
    put(f"B{r}",nm,Ff(9),al=LEF,fl=INP)
    for i,v in enumerate(vals): put(f"{get_column_letter(3+i)}{r}",v,Fi(9),PCT,fl=INP)
put("B61","終值成長 g (熊/基/牛)",Ff(9,True),al=LEF,fl=SUB)
for i,(lab,v) in enumerate([("熊",0.025),("基",0.035),("牛",0.040)]):
    put(f"{get_column_letter(3+i)}62",v,Fi(9),PCT,al=CEN,fl=INP)
put("B62","  g →",Ff(9),al=LEF)

put("B64","註:情境驅動 營收成長/利潤率/Capex/終值g;WACC 固定(其變化見上方敏感度表)。中心格=基準應≈C29。",Ff(8),al=LEF)
ws.merge_cells("B64:H65")
ws.column_dimensions["A"].width=2; ws.column_dimensions["B"].width=22
for c in "CDEFGH": ws.column_dimensions[c].width=11
ws.column_dimensions["J"].width=2; ws.column_dimensions["K"].width=18; ws.column_dimensions["L"].width=12

# ===================== Comps =====================
cp=wb.create_sheet("Comps"); cp.sheet_view.showGridLines=False
def cput(cell,v,f,fmt=None,al=RGT,fl=None):
    c=cp[cell]; c.value=v; c.font=f; c.border=cell_b; c.alignment=al
    if fmt:c.number_format=fmt
    if fl:c.fill=fill(fl)
cp.merge_cells("A1:G1"); cp["A1"]="晶圓代工/半導體 同業比較(概略,資料約 2026/6,估計值)"; cp["A1"].font=Fh(12); cp["A1"].fill=fill(HDR)
cp["A1"].alignment=Alignment(horizontal="left",vertical="center")
for i,h in enumerate(["公司","代碼","市值(US$bn)","營收YoY","毛利率","Fwd P/E","EV/EBITDA"]):
    cput(f"{get_column_letter(1+i)}3",h,Fh(9),al=CEN if i else LEF,fl=HDR)
data=[("台積電 TSMC","TSM","~985","+35%","66%","~20x","~12x"),
 ("Samsung(代工+記憶體)","005930.KS","~430","+10%","~38%","~13x","~5x"),
 ("UMC 聯電","2303.TW","~18","+5%","~32%","~13x","~5x"),
 ("GlobalFoundries","GFS","~24","+6%","~25%","~18x","~8x"),
 ("SMIC 中芯","0981.HK","~75","+20%","~22%","~35x","~9x"),
 ("Intel(含晶圓)","INTC","~120","-3%","~32%","~25x","~8x")]
for r,row in enumerate(data,start=4):
    for i,v in enumerate(row):
        cput(f"{get_column_letter(1+i)}{r}",v,Ff(9,row[0].startswith("台積")),al=LEF if i==0 else CEN,
             fl=(OUT if row[0].startswith("台積") else (INP if r%2 else WHITE)))
cput("A11","解讀:TSMC 毛利率 66% 遠高於同業(22–38%)+ 技術領先,Fwd P/E ~20x / EV/EBITDA ~12x 溢價由獲利品質與 AI 領導地位支撐。",Ff(9,True),al=LEF)
cp.merge_cells("A11:G11")
cput("A12","純代工可比者少(Samsung/Intel 為綜合體);倍數為概略估計,發布前以 FactSet/Bloomberg 校正。",Ff(8),al=LEF); cp.merge_cells("A12:G12")
cp.column_dimensions["A"].width=22
for c in "BCDEFG": cp.column_dimensions[c].width=13

if "Sheet" in wb.sheetnames: wb.remove(wb["Sheet"])
wb.save("./out/TSMC_DCF_Valuation.xlsx")
print("SAVED ./out/TSMC_DCF_Valuation.xlsx  sheets:", wb.sheetnames)
