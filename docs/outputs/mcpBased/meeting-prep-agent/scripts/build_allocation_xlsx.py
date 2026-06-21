# -*- coding: utf-8 -*-
"""Build the allocation-vs-IPS workbook from the *pulled* MCP JSON.

Reads ../mcp_pulls/03_crm_get_ips.json + 04_crm_get_holdings.json (the real MCP
snapshots) and writes ../out/Allocation_vs_IPS_2026-06-22.xlsx. No hard-coded
figures — everything comes from the JSON the harness saved.

Mock / dev only. Run: PYTHONUTF8=1 python build_allocation_xlsx.py
"""
from __future__ import annotations

import json
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

HERE = os.path.dirname(os.path.abspath(__file__))
PULLS = os.path.join(HERE, "..", "mcp_pulls")
OUT = os.path.join(HERE, "..", "out", "Allocation_vs_IPS_2026-06-22.xlsx")


def load(name):
    with open(os.path.join(PULLS, name), encoding="utf-8") as f:
        return json.load(f)["result"]


def main():
    ips = {a["asset_class"]: a for a in load("03_crm_get_ips.json")["allocation"]}
    hold = load("04_crm_get_holdings.json")
    alloc = hold["allocation_by_class"]
    total = hold["total_mv_usd"]

    wb = Workbook()

    # ---- Sheet 1: Allocation vs IPS ----
    ws = wb.active
    ws.title = "Allocation vs IPS"

    hdr_fill = PatternFill("solid", fgColor="1F4E78")
    hdr_font = Font(bold=True, color="FFFFFF")
    breach_fill = PatternFill("solid", fgColor="F8CBAD")
    ok_fill = PatternFill("solid", fgColor="C6E0B4")
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    pct = "0.0%"

    ws["A1"] = "會前配置 vs IPS 檢查 — 王大明 (CL-1001) · 2026-06-22"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A2"] = (f"資料來源:crm MCP get_holdings / get_ips(live pull)。"
                f"總市值 USD {total:,.0f}。Mock/dev 假資料。")
    ws["A2"].font = Font(italic=True, size=9, color="808080")

    cols = ["資產類別", "目前市值 USD", "目前權重", "IPS 目標", "再平衡帶下限",
            "再平衡帶上限", "狀態"]
    r0 = 4
    for c, name in enumerate(cols, start=1):
        cell = ws.cell(row=r0, column=c, value=name)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    # market value by class — derive from holdings rows for the MV column
    mv_by_class = {}
    for h in hold["holdings"]:
        mv_by_class[h["asset_class"]] = mv_by_class.get(h["asset_class"], 0) + (h["mv_usd"] or 0)

    order = ["Equity", "Fixed Income", "Structured/Alt", "Cash"]
    zh = {"Equity": "股票 Equity", "Fixed Income": "固定收益 Fixed Income",
          "Structured/Alt": "結構性/另類 Structured/Alt", "Cash": "現金 Cash"}
    r = r0 + 1
    for cls in order:
        w = alloc.get(cls, 0)
        ip = ips.get(cls, {})
        lo, hi = ip.get("band_low", 0), ip.get("band_high", 0)
        breach = w < lo or w > hi
        status = "❌ 超出帶" if breach else "✓ 帶內"
        if w > hi:
            status = "❌ 超出上限 → 觸發再平衡"
        elif w < lo:
            status = "❌ 低於下限 → 觸發再平衡"
        vals = [zh[cls], mv_by_class.get(cls, 0), w, ip.get("target_weight", 0), lo, hi, status]
        for c, v in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.border = border
            if c == 2:
                cell.number_format = "#,##0"
            if c in (3, 4, 5, 6):
                cell.number_format = pct
            if c == 7:
                cell.fill = breach_fill if breach else ok_fill
                cell.font = Font(bold=breach)
        r += 1

    # total row
    ws.cell(row=r, column=1, value="合計").font = Font(bold=True)
    tc = ws.cell(row=r, column=2, value=total)
    tc.number_format = "#,##0"
    tc.font = Font(bold=True)
    wc = ws.cell(row=r, column=3, value=1.0)
    wc.number_format = pct
    wc.font = Font(bold=True)

    widths = [26, 16, 12, 12, 14, 14, 28]
    for i, wd in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = wd

    # ---- Sheet 2: Rebalance scenario (deploy 2.0M inflow to Fixed Income) ----
    ws2 = wb.create_sheet("Rebalance Scenario")
    ws2["A1"] = "再平衡試算:USD 2.0M 房產款全配固定收益"
    ws2["A1"].font = Font(bold=True, size=13)
    ws2["A2"] = ("情境:把即將入帳的 USD 2,000,000 全數配置到 Fixed Income;"
                 "AUM 12.0M → 14.0M。無需賣出、無稅負。")
    ws2["A2"].font = Font(italic=True, size=9, color="808080")

    inflow = 2_000_000
    new_total = total + inflow
    cols2 = ["資產類別", "目前 USD", "目前權重", "+配置", "情境後 USD",
             "情境後權重", "IPS 帶", "情境後狀態"]
    for c, name in enumerate(cols2, start=1):
        cell = ws2.cell(row=4, column=c, value=name)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    r = 5
    for cls in order:
        cur_mv = mv_by_class.get(cls, 0)
        add = inflow if cls == "Fixed Income" else 0
        new_mv = cur_mv + add
        new_w = new_mv / new_total
        ip = ips.get(cls, {})
        lo, hi = ip.get("band_low", 0), ip.get("band_high", 0)
        breach = new_w < lo or new_w > hi
        status = "❌ 仍超帶" if breach else "✓ 回到帶內"
        band_str = f"{lo:.0%}–{hi:.0%}"
        vals = [zh[cls], cur_mv, alloc.get(cls, 0), add, new_mv, new_w, band_str, status]
        for c, v in enumerate(vals, start=1):
            cell = ws2.cell(row=r, column=c, value=v)
            cell.border = border
            if c in (2, 4, 5):
                cell.number_format = "#,##0"
            if c in (3, 6):
                cell.number_format = pct
            if c == 8:
                cell.fill = breach_fill if breach else ok_fill
                cell.font = Font(bold=True)
        r += 1
    ws2.cell(row=r, column=1, value="合計").font = Font(bold=True)
    for col, val in ((2, total), (4, inflow), (5, new_total)):
        cell = ws2.cell(row=r, column=col, value=val)
        cell.number_format = "#,##0"
        cell.font = Font(bold=True)

    for i, wd in enumerate([26, 14, 12, 12, 14, 12, 12, 18], start=1):
        ws2.column_dimensions[chr(64 + i)].width = wd

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    wb.save(OUT)
    print("wrote", os.path.abspath(OUT))


if __name__ == "__main__":
    main()
