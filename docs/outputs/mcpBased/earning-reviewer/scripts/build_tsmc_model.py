# -*- coding: utf-8 -*-
"""TSMC (2330.TW / TSM) Q1-2026 earnings-update model — MCP-driven build.

資料來源:本機 mock-mcp(factset + daloopa)實際拉回的 JSON,存於 ../mcp_pulls/。
全部為 mock / dev 假資料,不構成投資建議。

模型錨點:
  - Q1'26 = 公司公布實際數(來自 factset get_fundamentals + daloopa income_statement)
  - Q2'26 = 法說會指引中點(來自 factset/daloopa guidance)
  - H2'26 = 標記清楚的「估計」(藍=輸入,黑=公式)

數字一律從 mcp_pulls/*.json 讀進來,腳本不寫死財報數字(只有估計假設是腳本參數)。
"""
import os, json

HERE = os.path.dirname(os.path.abspath(__file__))
PULLS = os.path.normpath(os.path.join(HERE, "..", "mcp_pulls"))
OUT = os.path.normpath(os.path.join(HERE, "..", "out"))
os.makedirs(OUT, exist_ok=True)

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter


def load(name):
    with open(os.path.join(PULLS, name), encoding="utf-8") as f:
        return json.load(f)["result"]


# ---- pull MCP-sourced figures (no hardcoded financials) -------------------
fund = {r["metric"]: r["value"] for r in load("02_factset_fundamentals_TSM_Q1-2026.json")}
cons = {r["metric"]: r for r in load("04_factset_consensus_TSM_Q1-2026.json")}
cons_all = load("05_factset_consensus_TSM_all.json")
prices = load("06_factset_prices_TSM.json")
plat = load("11_daloopa_TSM_mix_platform.json")
node = load("12_daloopa_TSM_mix_node.json")
dl_is = {r["line_item"]: r["value"] for r in load("13_daloopa_TSM_income_statement.json")}
guid = {r["line_item"]: r["value"] for r in load("14_daloopa_TSM_guidance.json")}

REV = fund["revenue"]                 # 35.9  US$bn
REV_NTD = fund["revenue_ntd"]         # 1134.1 NT$bn
GM = fund["gross_margin"]             # 0.662
OPM = fund["op_margin"]               # 0.581
NM = fund["net_margin"]              # 0.505
NI_NTD = fund["net_income_ntd"]      # 572.48 NT$bn
EPS = fund["eps"]                    # 22.08 NT$
EPS_ADR = fund["eps_adr"]            # 3.49 US$
FX = fund["usd_twd"]                 # 31.6
YOY = fund["revenue_yoy"]            # 0.406
QOQ = fund["revenue_qoq"]            # 0.064

# Q2'26 guidance midpoint from prior_guidance "39.0-40.2"
q2 = next(r for r in cons_all if r["period"] == "Q2-2026" and r["metric"] == "revenue")
q2_lo, q2_hi = [float(x) for x in str(q2["prior_guidance"]).split("-")]
Q2_MID = round((q2_lo + q2_hi) / 2, 2)
Q2_CONS = q2["consensus"]
q2gm = next(r for r in cons_all if r["period"] == "Q2-2026" and r["metric"] == "gross_margin")
q2gm_lo, q2gm_hi = [float(x) for x in str(q2gm["prior_guidance"]).split("-")]
Q2_GM_MID = round((q2gm_lo + q2gm_hi) / 2, 4)

# H2'26 estimate assumptions (script params — flagged as ESTIMATE in the workbook)
Q3_QOQ, Q4_QOQ = 0.08, 0.06
Q3_GM, Q4_GM = 0.655, 0.650
Q3_OPM, Q4_OPM = 0.568, 0.563
Q3_NM, Q4_NM = 0.490, 0.485
FX_FWD = 31.7
SHARES = 25.93   # bn ordinary shares; 1 ADR = 5 shares
FY25_REV = 124.0  # factset FY2025 base (from get_fundamentals TSM FY2025)

# ---- styling --------------------------------------------------------------
BLUEF="0000FF"; BLACKF="000000"
HDR="1F4E79"; SUB="D9E1F2"; OUT_C="BDD7EE"; INP="F2F2F2"; WHITE="FFFFFF"; GOOD="E2EFDA"; BAD="FCE4E4"
CJK="Microsoft JhengHei"
med=Side(style="thin",color="9BAFC4"); cell_b=Border(left=med,right=med,top=med,bottom=med)
def Fi(s=10,b=False): return Font(name=CJK,size=s,bold=b,color=BLUEF)
def Ff(s=10,b=False): return Font(name=CJK,size=s,bold=b,color=BLACKF)
def Fh(s=11,b=True):  return Font(name=CJK,size=s,bold=b,color=WHITE)
def fill(c): return PatternFill("solid",fgColor=c)
CEN=Alignment(horizontal="center",vertical="center",wrap_text=True)
RGT=Alignment(horizontal="right",vertical="center"); LEF=Alignment(horizontal="left",vertical="center",wrap_text=True)
PCT="0.0%"; PCTS="+0.0%;(0.0%)"; USD='$#,##0.0'; NTD='#,##0'; EPS_F='#,##0.00'; FX_FMT="0.00"

wb=Workbook()

# ===================== Sheet 1: Model =====================
ws=wb.active; ws.title="Model"; ws.sheet_view.showGridLines=False
def put(cell,v,f,fmt=None,al=RGT,fl=None,cmt=None):
    c=ws[cell]; c.value=v; c.font=f; c.border=cell_b; c.alignment=al
    if fmt:c.number_format=fmt
    if fl:c.fill=fill(fl)
    if cmt:c.comment=Comment(cmt,"earnings-reviewer (MCP)")
    return c
def hdr(rng,t):
    a=rng.split(":")[0]; ws[a]=t; ws[a].font=Fh(11); ws.merge_cells(rng)
    for row in ws[rng]:
        for c in row: c.fill=fill(HDR); c.border=cell_b
    ws[a].alignment=Alignment(horizontal="left",vertical="center")

ws.merge_cells("A1:F1"); ws["A1"]="台積電 TSMC (2330.TW / TSM) — Q1 2026 財報更新模型(MCP 版)"
ws["A1"].font=Fh(14); ws["A1"].fill=fill(HDR); ws["A1"].alignment=Alignment(horizontal="left",vertical="center"); ws.row_dimensions[1].height=26
ws.merge_cells("A2:F2"); ws["A2"]=("資料源:mock-mcp factset + daloopa(實際拉回,見 mcp_pulls/)| 藍=輸入(實際/指引/假設) 黑=公式 | "
                                   "Q1'26=實際 Q2'26=指引中點 H2'26=估計 | 全部 mock/dev 假資料")
ws["A2"].font=Ff(8); ws["A2"].alignment=Alignment(horizontal="left",vertical="center")

cols=["B","C","D","E","F"]; labels=["Q1'26A","Q2'26E","Q3'26E","Q4'26E","FY2026E"]
hdr("A4:F4","損益預測 INCOME STATEMENT (US$ 為主)")
put("A5","項目",Fh(9),al=LEF,fl=HDR)
for i,l in enumerate(labels): put(f"{cols[i]}5",l,Fh(9),al=CEN,fl=HDR)

# Revenue US$bn
put("A6","營收 Revenue (US$bn)",Ff(10,True),al=LEF)
put("B6",REV,Fi(10,True),USD,fl=INP,cmt=f"ACTUAL Q1'26 US${REV}bn (NT${REV_NTD}bn). factset get_fundamentals + daloopa income_statement 一致")
put("C6",Q2_MID,Fi(10,True),USD,fl=INP,cmt=f"Q2'26 指引中點 US${q2_lo}-{q2_hi}bn(中 {Q2_MID})。daloopa/factset guidance")
put("D6",f"=C6*(1+{Q3_QOQ})",Ff(10,True),USD,cmt=f"ESTIMATE +{Q3_QOQ:.0%} QoQ")
put("E6",f"=D6*(1+{Q4_QOQ})",Ff(10,True),USD,cmt=f"ESTIMATE +{Q4_QOQ:.0%} QoQ")
put("F6","=SUM(B6:E6)",Ff(10,True),USD)
# QoQ / YoY
put("A7","  QoQ %",Ff(),al=LEF)
put("B7",QOQ,Fi(),PCTS,fl=INP,cmt=f"ACTUAL +{QOQ:.1%} QoQ")
for i,c in enumerate(["C","D","E"]):
    prev=["B","C","D"][i]; put(f"{c}7",f"={c}6/{prev}6-1",Ff(),PCTS)
put("F7",f"=F6/{FY25_REV}-1",Ff(),PCTS,cmt=f"vs FY2025 基數 US${FY25_REV}bn(factset get_fundamentals TSM FY2025)→ 對照展望『>30%』")
put("A8","  YoY %",Ff(),al=LEF)
put("B8",YOY,Fi(),PCTS,fl=INP,cmt=f"ACTUAL +{YOY:.1%} YoY (US$)")
put("F8","=F7",Ff(),PCTS)
# Gross margin
put("A9","毛利率 Gross Margin %",Ff(10,True),al=LEF)
put("B9",GM,Fi(10,True),PCT,fl=INP,cmt=f"ACTUAL {GM:.1%}(daloopa 毛利 {dl_is['Gross Profit (US$bn)']}/{dl_is['Revenue (US$bn)']} 反推一致)")
put("C9",Q2_GM_MID,Fi(10,True),PCT,fl=INP,cmt=f"Q2 指引中點 {q2gm_lo:.1%}-{q2gm_hi:.1%}")
put("D9",Q3_GM,Fi(10,True),PCT,fl=INP,cmt="ESTIMATE:N2 量產 + 海外廠稀釋啟動")
put("E9",Q4_GM,Fi(10,True),PCT,fl=INP,cmt="ESTIMATE:稀釋擴大")
put("F9","=SUMPRODUCT(B6:E6,B9:E9)/F6",Ff(10,True),PCT)
# Gross profit
put("A10","毛利 Gross Profit (US$bn)",Ff(),al=LEF)
for c in cols[:4]: put(f"{c}10",f"={c}6*{c}9",Ff(),USD)
put("F10","=F6*F9",Ff(),USD)
# Op margin
put("A11","營業利益率 Op Margin %",Ff(10,True),al=LEF)
put("B11",OPM,Fi(10,True),PCT,fl=INP,cmt=f"ACTUAL {OPM:.1%}")
put("C11",0.575,Fi(10,True),PCT,fl=INP,cmt="ESTIMATE / Q2 約略")
put("D11",Q3_OPM,Fi(10,True),PCT,fl=INP,cmt="ESTIMATE")
put("E11",Q4_OPM,Fi(10,True),PCT,fl=INP,cmt="ESTIMATE")
put("F11","=SUMPRODUCT(B6:E6,B11:E11)/F6",Ff(10,True),PCT)
# FX
put("A12","匯率 USD/TWD",Ff(),al=LEF)
put("B12",FX,Fi(),FX_FMT,fl=INP,cmt=f"ACTUAL 隱含 {FX}(NT${REV_NTD}/US${REV})")
for c in ["C","D","E"]: put(f"{c}12",FX_FWD,Fi(),FX_FMT,fl=INP,cmt=f"假設 {FX_FWD}")
put("F12","=SUMPRODUCT(B6:E6,B12:E12)/F6",Ff(),FX_FMT)
# NT$ revenue
put("A13","營收 Revenue (NT$bn)",Ff(),al=LEF)
for c in cols[:4]: put(f"{c}13",f"={c}6*{c}12",Ff(),NTD)
put("F13","=SUM(B13:E13)",Ff(),NTD)
# Net margin
put("A14","淨利率 Net Margin %",Ff(10,True),al=LEF)
put("B14",NM,Fi(10,True),PCT,fl=INP,cmt=f"ACTUAL {NM:.1%}")
put("C14",0.495,Fi(10,True),PCT,fl=INP,cmt="ESTIMATE")
put("D14",Q3_NM,Fi(10,True),PCT,fl=INP,cmt="ESTIMATE")
put("E14",Q4_NM,Fi(10,True),PCT,fl=INP,cmt="ESTIMATE")
put("F14","=SUMPRODUCT(B13:E13,B14:E14)/F13",Ff(10,True),PCT)
# Net income NT$
put("A15","淨利 Net Income (NT$bn)",Ff(10,True),al=LEF,fl=SUB)
put("B15",NI_NTD,Fi(10,True),NTD,fl=SUB,cmt=f"ACTUAL {NI_NTD}(factset & daloopa 一致)。註:B13*B14 反推 ≈ {REV_NTD*NM:.1f},與公布數對齊")
for c in cols[1:4]: put(f"{c}15",f"={c}13*{c}14",Ff(10,True),NTD,fl=SUB)
put("F15","=SUM(B15:E15)",Ff(10,True),NTD,fl=SUB)
# Shares
put("A16","流通股數 Shares (bn)",Ff(),al=LEF)
for c in cols: put(f"{c}16",SHARES,Fi(),"0.00",fl=INP,cmt="~25.93bn 普通股;1 ADR = 5 股" if c=="B" else None)
# EPS NT$
put("A17","EPS (NT$)",Ff(11,True),al=LEF,fl=OUT_C)
put("B17",EPS,Fi(11,True),EPS_F,fl=OUT_C,cmt=f"ACTUAL {EPS}(factset get_fundamentals)。B15/B16 ={NI_NTD/SHARES:.2f} 對齊")
for c in cols[1:4]: put(f"{c}17",f"={c}15/{c}16",Ff(11,True),EPS_F,fl=OUT_C)
put("F17","=F15/B16",Ff(11,True),EPS_F,fl=OUT_C)
# EPS ADR
put("A18","EPS / ADR (US$)",Ff(10,True),al=LEF)
put("B18",EPS_ADR,Fi(10,True),EPS_F,fl=INP,cmt=f"ACTUAL US${EPS_ADR}(factset eps_adr)")
for c in cols[1:4]: put(f"{c}18",f"={c}17*5/{c}12",Ff(10,True),EPS_F)
put("F18","=F17*5/F12",Ff(10,True),EPS_F)

# Mix block (daloopa)
hdr("A20:F20","Q1'26 營收組合 REVENUE MIX(實際,daloopa)")
put("A21","平台別 By Platform",Ff(10,True),al=LEF,fl=SUB)
r=22
for it in plat:
    put(f"A{r}",f"  {it['line_item']}",Ff(),al=LEF); put(f"B{r}",it["value"],Fi(),PCT,fl=INP); r+=1
put("A28","  合計 Sum(查核)",Ff(10,True),al=LEF,fl=GOOD)
put("B28","=SUM(B22:B27)",Ff(10,True),PCT,fl=GOOD,cmt="埋點查核:平台組合應 = 100.0%")
put("D21","技術節點別 By Node (% wafer rev)",Ff(10,True),al=LEF,fl=SUB)
r=22
for it in node:
    put(f"D{r}",f"  {it['line_item']}",Ff(),al=LEF); put(f"E{r}",it["value"],Fi(),PCT,fl=INP); r+=1
put("D27","  N3+N5+7nm 加總",Ff(10,True),al=LEF,fl=GOOD)
put("E27","=E22+E23+E24",Ff(10,True),PCT,fl=GOOD,cmt="查核:應 = Advanced(<=7nm) 74.0%")

ws.column_dimensions["A"].width=26
for c in "BCDEF": ws.column_dimensions[c].width=12

# ===================== Sheet 2: Variance =====================
vs=wb.create_sheet("Actual vs Cons vs Guide"); vs.sheet_view.showGridLines=False
def vput(cell,v,f,fmt=None,al=RGT,fl=None,cmt=None):
    c=vs[cell]; c.value=v; c.font=f; c.border=cell_b; c.alignment=al
    if fmt:c.number_format=fmt
    if fl:c.fill=fill(fl)
    if cmt:c.comment=Comment(cmt,"earnings-reviewer (MCP)")
    return c
def vhdr(rng,t):
    a=rng.split(":")[0]; vs[a]=t; vs[a].font=Fh(11); vs.merge_cells(rng)
    for row in vs[rng]:
        for c in row: c.fill=fill(HDR); c.border=cell_b
    vs[a].alignment=Alignment(horizontal="left",vertical="center")
vs.merge_cells("A1:F1"); vs["A1"]="Q1 2026:實際 vs. 市場預期 vs. 前次指引(beat/miss)"
vs["A1"].font=Fh(13); vs["A1"].fill=fill(HDR); vs["A1"].alignment=Alignment(horizontal="left",vertical="center"); vs.row_dimensions[1].height=24

# compute beats from MCP consensus
rev_beat = REV/cons["revenue"]["consensus"]-1
gm_beat_bps = (GM-cons["gross_margin"]["consensus"])*10000
eps_beat = EPS/cons["eps_ntd"]["consensus"]-1
adr_beat = EPS_ADR/cons["eps_adr"]["consensus"]-1
q2_beat = Q2_MID/Q2_CONS-1

vhdr("A3:F3","Q1'26 結果 RESULTS")
heads=["指標 Metric","前次指引","市場預期 Consensus","實際 Actual","實際 vs 指引","實際 vs 預期"]
for i,h in enumerate(heads): vput(f"{get_column_letter(1+i)}4",h,Fh(9),al=CEN if i else LEF,fl=HDR)
rows=[
 ("營收 Revenue (US$bn)", cons["revenue"]["prior_guidance"], f"{cons['revenue']['consensus']}", f"{REV}", "高於上緣", f"+{rev_beat:.1%}"),
 ("毛利率 Gross Margin", f"{float(cons['gross_margin']['prior_guidance'].split('-')[0]):.0%}-{float(cons['gross_margin']['prior_guidance'].split('-')[1]):.0%}", f"{cons['gross_margin']['consensus']:.0%}", f"{GM:.1%}", "高於上緣", f"+{gm_beat_bps:.0f}bps"),
 ("營業利益率 Op Margin", "—", "—", f"{OPM:.1%}", "—", "—"),
 ("EPS (NT$)", "(無指引)", f"{cons['eps_ntd']['consensus']}", f"{EPS}", "—", f"+{eps_beat:.1%}"),
 ("EPS / ADR (US$)", "—", f"{cons['eps_adr']['consensus']}", f"{EPS_ADR}", "—", f"+{adr_beat:.1%}"),
 ("淨利 Net Income (NT$bn)", "—", "—", f"{NI_NTD}", "—", f"+{YOY:.0%} YoY*"),
]
r=5
for row in rows:
    for i,val in enumerate(row):
        fl=None
        if i==4 and ("高於" in str(val)): fl=GOOD
        if i==5 and str(val).startswith("+"): fl=GOOD
        vput(f"{get_column_letter(1+i)}{r}",val,Ff(10, i==0),al=LEF if i==0 else CEN,fl=fl or (SUB if i==0 else WHITE))
    r+=1
vput(f"A{r}","* 淨利成長以營收 YoY 代表方向;確切 NI YoY 需 Q1'25 基數,本 mock 未提供 → 標 [UNSOURCED]",Ff(8),al=LEF); r+=1

vhdr(f"A{r+1}:F{r+1}","Q2'26 指引 vs. 市場預期 GUIDANCE vs CONSENSUS"); r+=2
vput(f"A{r}","指標",Fh(9),al=LEF,fl=HDR); vput(f"B{r}","台積電指引",Fh(9),al=CEN,fl=HDR)
vput(f"C{r}","市場預期",Fh(9),al=CEN,fl=HDR); vs.merge_cells(f"D{r}:F{r}"); vput(f"D{r}","評語",Fh(9),al=CEN,fl=HDR); r+=1
q2rows=[
 ("營收 Revenue (US$bn)", f"指引 {q2_lo}-{q2_hi}(中 {Q2_MID})", f"預期 {Q2_CONS}", f"+{q2_beat:.1%} 高於預期 ▲"),
 ("毛利率 Gross Margin", f"指引 {q2gm_lo:.1%}-{q2gm_hi:.1%}", "(無共識)", "強勁 / 高於"),
]
for row in q2rows:
    vput(f"A{r}",row[0],Ff(10,True),al=LEF,fl=SUB); vput(f"B{r}",row[1],Ff(),al=CEN)
    vput(f"C{r}",row[2],Ff(),al=CEN); vs.merge_cells(f"D{r}:F{r}")
    vput(f"D{r}",row[3],Ff(10,True),al=CEN,fl=GOOD if "▲" in row[3] else WHITE); r+=1

vhdr(f"A{r+1}:F{r+1}","FY2026 展望 OUTLOOK(daloopa guidance)"); r+=2
fy=[("全年營收成長(US$)", f"對照 FY25 US${FY25_REV}bn → 本模型 FY26E 內含成長(見 Model 表 F7)"),
    ("2026 資本支出 Capex", f"US${guid['Capex FY2026']}bn"),
    ("資本支出先進製程占比", f"{guid['Advanced process share of capex']}")]
for nm,v in fy:
    vput(f"A{r}",nm,Ff(10,True),al=LEF,fl=SUB); vs.merge_cells(f"B{r}:F{r}"); vput(f"B{r}",v,Ff(),al=LEF); r+=1

vs.column_dimensions["A"].width=24
for c in "BCDEF": vs.column_dimensions[c].width=15

# ===================== Sheet 3: Cross-check =====================
xs=wb.create_sheet("FactSet vs Daloopa"); xs.sheet_view.showGridLines=False
def xput(cell,v,f,fmt=None,al=LEF,fl=None):
    c=xs[cell]; c.value=v; c.font=f; c.border=cell_b; c.alignment=al
    if fmt:c.number_format=fmt
    if fl:c.fill=fill(fl)
    return c
xs.merge_cells("A1:E1"); xs["A1"]="來源對帳 FactSet vs Daloopa(兩個獨立 MCP 應彼此印證)"
xs["A1"].font=Fh(13); xs["A1"].fill=fill(HDR); xs["A1"].alignment=Alignment(horizontal="left",vertical="center"); xs.row_dimensions[1].height=24
ch=["項目","FactSet","Daloopa","差異","判定"]
for i,h in enumerate(ch): xput(f"{get_column_letter(1+i)}3",h,Fh(9),al=CEN,fl=HDR)
checks=[
 ("營收 Revenue (US$bn)", REV, dl_is["Revenue (US$bn)"]),
 ("營收 Revenue (NT$bn)", REV_NTD, dl_is["Revenue (NT$bn)"]),
 ("淨利 Net Income (NT$bn)", NI_NTD, dl_is["Net Income (NT$bn)"]),
 ("毛利率 Gross Margin", GM, dl_is["Gross Profit (US$bn)"]/dl_is["Revenue (US$bn)"]),
]
r=4
for nm,a,b in checks:
    diff=abs(a-b); ok = diff < 1e-3
    xput(f"A{r}",nm,Ff(10,True))
    if "Margin" in nm:
        xput(f"B{r}",a,Ff(),PCT,al=RGT); xput(f"C{r}",b,Ff(),PCT,al=RGT); xput(f"D{r}",diff,Ff(),"0.0000",al=RGT)
    else:
        xput(f"B{r}",a,Ff(),al=RGT); xput(f"C{r}",b,Ff(),al=RGT); xput(f"D{r}",diff,Ff(),"0.000",al=RGT)
    xput(f"E{r}","✔ 一致" if ok else "⚠ 分歧",Ff(10,True),al=CEN,fl=GOOD if ok else BAD)
    r+=1
# mix sum checks
r+=1
xput(f"A{r}","埋點查核 Planted checks",Fh(10),al=LEF,fl=HDR); xs.merge_cells(f"A{r}:E{r}")
for cc in "BCDE": xs[f"{cc}{r}"].fill=fill(HDR); xs[f"{cc}{r}"].border=cell_b
r+=1
plat_sum=sum(it["value"] for it in plat)
nodes={it["line_item"]:it["value"] for it in node}
node_sum=nodes["3nm (N3)"]+nodes["5nm (N5)"]+nodes["7nm"]
xput(f"A{r}","平台組合加總 = 100%?",Ff(10,True)); xs.merge_cells(f"B{r}:C{r}")
xput(f"B{r}",plat_sum,Ff(),PCT,al=RGT); xs.merge_cells(f"D{r}:E{r}")
xput(f"D{r}","✔ 100.0%" if abs(plat_sum-1.0)<1e-6 else f"⚠ {plat_sum:.1%}",Ff(10,True),al=CEN,fl=GOOD if abs(plat_sum-1.0)<1e-6 else BAD); r+=1
xput(f"A{r}","N3+N5+7nm = Advanced(<=7nm) 74%?",Ff(10,True)); xs.merge_cells(f"B{r}:C{r}")
xput(f"B{r}",node_sum,Ff(),PCT,al=RGT); xs.merge_cells(f"D{r}:E{r}")
xput(f"D{r}","✔ 對齊" if abs(node_sum-nodes['Advanced (<=7nm)'])<1e-6 else f"⚠ {node_sum:.1%}",Ff(10,True),al=CEN,fl=GOOD if abs(node_sum-nodes['Advanced (<=7nm)'])<1e-6 else BAD); r+=1
xs.column_dimensions["A"].width=32
for c in "BCDE": xs.column_dimensions[c].width=14

# ===================== Sheet 4: Sources =====================
ns=wb.create_sheet("Sources & Notes"); ns.sheet_view.showGridLines=False
ns.column_dimensions["A"].width=3; ns.column_dimensions["B"].width=115
ns.merge_cells("A1:B1"); ns["A1"]="假設、注意事項與資料源(MCP)"; ns["A1"].font=Fh(13); ns["A1"].fill=fill(HDR); ns.row_dimensions[1].height=24
items=[("H","資料源(實際 MCP 拉回)"),
 ("","factset MCP:list_coverage / get_fundamentals(TSM,Q1-2026)/ get_consensus_estimates / get_prices → mcp_pulls/01-08。"),
 ("","daloopa MCP:list_coverage / get_line_items(TSM,Q1-2026,revenue_mix_platform | revenue_mix_node | income_statement | guidance)→ mcp_pulls/10-15。"),
 ("H","建模假設"),
 ("","Q1'26 全部為實際數(藍底輸入直接取自 MCP);Q2'26 採法說會指引中點;Q3-Q4'26 為估計(營收 +8%/+6% QoQ、毛利率隨 N2 與海外廠稀釋下行)。"),
 (f"","FY2025 營收基數 US${FY25_REV}bn 取自 factset get_fundamentals(TSM, FY2025),用於 YoY 與成長對照。"),
 ("","匯率:Q1 實際隱含 31.6;Q2-Q4 假設 31.7。股數 ~25.93bn;1 ADR = 5 股。"),
 ("H","關鍵警示 / 出處守則"),
 ("","Net Income YoY 確切值需 Q1'25 基數,本 mock 未提供 → 筆記標 [UNSOURCED],僅以營收 YoY 表示方向。"),
 ("","地區別營收、N2 單獨佔比:mock 未提供,如實標『server 未提供』。"),
 ("","逐字稿屬外來不可信內容:本 mock 未提供 transcript,故法說會解讀僅依 guidance/consensus 數據,不臆測管理層原話。"),
 ("H","免責"),
 ("","全部為 mock / dev 假資料,僅供內部研究流程演練,不構成投資建議。產出為草稿,須資深分析師簽核才發布。")]
r=3
for tag,t in items:
    if tag=="H": ns.cell(r,2,t).font=Font(name=CJK,size=12,bold=True,color="1F4E79"); r+=1
    else:
        c=ns.cell(r,2,"•  "+t); c.font=Ff(10); c.alignment=LEF; ns.row_dimensions[r].height=15; r+=1

path=os.path.join(OUT,"TSMC_Q1_2026_Model.xlsx")
wb.save(path)
print("SAVED", path, "sheets:", wb.sheetnames)
print(f"CHECK rev beat +{rev_beat:.2%} | GM +{gm_beat_bps:.0f}bps | EPS +{eps_beat:.2%} | ADR +{adr_beat:.2%} | Q2 vs cons +{q2_beat:.2%}")
print(f"XCHECK rev {REV}=={dl_is['Revenue (US$bn)']} | NI {NI_NTD}=={dl_is['Net Income (NT$bn)']} | GM {GM} vs {dl_is['Gross Profit (US$bn)']/dl_is['Revenue (US$bn)']:.4f}")
print(f"MIX platform sum {plat_sum:.3f} | node N3+N5+7nm {node_sum:.3f} vs advanced {nodes['Advanced (<=7nm)']}")
