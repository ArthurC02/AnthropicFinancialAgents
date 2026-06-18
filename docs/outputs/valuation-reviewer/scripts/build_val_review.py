# -*- coding: utf-8 -*-
"""Aurora Growth Fund III — Q1 2026 valuation review workpaper (live formulas)."""
import os
os.makedirs("./out", exist_ok=True)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter
import datetime

BLUEF="0000FF"; BLACKF="000000"; GREENF="008000"
HDR="1F4E79"; SUB="D9E1F2"; OUT="BDD7EE"; INP="F2F2F2"; WHITE="FFFFFF"; WARN="FCE4D6"; OK="E2EFDA"
CJK="Microsoft JhengHei"
med=Side(style="thin",color="9BAFC4"); cell_b=Border(left=med,right=med,top=med,bottom=med)
def Fi(s=10,b=False): return Font(name=CJK,size=s,bold=b,color=BLUEF)
def Ff(s=10,b=False): return Font(name=CJK,size=s,bold=b,color=BLACKF)
def Fl(s=10,b=False): return Font(name=CJK,size=s,bold=b,color=GREENF)
def Fh(s=11,b=True):  return Font(name=CJK,size=s,bold=b,color=WHITE)
def fill(c): return PatternFill("solid",fgColor=c)
CEN=Alignment(horizontal="center",vertical="center",wrap_text=True)
RGT=Alignment(horizontal="right",vertical="center"); LEF=Alignment(horizontal="left",vertical="center",wrap_text=True)
USD='#,##0;(#,##0);"-"'; MULT='0.000"x"'; PCT='0.00%'; DATEF='yyyy-mm-dd'

wb=Workbook()
def sheet(name,idx=None):
    s=wb.create_sheet(name) if idx is None else wb.create_sheet(name,idx); s.sheet_view.showGridLines=False; return s
def W(ws,cell,v,f,fmt=None,al=RGT,fl=None,cmt=None):
    c=ws[cell]; c.value=v; c.font=f; c.border=cell_b; c.alignment=al
    if fmt:c.number_format=fmt
    if fl:c.fill=fill(fl)
    if cmt:c.comment=Comment(cmt,"valuation-reviewer")
def hdr(ws,rng,t):
    a=rng.split(":")[0]; ws[a]=t; ws[a].font=Fh(11); ws.merge_cells(rng)
    for row in ws[rng]:
        for c in row: c.fill=fill(HDR); c.border=cell_b
    ws[a].alignment=Alignment(horizontal="left",vertical="center")

# ===================== 1. Valuation Review =====================
vr=sheet("Valuation Review")
vr.merge_cells("A1:I1"); vr["A1"]="GP 估值複核 Valuation Review（對照政策 P1–P4）· Aurora Growth Fund III · 2026-Q1"
vr["A1"].font=Fh(13); vr["A1"].fill=fill(HDR); vr["A1"].alignment=Alignment(horizontal="left",vertical="center"); vr.row_dimensions[1].height=24
heads=["被投資公司","成本","上季FV","GP報告FV","方法","政策檢核(P1–P4)","判定","政策調整","調整後FV"]
for i,h in enumerate(heads): W(vr,f"{get_column_letter(1+i)}3",h,Fh(9),al=CEN if i>0 else LEF,fl=HDR)
data=[
 ("Helios Software",60_000_000,90_000_000,120_000_000,"可比公司","P2✓；P3待委員會核准文件","有條件放行",0,"P3：確認是否高於2025-09輪價"),
 ("Orion Logistics",50_000_000,55_000_000,55_000_000,"近期交易","P1本季已複核✓","放行",0,"輪價漸老化(>9月)"),
 ("Nova Biotech",40_000_000,40_000_000,65_000_000,"管理層估計","❌P2無方法/無觸發；❌P3無核准","駁回上調",-25_000_000,"撤回+25M至prior/cost 40M"),
 ("Titan Manufacturing",70_000_000,60_000_000,60_000_000,"DCF","P1✓ P2✓(下修,附DCF)","放行",0,"已減損至成本以下"),
 ("Zephyr Retail",30_000_000,30_000_000,30_000_000,"成本法","❌P1：4季未複核(stale)","需重新複核",0,"流程flag；值未驗證"),
]
r0=4
for i,(nm,cost,prior,rep,method,chk,verd,adj,note) in enumerate(data):
    r=r0+i; warn=("❌" in chk)
    W(vr,f"A{r}",nm,Ff(9,True),al=LEF)
    W(vr,f"B{r}",cost,Fi(9),USD); W(vr,f"C{r}",prior,Fi(9),USD); W(vr,f"D{r}",rep,Fi(9),USD)
    W(vr,f"E{r}",method,Ff(9),al=CEN); W(vr,f"F{r}",chk,Ff(8,warn),al=LEF,fl=(WARN if warn else None))
    W(vr,f"G{r}",verd,Ff(9,warn),al=CEN,fl=(WARN if warn else OK))
    W(vr,f"H{r}",adj if adj else None,Fi(9,True) if adj else Ff(9),USD,fl=(WARN if adj else None))
    W(vr,f"I{r}",f"=D{r}+N(H{r})",Ff(9,True),USD)
    vr[f"J{r}"]=note; vr[f"J{r}"].font=Ff(8); vr[f"J{r}"].alignment=LEF
rN=r0+len(data)
W(vr,f"A{rN}","合計 Total",Ff(10,True),al=LEF,fl=SUB)
for col in ["B","C","D","H","I"]:
    W(vr,f"{col}{rN}",f"=SUM({col}{r0}:{col}{rN-1})",Ff(10,True),USD,fl=SUB)
vr.column_dimensions["A"].width=20; vr.column_dimensions["E"].width=10; vr.column_dimensions["F"].width=26
vr.column_dimensions["G"].width=11; vr.column_dimensions["J"].width=28
for c in "BCDHI": vr.column_dimensions[c].width=13

# ===================== 2. NAV Bridge =====================
nb=sheet("NAV Bridge",0)
nb.merge_cells("A1:C1"); nb["A1"]="NAV 橋接 NAV Bridge · 2026-Q1"; nb["A1"].font=Fh(12); nb["A1"].fill=fill(HDR)
nb["A1"].alignment=Alignment(horizontal="left",vertical="center")
rows=[("投資組合FV（GP報告）","='Valuation Review'!D9","link"),
 ("現金 Cash",20_000_000,"input"),("其他資產 Other Assets",5_000_000,"input"),
 ("應付費用 Accrued Liabilities",-5_000_000,"input"),
 ("＝ NAV（as-reported）",None,"=C3+C4+C5+C6"),
 ("減：Nova 不符政策上調（P2/P3）","='Valuation Review'!H6","link"),
 ("Zephyr stale（流程flag,未調金額）",0,"input"),
 ("＝ NAV（policy-adjusted）",None,"=C7+C8+C9")]
r=3
for lab,val,how in rows:
    bold=("＝" in lab); W(nb,f"B{r}",lab,Ff(10,bold),al=LEF,fl=(SUB if bold else None))
    if how=="input": W(nb,f"C{r}",val,Fi(10),USD)
    elif how=="link": W(nb,f"C{r}",val,Fl(10),USD)
    else: W(nb,f"C{r}",how,Ff(10,True),USD,fl=(OUT if "policy-adjusted" in lab else (SUB if bold else None)))
    r+=1
W(nb,"B13","檢查：調整後投組 305M",Ff(9),al=LEF); W(nb,"C13","='Valuation Review'!I9",Fl(9),USD)
nb.column_dimensions["A"].width=2; nb.column_dimensions["B"].width=32; nb.column_dimensions["C"].width=16

# ===================== 3. Returns =====================
rt=sheet("Returns")
rt.merge_cells("A1:D1"); rt["A1"]="報酬指標 Return Metrics"; rt["A1"].font=Fh(12); rt["A1"].fill=fill(HDR)
rt["A1"].alignment=Alignment(horizontal="left",vertical="center")
W(rt,"A3","項目",Fh(9),al=LEF,fl=HDR); W(rt,"C3","Reported",Fh(9),al=CEN,fl=HDR); W(rt,"D3","Adjusted",Fh(9),al=CEN,fl=HDR)
W(rt,"A4","累計出資 Contributed",Ff(10),al=LEF); W(rt,"C4",300_000_000,Fi(10),USD); W(rt,"D4","=C4",Ff(10),USD)
W(rt,"A5","累計分配 Distributions",Ff(10),al=LEF); W(rt,"C5",40_000_000,Fi(10),USD); W(rt,"D5","=C5",Ff(10),USD)
W(rt,"A6","NAV",Ff(10),al=LEF); W(rt,"C6","='NAV Bridge'!C7",Fl(10),USD); W(rt,"D6","='NAV Bridge'!C11",Fl(10),USD)
W(rt,"A7","TVPI",Ff(10,True),al=LEF,fl=SUB); W(rt,"C7","=(C5+C6)/C4",Ff(10,True),MULT,fl=SUB); W(rt,"D7","=(D5+D6)/D4",Ff(10,True),MULT,fl=SUB)
W(rt,"A8","DPI",Ff(10),al=LEF); W(rt,"C8","=C5/C4",Ff(10),MULT); W(rt,"D8","=D5/D4",Ff(10),MULT)
W(rt,"A9","RVPI",Ff(10),al=LEF); W(rt,"C9","=C6/C4",Ff(10),MULT); W(rt,"D9","=D6/D4",Ff(10),MULT)
# cashflow table for XIRR
hdr(rt,"A11:D11","現金流 (XIRR)")
W(rt,"A12","日期",Fh(9),al=CEN,fl=HDR); W(rt,"B12","事件",Fh(9),al=LEF,fl=HDR)
W(rt,"C12","Reported",Fh(9),al=CEN,fl=HDR); W(rt,"D12","Adjusted",Fh(9),al=CEN,fl=HDR)
cf=[(datetime.date(2022,6,30),"出資",-100_000_000,-100_000_000),
    (datetime.date(2023,6,30),"出資",-100_000_000,-100_000_000),
    (datetime.date(2024,6,30),"出資",-100_000_000,-100_000_000),
    (datetime.date(2025,6,30),"分配",40_000_000,40_000_000),
    (datetime.date(2026,3,31),"期末NAV",None,None)]
for i,(d,ev,cr,ca) in enumerate(cf):
    r=13+i; W(rt,f"A{r}",d,Fi(9),DATEF,al=CEN); W(rt,f"B{r}",ev,Ff(9),al=LEF)
    if cr is None:
        W(rt,f"C{r}","=C6",Fl(9),USD); W(rt,f"D{r}","=D6",Fl(9),USD)
    else:
        W(rt,f"C{r}",cr,Fi(9),USD); W(rt,f"D{r}",ca,Fi(9),USD)
W(rt,"A19","Net IRR (XIRR)",Ff(10,True),al=LEF,fl=OUT)
W(rt,"C19","=XIRR(C13:C17,A13:A17)",Ff(10,True),PCT,fl=OUT)
W(rt,"D19","=XIRR(D13:D17,A13:A17)",Ff(10,True),PCT,fl=OUT)
W(rt,"A20","vs 8% hurdle",Ff(9),al=LEF)
W(rt,"D20",'=IF(D19<0.08,"❌ 低於hurdle","✓")',Ff(9,True),al=CEN,fl=WARN)
rt.column_dimensions["A"].width=16; rt.column_dimensions["B"].width=12
rt.column_dimensions["C"].width=15; rt.column_dimensions["D"].width=15

# ===================== 4. Preferred Recompute =====================
pf=sheet("Pref Recompute")
pf.merge_cells("A1:E1"); pf["A1"]="優先報酬獨立重算 Preferred Return Recompute（8% 複利）"; pf["A1"].font=Fh(12); pf["A1"].fill=fill(HDR)
pf["A1"].alignment=Alignment(horizontal="left",vertical="center")
W(pf,"A3","利率 Rate",Ff(10),al=LEF); W(pf,"B3",0.08,Fi(10),PCT)
W(pf,"A4","估值日 As-of",Ff(10),al=LEF); W(pf,"B4",datetime.date(2026,3,31),Fi(10),DATEF,al=CEN)
W(pf,"A6","現金流",Fh(9),al=LEF,fl=HDR); W(pf,"B6","日期",Fh(9),al=CEN,fl=HDR); W(pf,"C6","金額",Fh(9),al=CEN,fl=HDR)
W(pf,"D6","年數t",Fh(9),al=CEN,fl=HDR); W(pf,"E6","FV@8%",Fh(9),al=CEN,fl=HDR)
flows=[("出資1",datetime.date(2022,6,30),100_000_000),("出資2",datetime.date(2023,6,30),100_000_000),
       ("出資3",datetime.date(2024,6,30),100_000_000),("分配",datetime.date(2025,6,30),-40_000_000)]
for i,(nm,d,amt) in enumerate(flows):
    r=7+i; W(pf,f"A{r}",nm,Ff(9),al=LEF); W(pf,f"B{r}",d,Fi(9),DATEF,al=CEN); W(pf,f"C{r}",amt,Fi(9),USD)
    W(pf,f"D{r}",f"=YEARFRAC(B{r},$B$4)",Ff(9),"0.00",al=CEN)
    W(pf,f"E{r}",f"=C{r}*(1+$B$3)^D{r}",Ff(9),USD)
W(pf,"A12","FV 出資合計",Ff(10),al=LEF); W(pf,"E12","=SUM(E7:E9)",Ff(10),USD)
W(pf,"A13","FV 分配",Ff(10),al=LEF); W(pf,"E13","=E10",Ff(10),USD)
W(pf,"A14","淨出資（300−40）",Ff(10),al=LEF); W(pf,"E14",260_000_000,Fi(10),USD)
W(pf,"A15","＝ 獨立重算優先報酬",Ff(10,True),al=LEF,fl=OUT); W(pf,"E15","=E12+E13-E14",Ff(10,True),USD,fl=OUT)
W(pf,"A16","GP 包提供之 pref",Ff(10),al=LEF); W(pf,"E16",35_000_000,Fi(10),USD,cmt="fund_terms Section C 提供值")
W(pf,"A17","差異（重算 − 提供）",Ff(10,True),al=LEF); W(pf,"E17","=E15-E16",Ff(10,True),USD,fl=WARN)
pf.column_dimensions["A"].width=20
for c in "BCDE": pf.column_dimensions[c].width=14

# ===================== 5. Waterfall (3 scenarios) =====================
wf=sheet("Waterfall")
wf.merge_cells("A1:D1"); wf["A1"]="收益分配 Waterfall（歐式/whole-fund；8% pref、100% catch-up、20% carry）"
wf["A1"].font=Fh(12); wf["A1"].fill=fill(HDR); wf["A1"].alignment=Alignment(horizontal="left",vertical="center")
W(wf,"A3","層級 / 情境",Fh(9),al=LEF,fl=HDR)
W(wf,"B3","A: GP包(報告)",Fh(9),al=CEN,fl=HDR); W(wf,"C3","B: 調整+pref35",Fh(9),al=CEN,fl=HDR); W(wf,"D3","C: 調整+獨立pref",Fh(9),al=CEN,fl=HDR)
# inputs row
W(wf,"A4","總價值（分配+NAV）",Ff(10,True),al=LEF)
W(wf,"B4","=Returns!C5+Returns!C6",Fl(10),USD); W(wf,"C4","=Returns!D5+Returns!D6",Fl(10),USD); W(wf,"D4","=Returns!D5+Returns!D6",Fl(10),USD)
W(wf,"A5","承諾出資 Contributed",Ff(10),al=LEF)
for col in "BCD": W(wf,f"{col}5","=Returns!C4",Fl(10),USD)
W(wf,"A6","優先報酬 Preferred",Ff(10),al=LEF)
W(wf,"B6",35_000_000,Fi(10),USD); W(wf,"C6",35_000_000,Fi(10),USD); W(wf,"D6","='Pref Recompute'!E15",Fl(10),USD)
# tiers
W(wf,"A8","① 返還出資",Ff(10),al=LEF)
for col in "BCD": W(wf,f"{col}8",f"=MIN({col}4,{col}5)",Ff(10),USD)
W(wf,"A9","  餘 after capital",Ff(9),al=LEF)
for col in "BCD": W(wf,f"{col}9",f"={col}4-{col}8",Ff(9),USD)
W(wf,"A10","② LP 優先報酬(付)",Ff(10),al=LEF)
for col in "BCD": W(wf,f"{col}10",f"=MIN({col}9,{col}6)",Ff(10),USD)
W(wf,"A11","  餘 after pref",Ff(9),al=LEF)
for col in "BCD": W(wf,f"{col}11",f"={col}9-{col}10",Ff(9),USD)
W(wf,"A12","③ GP 追補 catch-up",Ff(10),al=LEF)
for col in "BCD":
    # catch-up target only if pref fully paid: 0.25*pref ; capped by remaining
    W(wf,f"{col}12",f'=MIN({col}11,IF({col}10>={col}6,0.25*{col}6,0))',Ff(10),USD)
W(wf,"A13","  餘 after catch-up",Ff(9),al=LEF)
for col in "BCD": W(wf,f"{col}13",f"={col}11-{col}12",Ff(9),USD)
W(wf,"A14","④ 80/20 — LP 80%",Ff(10),al=LEF)
for col in "BCD": W(wf,f"{col}14",f"={col}13*0.8",Ff(10),USD)
W(wf,"A15","④ 80/20 — GP 20%",Ff(10),al=LEF)
for col in "BCD": W(wf,f"{col}15",f"={col}13*0.2",Ff(10),USD)
W(wf,"A17","GP Carry 合計",Ff(11,True),al=LEF,fl=OUT)
for col in "BCD": W(wf,f"{col}17",f"={col}12+{col}15",Ff(11,True),USD,fl=OUT)
W(wf,"A18","LP 合計",Ff(11,True),al=LEF,fl=SUB)
for col in "BCD": W(wf,f"{col}18",f"={col}8+{col}10+{col}14",Ff(11,True),USD,fl=SUB)
W(wf,"A19","檢查 LP+GP=總價值",Ff(9),al=LEF)
for col in "BCD": W(wf,f"{col}19",f'=IF(ROUND({col}17+{col}18-{col}4,0)=0,"✓","✗")',Ff(9),al=CEN,fl=OK)
W(wf,"A21","註：情境C用獨立重算pref(~69M);若調整後NAV未達 capital+pref → 無 catch-up、carry=0。pref基準須CCO對帳後定案。",Ff(8),al=LEF)
wf.merge_cells("A21:D22")
wf.column_dimensions["A"].width=24
for c in "BCD": wf.column_dimensions[c].width=16

if "Sheet" in wb.sheetnames: wb.remove(wb["Sheet"])
wb.save("./out/Fund_III_Valuation_Review_2026-Q1.xlsx")
print("SAVED ./out/Fund_III_Valuation_Review_2026-Q1.xlsx  sheets:", wb.sheetnames)
