# -*- coding: utf-8 -*-
"""Vertiv (VRT) M&A proposal — valuation workbook: Comps, Precedents, DCF summary,
illustrative LBO (live formulas), Football Field summary."""
import os
os.makedirs("./out", exist_ok=True)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter

BLUEF="0000FF"; BLACKF="000000"; PURP="800080"; GREENF="008000"
HDR="1F4E79"; SUB="D9E1F2"; OUT="BDD7EE"; INP="F2F2F2"; WHITE="FFFFFF"
CJK="Microsoft JhengHei"
med=Side(style="thin",color="9BAFC4"); cell_b=Border(left=med,right=med,top=med,bottom=med)
def Fi(s=10,b=False): return Font(name=CJK,size=s,bold=b,color=BLUEF)
def Ff(s=10,b=False): return Font(name=CJK,size=s,bold=b,color=BLACKF)
def Fp(s=10,b=False): return Font(name=CJK,size=s,bold=b,color=PURP)
def Fh(s=11,b=True):  return Font(name=CJK,size=s,bold=b,color=WHITE)
def fill(c): return PatternFill("solid",fgColor=c)
CEN=Alignment(horizontal="center",vertical="center",wrap_text=True)
RGT=Alignment(horizontal="right",vertical="center"); LEF=Alignment(horizontal="left",vertical="center",wrap_text=True)
PCT="0.0%"; PCTS="+0.0%;(0.0%)"; USD='$#,##0.0'; SH2="$#,##0.00"; SH0='$#,##0'; NUM="0.000"; MULT='0.0"x"'; MO='0.00"x"'

wb=Workbook()
def sheet(name,idx=None):
    s=wb.create_sheet(name) if idx is None else wb.create_sheet(name,idx)
    s.sheet_view.showGridLines=False; return s
def W(ws,cell,v,f,fmt=None,al=RGT,fl=None,cmt=None):
    c=ws[cell]; c.value=v; c.font=f; c.border=cell_b; c.alignment=al
    if fmt:c.number_format=fmt
    if fl:c.fill=fill(fl)
    if cmt:c.comment=Comment(cmt,"pitch")
def hdr(ws,rng,t):
    a=rng.split(":")[0]; ws[a]=t; ws[a].font=Fh(11); ws.merge_cells(rng)
    for row in ws[rng]:
        for c in row: c.fill=fill(HDR); c.border=cell_b
    ws[a].alignment=Alignment(horizontal="left",vertical="center")

# ============================ LBO (centerpiece) ============================
lb=sheet("LBO")
lb.merge_cells("A1:H1"); lb["A1"]="Vertiv (VRT) — 示意槓桿收購 ILLUSTRATIVE LBO"; lb["A1"].font=Fh(13); lb["A1"].fill=fill(HDR)
lb["A1"].alignment=Alignment(horizontal="left",vertical="center"); lb.row_dimensions[1].height=24
lb.merge_cells("A2:H2"); lb["A2"]="US$bn(每股除外)| 藍=輸入 黑=公式 紫=同頁連結 | 示意性:VRT 規模 ~$122bn,傳統 LBO 不可行,本表僅示範報酬數學"
lb["A2"].font=Ff(8); lb["A2"].alignment=Alignment(horizontal="left",vertical="center")

hdr(lb,"A4:C4","假設 ASSUMPTIONS")
W(lb,"A5","報價/股 Offer ($)",Ff(),al=LEF); W(lb,"B5",400.0,Fi(),SH2,fl=INP,cmt="示意:現價 $320 +25% 控制溢價")
W(lb,"A6","對現價溢價",Ff(),al=LEF); W(lb,"B6","=B5/320-1",Ff(),PCTS)
W(lb,"A7","股數 (bn)",Ff(),al=LEF); W(lb,"B7",0.382,Fi(),NUM,fl=INP)
W(lb,"A8","現有淨負債",Ff(),al=LEF); W(lb,"B8",1.9,Fi(),USD,fl=INP)
W(lb,"A9","FY26E EBITDA",Ff(),al=LEF); W(lb,"B9",3.11,Fi(),USD,fl=INP,cmt="13.75 rev x 22.6% EBITDA margin")
W(lb,"A10","進場槓桿 (x EBITDA)",Ff(),al=LEF); W(lb,"B10",6.0,Fi(),MULT,fl=INP)
W(lb,"A11","負債利率",Ff(),al=LEF); W(lb,"B11",0.08,Fi(),PCT,fl=INP)
W(lb,"A12","稅率",Ff(),al=LEF); W(lb,"B12",0.21,Fi(),PCT,fl=INP)
W(lb,"A13","交易費用 %股權",Ff(),al=LEF); W(lb,"B13",0.025,Fi(),PCT,fl=INP)
W(lb,"A14","出場 EV/EBITDA",Ff(),al=LEF); W(lb,"B14",22.0,Fi(),MULT,fl=INP)

hdr(lb,"E4:H4","資金來源與用途 SOURCES & USES")
W(lb,"E5","用途 Uses",Ff(10,True),al=LEF,fl=SUB); W(lb,"F5","",Ff(),fl=SUB); W(lb,"G5","來源 Sources",Ff(10,True),al=LEF,fl=SUB); W(lb,"H5","",Ff(),fl=SUB)
W(lb,"E6","購買股權",Ff(),al=LEF); W(lb,"F6","=B5*B7",Ff(),USD)
W(lb,"G6","新負債",Ff(),al=LEF); W(lb,"H6","=B10*B9",Ff(),USD)
W(lb,"E7","再融資淨負債",Ff(),al=LEF); W(lb,"F7","=B8",Fp(),USD)
W(lb,"G7","發起人股權(plug)",Ff(),al=LEF); W(lb,"H7","=F9-H6",Ff(),USD)
W(lb,"E8","交易費用",Ff(),al=LEF); W(lb,"F8","=F6*B13",Ff(),USD)
W(lb,"E9","用途合計",Ff(10,True),al=LEF,fl=SUB); W(lb,"F9","=SUM(F6:F8)",Ff(10,True),USD,fl=SUB)
W(lb,"G9","來源合計",Ff(10,True),al=LEF,fl=SUB); W(lb,"H9","=H6+H7",Ff(10,True),USD,fl=SUB)
W(lb,"E10","進場 EV",Ff(),al=LEF); W(lb,"F10","=F6+B8",Ff(),USD)
W(lb,"E11","進場 EV/EBITDA",Ff(),al=LEF); W(lb,"F11","=F10/B9",Ff(),MULT,cmt="示意報價下進場倍數—極高,凸顯不可行")

# operating model
hdr(lb,"A16:H16","營運模型 OPERATING MODEL (FY27→FY31E)")
cols=["D","E","F","G","H"]; yrs=["FY27E","FY28E","FY29E","FY30E","FY31E"]
W(lb,"A17","項目",Fh(9),al=LEF,fl=HDR); W(lb,"C17","進場FY26",Fh(9),al=CEN,fl=HDR)
for i,y in enumerate(yrs): W(lb,f"{cols[i]}17",y,Fh(9),al=CEN,fl=HDR)
g=[.22,.16,.12,.09,.07]; em=[.235,.245,.255,.26,.27]
W(lb,"A18","營收成長 %",Ff(),al=LEF)
for i in range(5): W(lb,f"{cols[i]}18",g[i],Fi(),PCTS,fl=INP)
W(lb,"A19","營收 Revenue",Ff(10,True),al=LEF); W(lb,"C19",13.75,Fi(),USD,fl=INP)
for i in range(5):
    prev="C" if i==0 else cols[i-1]; W(lb,f"{cols[i]}19",f"={prev}19*(1+{cols[i]}18)",Ff(10,True),USD)
W(lb,"A20","EBITDA 利潤率 %",Ff(),al=LEF)
for i in range(5): W(lb,f"{cols[i]}20",em[i],Fi(),PCT,fl=INP)
W(lb,"A21","EBITDA",Ff(10,True),al=LEF)
for i in range(5): W(lb,f"{cols[i]}21",f"={cols[i]}19*{cols[i]}20",Ff(10,True),USD)
W(lb,"A22","  減：D&A(2.6%)",Ff(),al=LEF)
for i in range(5): W(lb,f"{cols[i]}22",f"={cols[i]}19*0.026",Ff(),USD)
W(lb,"A23","EBIT",Ff(),al=LEF)
for i in range(5): W(lb,f"{cols[i]}23",f"={cols[i]}21-{cols[i]}22",Ff(),USD)
W(lb,"A24","  減：Capex(2.3%)",Ff(),al=LEF)
for i in range(5): W(lb,f"{cols[i]}24",f"={cols[i]}19*0.023",Ff(),USD)
W(lb,"A25","  減：現金稅",Ff(),al=LEF)
for i in range(5): W(lb,f"{cols[i]}25",f"={cols[i]}23*$B$12",Ff(),USD)
W(lb,"A26","  減：ΔNWC(8%)",Ff(),al=LEF)
for i in range(5):
    prev="C" if i==0 else cols[i-1]; W(lb,f"{cols[i]}26",f"=({cols[i]}19-{prev}19)*0.08",Ff(),USD)

# debt schedule
W(lb,"A28","負債期初",Ff(),al=LEF)
W(lb,"D28","=H6",Fp(),USD)
for i in range(1,5): W(lb,f"{cols[i]}28",f"={cols[i-1]}31",Ff(),USD)
W(lb,"A29","  利息(期初×利率)",Ff(),al=LEF)
for i in range(5): W(lb,f"{cols[i]}29",f"={cols[i]}28*$B$11",Ff(),USD)
W(lb,"A30","FCF 可償債",Ff(),al=LEF)
for i in range(5): W(lb,f"{cols[i]}30",f"={cols[i]}21-{cols[i]}24-{cols[i]}25-{cols[i]}29-{cols[i]}26",Ff(),USD)
W(lb,"A31","負債期末(掃債)",Ff(10,True),al=LEF,fl=SUB)
for i in range(5): W(lb,f"{cols[i]}31",f"=MAX({cols[i]}28-{cols[i]}30,0)",Ff(10,True),USD,fl=SUB)

# exit & returns
hdr(lb,"A33:H33","出場與報酬 EXIT & RETURNS")
W(lb,"A34","出場 EBITDA (FY31)",Ff(),al=LEF); W(lb,"B34","=H21",Fp(),USD)
W(lb,"A35","出場 EV (×倍數)",Ff(),al=LEF); W(lb,"B35","=B34*B14",Ff(),USD)
W(lb,"A36","減：出場淨負債",Ff(),al=LEF); W(lb,"B36","=H31",Fp(),USD)
W(lb,"A37","出場股權價值",Ff(10,True),al=LEF,fl=SUB); W(lb,"B37","=B35-B36",Ff(10,True),USD,fl=SUB)
W(lb,"A38","發起人進場股權",Ff(),al=LEF); W(lb,"B38","=H7",Fp(),USD)
W(lb,"A39","MOIC",Ff(11,True),al=LEF,fl=OUT); W(lb,"B39","=B37/B38",Ff(11,True),MO,fl=OUT)
W(lb,"A40","IRR (5年)",Ff(11,True),al=LEF,fl=OUT); W(lb,"B40","=(B37/B38)^(1/5)-1",Ff(11,True),PCT,fl=OUT)

# ability to pay
W(lb,"D34","Ability-to-Pay(反推進場)",Ff(10,True),al=LEF,fl=SUB); lb.merge_cells("D34:H34")
W(lb,"D35","目標 IRR",Ff(9,True),al=CEN,fl=SUB); W(lb,"E35","進場股權",Ff(9,True),al=CEN,fl=SUB)
W(lb,"F35","進場 EV",Ff(9,True),al=CEN,fl=SUB); W(lb,"G35","每股 $",Ff(9,True),al=CEN,fl=SUB)
for r,irr in enumerate([0.15,0.20,0.25]):
    rr=36+r; W(lb,f"D{rr}",irr,Fi(),PCT,al=CEN,fl=INP)
    W(lb,f"E{rr}",f"=$B$37/(1+D{rr})^5",Ff(),USD,al=CEN)
    W(lb,f"F{rr}",f"=E{rr}+$H$6",Ff(),USD,al=CEN)
    W(lb,f"G{rr}",f"=(F{rr}-$B$8)/$B$7",Ff(10,True),SH2,al=CEN,fl=(OUT if r==1 else WHITE))
W(lb,"D40","解讀:示意 $400 報價 IRR ~0–1%(不可行);要達 20–25% IRR 進場價需 ~$168–$196,即現價 $320 的大幅折讓。",Ff(8),al=LEF)
lb.merge_cells("D40:H41")
lb.column_dimensions["A"].width=20
for c in "BCDEFGH": lb.column_dimensions[c].width=11.5

# ============================ Comps ============================
cp=sheet("Comps")
cp.merge_cells("A1:H1"); cp["A1"]="同業比較 Trading Comps（散熱/電力基礎設施,資料約 2026/6,估計值）"; cp["A1"].font=Fh(12); cp["A1"].fill=fill(HDR)
cp["A1"].alignment=Alignment(horizontal="left",vertical="center")
for i,h in enumerate(["公司","代碼","市值$bn","營收YoY","EBITDA利潤率","Fwd P/E","EV/EBITDA","純度"]):
    W(cp,f"{get_column_letter(1+i)}3",h,Fh(9),al=CEN if i else LEF,fl=HDR)
data=[("Vertiv ★","VRT","~122","+28%","~23%","~46x","~40-51x","純(高)"),
 ("nVent","NVT","~28","+11%","~20%","~35x","~32x","多元"),
 ("Schneider","SU.PA","~160","+5%","~19%","~28x","~21x","多元"),
 ("Delta 台達電","2308","~187","+33%","~18%","~46x","n/a","多元"),
 ("Supermicro","SMCI","~18","+123%","~6%","~9x","~16x","OEM"),
 ("nVent/同業中位","—","—","—","—","~30x","~26x","—")]
for r,row in enumerate(data,start=4):
    for i,v in enumerate(row):
        W(cp,f"{get_column_letter(1+i)}{r}",v,Ff(9,row[0].startswith("Vertiv")),al=LEF if i==0 else CEN,
          fl=(OUT if row[0].startswith("Vertiv") else (INP if r%2 else WHITE)))
W(cp,"A11","Comps 隱含 VRT 區間(套 25–40x FY26E EBITDA $3.1bn):每股約 $198–$320。VRT 本身即同業最高倍數,溢價反映 AI 純度與龍頭地位。",Ff(9,True),al=LEF)
cp.merge_cells("A11:H11")
cp.column_dimensions["A"].width=18
for c in "BCDEFGH": cp.column_dimensions[c].width=12

# ============================ Precedent Transactions ============================
pt=sheet("Precedent Txns")
pt.merge_cells("A1:G1"); pt["A1"]="先例交易 Precedent Transactions（資料中心散熱/基礎設施 2022–2026）"; pt["A1"].font=Fh(12); pt["A1"].fill=fill(HDR)
pt["A1"].alignment=Alignment(horizontal="left",vertical="center")
for i,h in enumerate(["日期","買方","標的","EV (US$)","EV/EBITDA","EV/Revenue","來源旗標"]):
    W(pt,f"{get_column_letter(1+i)}3",h,Fh(9),al=CEN if i else LEF,fl=HDR)
deals=[("2026-03","Ecolab","CoolIT Systems","$4.75bn","29x NTM / 24x'27E","~8.6x","揭露/估"),
 ("2025-12","Eaton","Boyd Thermal","$9.5bn","22.5x '26E","~5.6x","揭露/估"),
 ("2024-10","Schneider","Motivair(75%)","$0.85bn","n/a","中個位數(rev)","揭露(質化)"),
 ("2025-11","Vertiv","PurgeRite","~$1.0bn","~10x '26E(含綜效)","n/a","揭露"),
 ("2023-04","Carrier","Viessmann(參考)","~$13bn","~13x(綜效)","~3.0x","揭露(非DC)"),
 ("2025-03","nVent","Avail EPG(電力)","$0.975bn","n/a","~2.6x","估"),
 ("2023","KKR","CoolIT(前手)","未揭露","n/a","n/a","KKR 出場 ~15x")]
for r,row in enumerate(deals,start=4):
    for i,v in enumerate(row):
        W(pt,f"{get_column_letter(1+i)}{r}",v,Ff(9,i==2),al=LEF if i in(1,2) else CEN, fl=(INP if r%2 else WHITE))
W(pt,"A12","純散熱控制交易 EV/EBITDA 區間 ~10x–29x(前瞻);套 20–28x 於 VRT FY26E EBITDA → 每股約 $158–$223。",Ff(9,True),al=LEF)
pt.merge_cells("A12:G12")
W(pt,"A13","注意:CoolIT/Boyd 為『前瞻』倍數且標的成長更快、規模更小;VRT 規模大、成長較緩,直接類比會高估。Carrier/Viessmann 為 HVAC,僅作規模參考。",Ff(8),al=LEF)
pt.merge_cells("A13:G13")
pt.column_dimensions["A"].width=11; pt.column_dimensions["B"].width=12; pt.column_dimensions["C"].width=18
for c in "DEFG": pt.column_dimensions[c].width=15

# ============================ DCF summary ============================
ds=sheet("DCF Summary")
ds.merge_cells("A1:E1"); ds["A1"]="DCF 估值彙總（詳見 Vertiv_VRT_DCF.xlsx）"; ds["A1"].font=Fh(12); ds["A1"].fill=fill(HDR)
ds["A1"].alignment=Alignment(horizontal="left",vertical="center")
for i,h in enumerate(["方法","WACC","終值","隱含每股","備註"]):
    W(ds,f"{get_column_letter(1+i)}3",h,Fh(9),al=CEN if i else LEF,fl=HDR)
rows=[("永續成長法","11.4%","g=3.5%","~$102","紀律性;終值占EV偏高"),
 ("出場倍數法","11.4%","22x EBITDA","~$245","較寬鬆"),
 ("區間","—","—","$100–$245","DCF 中樞遠低於現價 $320")]
for r,row in enumerate(rows,start=4):
    for i,v in enumerate(row):
        W(ds,f"{get_column_letter(1+i)}{r}",v,Ff(9,r==6),al=LEF if i in(0,4) else CEN,fl=(OUT if r==6 else (INP if r%2 else WHITE)))
W(ds,"A8","結論:以紀律性 DCF,VRT 內在價值中樞約 $100–$245,顯著低於現價 $320——市場已 price in 高度完美成長。",Ff(9,True),al=LEF)
ds.merge_cells("A8:E8")
ds.column_dimensions["A"].width=16
for c in "BCDE": ds.column_dimensions[c].width=15

# ============================ Football Field (summary) ============================
ff=sheet("Football Field",0)
ff.merge_cells("A1:F1"); ff["A1"]="估值區間彙總 Football Field — Vertiv (VRT)"; ff["A1"].font=Fh(13); ff["A1"].fill=fill(HDR)
ff["A1"].alignment=Alignment(horizontal="left",vertical="center"); ff.row_dimensions[1].height=24
ff.merge_cells("A2:F2"); ff["A2"]="每股 US$;現價 $320;示意控制報價 $400(+25%)"; ff["A2"].font=Ff(9); ff["A2"].alignment=Alignment(horizontal="left",vertical="center")
for i,h in enumerate(["估值法","低端 $","高端 $","中點 $","vs現價","說明"]):
    W(ff,f"{get_column_letter(1+i)}4",h,Fh(9),al=CEN if i else LEF,fl=HDR)
methods=[("52週交易區間",130,400,"~$265","市場情緒(示意)"),
 ("分析師目標價",300,450,"~$375","賣方共識(示意);BofA等"),
 ("同業比較 Comps",198,320,"$259","25–40x FY26E EBITDA"),
 ("先例交易",158,223,"$191","20–28x;純散熱控制交易"),
 ("DCF",100,245,"$173","永續法~$102 / 出場22x~$245"),
 ("LBO ability-to-pay",168,196,"$182","20–25% IRR 反推")]
for r,(nm,lo,hi,mid,note) in enumerate(methods,start=5):
    W(ff,f"A{r}",nm,Ff(9,True),al=LEF,fl=INP)
    W(ff,f"B{r}",lo,Fi(),SH0,al=CEN); W(ff,f"C{r}",hi,Fi(),SH0,al=CEN)
    W(ff,f"D{r}",mid,Ff(),al=CEN); W(ff,f"E{r}",f"=({lo}+{hi})/2/320-1",Ff(),PCTS,al=CEN)
    W(ff,f"F{r}",note,Ff(8),al=LEF)
W(ff,"A12","現價 Current",Ff(9,True),al=LEF,fl=SUB); W(ff,"B12",320,Fi(),SH0,al=CEN,fl=SUB)
W(ff,"A13","示意報價 Offer",Ff(9,True),al=LEF,fl=SUB); W(ff,"B13",400,Fi(),SH0,al=CEN,fl=SUB)
W(ff,"A15","核心結論:現價 $320 已位於多數內在/交易法之上;控制溢價(→~$400)僅情緒類參考(52週高/樂觀PT)可及。",Ff(10,True),al=LEF)
ff.merge_cells("A15:F15")
W(ff,"A16","內在(DCF)、先例交易與 LBO ability-to-pay 一致指向 $158–$245——收購 VRT 需對已偏貴的價格再付溢價,且 LBO 規模/倍數不可行。",Ff(9),al=LEF)
ff.merge_cells("A16:F16")
ff.column_dimensions["A"].width=20; ff.column_dimensions["F"].width=30
for c in "BCDE": ff.column_dimensions[c].width=11

# ============================ Synergies & Sensitivity ============================
sy=sheet("Synergies & Sensitivity")
sy.merge_cells("A1:H1"); sy["A1"]="綜效 Ability-to-Pay 與 IRR 敏感度（連結 LBO 分頁）"; sy["A1"].font=Fh(12); sy["A1"].fill=fill(HDR)
sy["A1"].alignment=Alignment(horizontal="left",vertical="center")
sy.merge_cells("A2:H2"); sy["A2"]="US$bn(每股除外)| 綠=連結 LBO | 假設:出場淨負債保守維持基準(不計綜效額外掃債)"; sy["A2"].font=Ff(8)
sy["A2"].alignment=Alignment(horizontal="left",vertical="center")

# Synergy ability-to-pay @ target IRR
hdr(sy,"A4:F4","綜效 Ability-to-Pay（@ 20% IRR,出場 22x）")
W(sy,"A5","綜效(run-rate, $bn)",Ff(9,True),al=CEN,fl=SUB); W(sy,"B5","出場EBITDA",Ff(9,True),al=CEN,fl=SUB)
W(sy,"C5","出場股權",Ff(9,True),al=CEN,fl=SUB); W(sy,"D5","進場股權@20%",Ff(9,True),al=CEN,fl=SUB)
W(sy,"E5","進場EV",Ff(9,True),al=CEN,fl=SUB); W(sy,"F5","每股 $",Ff(9,True),al=CEN,fl=SUB)
for r,S in enumerate([0,0.5,1.0,1.5,2.0]):
    rr=6+r
    W(sy,f"A{rr}",S,Fi(),USD,al=CEN,fl=INP)
    W(sy,f"B{rr}",f"=LBO!$H$21+A{rr}",Ff(),USD,al=CEN)   # exit EBITDA + synergy
    W(sy,f"C{rr}",f"=B{rr}*LBO!$B$14-LBO!$H$31",Ff(),USD,al=CEN)  # exit equity
    W(sy,f"D{rr}",f"=C{rr}/(1+0.2)^5",Ff(),USD,al=CEN)
    W(sy,f"E{rr}",f"=D{rr}+LBO!$H$6",Ff(),USD,al=CEN)
    W(sy,f"F{rr}",f"=(E{rr}-LBO!$B$8)/LBO!$B$7",Ff(10,True),SH2,al=CEN,fl=(OUT if r==2 else WHITE))
W(sy,"A12","解讀:即使綜效 $2.0bn(≈進場 EBITDA 的 64%),@20% IRR 的進場價仍僅 ~$243,低於現價 $320。",Ff(9,True),al=LEF)
sy.merge_cells("A12:F12")
W(sy,"A13","反推:要在 20% IRR 下 justify $320,需 run-rate 綜效約 $5.3bn(≈進場 EBITDA 1.7 倍)——不切實際。",Ff(9,True),al=LEF)
sy.merge_cells("A13:F13")

# 5x5 IRR sensitivity: offer price x exit multiple (centered 400 / 22x)
hdr(sy,"A15:H15","IRR 敏感度:報價/股 × 出場 EV/EBITDA（中心=基準 $400 / 22x）")
W(sy,"A16","報價↓ / 倍數→",Ff(9,True),al=CEN,fl=SUB)
exits=[18,20,22,24,26]
for j,m in enumerate(exits): W(sy,f"{get_column_letter(2+j)}16",m,Fi(9),MULT,al=CEN,fl=SUB)
offers=[320,360,400,440,480]
for i,of in enumerate(offers):
    rr=17+i; W(sy,f"A{rr}",of,Fi(9),SH0,al=CEN,fl=SUB)
    for j,m in enumerate(exits):
        col=get_column_letter(2+j)
        # sponsor equity = offer*shares*(1+fee)+netdebt-debt ; exit equity = m*exitEBITDA - exitND
        f=(f"=(({m}*LBO!$H$21-LBO!$H$31)/($A{rr}*LBO!$B$7*(1+LBO!$B$13)+LBO!$B$8-LBO!$H$6))^(1/5)-1")
        ctr=(i==2 and j==2)
        W(sy,f"{col}{rr}",f,Ff(10,ctr),PCT,al=CEN,fl=(OUT if ctr else (INP if (i+j)%2 else WHITE)))
W(sy,"A23","中心格($400/22x)應≈LBO 分頁 IRR(~0.7%)。即便出場 26x + 報價降至 $320,IRR 仍遠低於 PE 門檻(~20–25%)。",Ff(8),al=LEF)
sy.merge_cells("A23:H24")
sy.column_dimensions["A"].width=16
for c in "BCDEFGH": sy.column_dimensions[c].width=12

if "Sheet" in wb.sheetnames: wb.remove(wb["Sheet"])
wb.save("./out/VRT_MA_Valuation.xlsx")
print("SAVED ./out/VRT_MA_Valuation.xlsx  sheets:", wb.sheetnames)
